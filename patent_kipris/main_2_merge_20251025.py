# -*- coding: utf-8 -*-
"""
기존 엑셀 파일에서 '패밀리정보(수)' 열의 해당 셀만 opFamilyCnt로 업데이트
- 비교 키: (출원번호(일자)=applno, 등록번호=rgstno)
- 전부 문자열 비교 (공백 제거)
- 다른 셀/행/스타일은 건드리지 않음
- 멀티스레드 10개
"""

import json
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, Tuple, List, Optional
from openpyxl import load_workbook

# ========= 사용자 설정 =========
EXCEL_PATH  = "요청2.xlsx"
JSON_PATH   = "data_list.json"
SHEET_NAME  = None          # None이면 첫 번째 시트 사용, 특정 시트면 문자열로 지정
THREADS     = 10
OVERWRITE_ALL = False        # 기존 값이 있어도 덮어쓸지 여부
PRINT_HIT   = False         # 매칭 성공 로그 출력
PRINT_MISS  = False         # 매칭 실패 로그 출력

# ========= 유틸 =========
def _s(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r", "").replace("\n", "").strip()
    return s

def build_mapping(j: Dict[str, Any]) -> Dict[Tuple[str, str], Dict[str, Any]]:
    m: Dict[Tuple[str, str], Dict[str, Any]] = {}
    for _, val in j.items():
        applno = _s(val.get("applno"))
        rgstno = _s(val.get("rgstno"))
        if applno and rgstno:
            m[(applno, rgstno)] = val
    print(f"[INFO] JSON 매핑 수: {len(m):,}")
    return m

# ========= 메인 로직 =========
def main():
    # 1) JSON 로드 → 매핑
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        j = json.load(f)
    mapping = build_mapping(j)

    # 2) 엑셀 로드
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.worksheets[0]

    # 3) 헤더 행 파악 및 컬럼 인덱스 매핑
    header_row = 1
    headers = {}
    for c in range(1, ws.max_column + 1):
        headers[_s(ws.cell(row=header_row, column=c).value)] = c

    need = ["출원번호(일자)", "등록번호", "패밀리정보(수)"]
    for h in need:
        if h not in headers:
            raise ValueError(f"엑셀에 필수 헤더 누락: {h}")

    col_appl = headers["출원번호(일자)"]
    col_rgst = headers["등록번호"]
    col_fam  = headers["패밀리정보(수)"]

    # 4) 대상 행(2행~max_row)에서 (row_idx, appl, rgst) 수집
    tasks: List[Tuple[int, str, str]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        appl = _s(ws.cell(row=r, column=col_appl).value)
        rgst = _s(ws.cell(row=r, column=col_rgst).value)
        tasks.append((r, appl, rgst))

    # 5) 멀티스레드 매칭 → (row_idx, opFamilyCnt or None)
    def check_task(t: Tuple[int, str, str]) -> Tuple[int, Optional[str]]:
        r, appl, rgst = t
        if not appl or not rgst:
            if PRINT_MISS:
                print(f"[MISS] r={r} (빈 값) {appl} / {rgst}")
            return (r, None)
        hit = mapping.get((appl, rgst))
        if hit is None:
            if PRINT_MISS:
                print(f"[MISS] r={r} {appl} / {rgst}")
            return (r, None)
        val = _s(hit.get("opFamilyCnt", ""))
        if PRINT_HIT:
            print(f"[HIT]  r={r} {appl} / {rgst} -> {val}")
        return (r, val if val != "" else None)

    results: Dict[int, str] = {}
    updated = 0
    with ThreadPoolExecutor(max_workers=THREADS) as ex:
        futures = [ex.submit(check_task, t) for t in tasks]
        for f in as_completed(futures):
            r, val = f.result()
            if val is None:
                continue
            cur = ws.cell(row=r, column=col_fam).value
            cur_s = _s(cur)
            if OVERWRITE_ALL or cur_s == "":
                ws.cell(row=r, column=col_fam, value=val)
                updated += 1

    # 6) 저장(원본만 갱신)
    wb.save(EXCEL_PATH)
    print(f"[DONE] 업데이트 완료: {updated:,} 셀 수정 (파일: {EXCEL_PATH})")

if __name__ == "__main__":
    main()

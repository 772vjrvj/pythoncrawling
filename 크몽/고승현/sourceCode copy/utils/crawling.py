import os
import re
import shutil
import openpyxl
from openpyxl.styles import PatternFill
import requests
from bs4 import BeautifulSoup
from .console import crwalingScreen
import logging
from datetime import datetime
logDir = './logs'
now = datetime.now()
formattedNow = now.strftime("%Y%m%d%H%M%S")
logFile = f"{formattedNow}.txt"

if not os.path.exists(logDir):
    os.makedirs(logDir)

logPath = os.path.join(logDir, logFile)

logging.basicConfig(
    level=logging.ERROR,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(logPath),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

MAIN_URL = "https://www.worldfootball.net/"
MAIN_CONTENT = BeautifulSoup(requests.get(MAIN_URL).content, "html.parser")

def addToList(list, datas, index):
    return list[:index] + datas + list[index:]


# 모든 나라를 가져온다.
# ('England', '/competition/eng-premier-league/') 형식의 list
def getCountryURL() -> list[tuple[str, str]]:
    country_url_list: list[tuple[str, str]] = []
    li_items = MAIN_CONTENT.find("div", class_="top").find("ul").find_all("li")
    # 상위 네이게이션 나라 링크를 가져온다.
    flag = True
    for li_item in li_items:
        if flag:
            flag = False
            continue
        countrys = li_item.find_all("a")
        for country in countrys:
            url = country.get("href");  countryName = country.text
            if(url == "javascript;;" or countryName == "FAQ"): continue
            country_url_list.append((countryName, url))
    return country_url_list


def getWhatURL(countryURL: str):
    main_content = BeautifulSoup(requests.get(countryURL).content, "html.parser")
    whats = []

    for el in main_content.select('.subnavi a'):
        if el.get('href') == "javascript:;":
            continue
        whats.append((el.text, el.get('href')))

    return whats

def getFirstTableData(Round: str, url: str):
    main_content = BeautifulSoup(requests.get(url).content, "html.parser")
    tables = main_content.select('.box .data .standard_tabelle')
    rows = tables[0].find_all("tr")
    hasSecondTable = False

    if (
        len(tables) > 1
        and len(tables[1].find_all("tr")) > 0
    ):
        hasSecondTable = True

    datas = []
    isFirst = True
    for row in rows:
        rowDatas = row.find_all("td")
        data = []
        for rowData in rowDatas:
            data.append(rowData.text.strip().replace("\n", ""))
        if len(data) < 3: continue
        if isFirst:
            isFirst = False
            if data[1].strip() == "":
                data[1] = "??:??"
        else:
            if data[0].strip() == "":
                data[0] = datas[-1][0]
            if data[1].strip() == "":
                data[1] = datas[-1][1]
        datas.append(data)

    datas = [[Round if i == 0 else "" for i in range(len(datas[0]))]] + datas
    return datas, hasSecondTable

def getSecondTableData(url: str):
    main_content = BeautifulSoup(requests.get(url).content, "html.parser")
    rows = main_content.find_all("table", class_="standard_tabelle")[1].find_all("tr")

    datas = []
    for row in rows[1:-1]:
        data = []
        rowDatas = row.find_all("th")
        for rowData in rowDatas:
            data.append(rowData.text.strip().replace("\n", ""))
        if len(data) <  3: data = []
        rowDatas = row.find_all("td")
        for index, rowData in enumerate(rowDatas):
            if index != 0 and rowData.text.strip() == "": continue
            data.append(rowData.text.strip().replace("\n", ""))
        if len(data) < 3: continue
        datas.append(data)

    for i in range(len(datas)):
        if(i == 0): continue
        if datas[i][0].strip() == "": datas[i][0] = datas[i-1][0]

    for data in datas:
        temp = data[0]
        data[0] = data[1]
        data[1] = temp

    return datas

def generateExcelFilePath(path, index):
    return f"{path}_{index}.xlsx"

def writeExcel(datas, path):
    fileIndex = 0
    pathWithIndex = generateExcelFilePath(path, fileIndex)

    while os.path.exists(pathWithIndex):
        fileIndex += 1
        pathWithIndex = generateExcelFilePath(path, fileIndex)

    shutil.copyfile("./template.xlsx", pathWithIndex)

    workbook = openpyxl.load_workbook(pathWithIndex)
    sheet = workbook.active

    # 2행부터 데이터 작성
    for i, data in enumerate(datas, start=2):
        for j, value in enumerate(data, start=1):
            sheet.cell(row=i, column=j).value = value

    greenFill = PatternFill(start_color="00FF00", fill_type="solid")
    redFill = PatternFill(start_color="FF0000", fill_type="solid")

    # au, aw열 초록색 적용, av, ax열 빨간색 적용
    for row in sheet.iter_rows(min_row=2, min_col=47, max_col=50):
        for cell in row:
            if(cell.column % 2 == 0):
                cell.fill = redFill
            else:
                cell.fill = greenFill

    workbook.save(pathWithIndex)

# @return {{type: "Results" | "Results & Tables", urls: [string, string][]}}
def makeYearsURL(Data) -> {
    "type": str,
    "urls": list[tuple[str, str]]
}:
    fromYear, toYear, url = Data
    yearsURL: list[tuple[str, str]] = []

    base_url = "https://www.worldfootball.net"

    main_content = BeautifulSoup(requests.get(url).content, "html.parser")
    try:
        result_url = main_content.find("a", text="Results").get("href")
        resultType = 'Results'
    except:
        result_url = main_content.find("a", text="Results & Tables").get("href")
        resultType = 'Results & Tables'

    enterTable_url = base_url + result_url

    main_content = BeautifulSoup(requests.get(enterTable_url).content, "html.parser")
    years = [str(year) for year in range(fromYear, toYear + 1)]
    for yearData in main_content.find("select", attrs={"name": "saison"}).find_all("option"):
        if any(year in yearData.text for year in years): yearsURL.append((yearData.text, base_url + yearData.get("value")))

    return {
        "type": resultType,
        "urls": yearsURL
    }

def makeRoundsURL(url: str) -> list[tuple[str, str]]:
    roundsURL: list[tuple[str, str]] = []

    base_url = "https://www.worldfootball.net"
    main_content = BeautifulSoup(requests.get(url).content, "html.parser")
    try:
        roundDatas = main_content.find("select", attrs={"name": "phase"}).find_all("option")
    except:
        try:
            roundDatas = main_content.find("select", attrs={"name": "runde"}).find_all("option")
        except:
            return [('', url)]

    for roundData in roundDatas:
        roundsURL.append((roundData.text, base_url + roundData.get("value")))
    return roundsURL


def selectTable(Data):
    Category, What, Year, Round, url = Data

    df = []
    df1, hasSecondTable = getFirstTableData(Round, url)
    if hasSecondTable:
        df2 = getSecondTableData(url)
        df1[0] += df2[0] + df2[0]
        TeamIndex = 0

        for index in range(len(df2[0])):
            if df2[0][index] == "Team":
                TeamIndex = index
                break

        for i in range(len(df1)):
            if(i == 0): continue
            FirstTeamIndex = 0;  SecondTeamIndex = 0
            for data in range(len(df1[i])):
                for team in range(len(df2)):
                    if(team == 0): continue
                    if(df1[i][data].strip() == df2[team][TeamIndex].strip()):
                        if(FirstTeamIndex == 0): FirstTeamIndex = team
                        else: SecondTeamIndex = team

            df1[i] += df2[FirstTeamIndex] + df2[SecondTeamIndex]

    isFirst = True
    for df_row in df1:
        if isFirst:
            isFirst = False
            continue

        if df_row[0] == "":
            new_df_row = ["","","","",""]
        else:
            new_df_row = [
                df_row[0].split("/")[2], df_row[0].split("/")[1], df_row[0].split("/")[0],
                df_row[1].split(":")[0], df_row[1].split(":")[1],
            ]


        if df_row[5] == "annull" or df_row[5] == "dnp" or df_row[5] == "WO":
            for _ in range(18): df_row.append("")
            df_row[17] = df_row[5]

        if len(df_row) < 10:
            for _ in range(18): df_row.append("")

        if hasSecondTable:
            for _ in range(25): new_df_row.append("0")
        else:
            for _ in range(25): new_df_row.append("")


        if hasSecondTable:
            # Pt. 3

            df_row[15] = int(df_row[15])
            if df_row[15] != 0:
                new_df_row[5] = df_row[15] // 100; new_df_row[6] = (df_row[15] % 100) // 10; new_df_row[7] = (df_row[15] % 10)
            # Dif. 4
            df_row[14] = int(df_row[14])
            if(df_row[14] != 0):
                new_df_row[8] = "+" if df_row[14] >= 0 else "-"
                if df_row[14] < 0: df_row[14] *= -1
                if df_row[14] != 0:
                    new_df_row[9] = df_row[14] // 100; new_df_row[10] = (df_row[14] % 100) // 10; new_df_row[11] = (df_row[14] % 10)
            # goals 3
            goal1 = int(df_row[13].split(":")[0])
            goal2 = int(df_row[13].split(":")[0])
            if goal1 != 0:
                new_df_row[12] = goal1 // 100; new_df_row[13] = (goal1 % 100) // 10; new_df_row[14] = (goal1 % 10)
            if goal2 != 0:
                new_df_row[15] = goal2 // 100; new_df_row[16] = (goal2 % 100) // 10; new_df_row[17] = (goal2 % 10)
            # L 3
            df_row[12] = int(df_row[12])
            if df_row[12] != 0:
                new_df_row[18] = df_row[12] // 100; new_df_row[19] = (df_row[12] % 100) // 10; new_df_row[20] = (df_row[12] % 10)
            # D 3
            df_row[11] = int(df_row[11])
            if df_row[11] != 0:
                new_df_row[21] = df_row[11] // 100; new_df_row[22] = (df_row[11] % 100) // 10; new_df_row[23] = (df_row[11] % 10)
            # W 3
            df_row[10] = int(df_row[10])
            if df_row[10] != 0:
                new_df_row[24] = df_row[10] // 100; new_df_row[25] = (df_row[10] % 100) // 10; new_df_row[26] = (df_row[10] % 10)
            # M. 3
            df_row[9] = int(df_row[9])
            if df_row[9] != 0:
                new_df_row[27] = df_row[9] // 100; new_df_row[28] = (df_row[9] % 100) // 10; new_df_row[29] = (df_row[9] % 10)

        # Team1
        if df_row[7]:
            new_df_row.append(df_row[7])
        else:
            new_df_row.append("")

        # Rank1
        if df_row[8]:
            df_row[8] = int(df_row[8])
            new_df_row += [df_row[8] // 100, (df_row[8] % 100) // 10, (df_row[8] % 10)]
        else:
            new_df_row += ["", "", ""]

        # Score
        if df_row[5].strip() == "-:-":
            for _ in range(4): new_df_row.append("-")
        else:
            # 4:1 (0:1 1:1 1:1) 경우 -> 01 11
            if df_row[5].count(':') == 4:
                score = re.split(r"[:,\(\)]", df_row[5].strip())
                new_df_row += [
                    score[2].strip(), score[3].strip(), score[4].strip(), score[5].strip()
                ]
            # 2:3 (0:1 1:3) 경우 -> 01 13
            elif df_row[5].count(':') == 3:
                score = re.split(r"[:,\(\)]", df_row[5].strip())
                new_df_row += [
                    score[2].strip(), score[3].strip(), score[4].strip(), score[5].strip()
                ]
            else:
                # '0:3 (0:2) 경우
                score = re.split(r"[:,\(\)]", df_row[5].strip())
                if df_row[5] == "annull" or df_row[5] == "dnp" or df_row[5] == "WO":
                    new_df_row += [df_row[5], '', '', '']
                elif df_row[5].count(':') == 1:
                    new_df_row += [
                        score[0].strip(), score[1].strip(), '', ''
                    ]
                else:
                    new_df_row += [
                        score[0].strip(), score[1].strip(), score[2].strip(), score[3].strip()
                    ]
        # Rank2
        if df_row[17]:

            if df_row[17] == "annull" or df_row[17] == "dnp" or df_row[17] == "WO":
                new_df_row += [df_row[17], '', '']
            else:
                df_row[17] = int(df_row[17])
                new_df_row += [df_row[17] // 100, (df_row[17] % 100) // 10, (df_row[17] % 10)]

        # Team2
        new_df_row.append(df_row[16])

        secondTeamStartIndex = len(new_df_row)

        if hasSecondTable:
            for _ in range(25): new_df_row.append("0")
        else:
            for _ in range(25): new_df_row.append("")

        if hasSecondTable:
            # Pt. 3
            df_row[-1] = int(df_row[-1])
            if df_row[-1] != 0:
                new_df_row[secondTeamStartIndex + 22] = df_row[-1] // 100
                new_df_row[secondTeamStartIndex + 23] = (df_row[-1] % 100) // 10
                new_df_row[secondTeamStartIndex + 24] = (df_row[-1] % 10)
            # Dif. 4
            df_row[-2] = int(df_row[-2])
            new_df_row[secondTeamStartIndex + 18] = "+" if df_row[-2] >= 0 else "-"
            if df_row[-2] < 0: df_row[-2] *= -1
            if df_row[-2] != 0:
                new_df_row[secondTeamStartIndex + 19] = df_row[-2] // 100
                new_df_row[secondTeamStartIndex + 20] = (df_row[-2] % 100) // 10
                new_df_row[secondTeamStartIndex + 21] = (df_row[-2] % 10)
            # goals 6
            goal1 = int(df_row[-3].split(":")[0])
            goal2 = int(df_row[-3].split(":")[0])
            if goal1 != 0:
                new_df_row[secondTeamStartIndex + 12] = goal1 // 100
                new_df_row[secondTeamStartIndex + 13] = (goal1 % 100) // 10
                new_df_row[secondTeamStartIndex + 14] = (goal1 % 10)
            if goal2 != 0:
                new_df_row[secondTeamStartIndex + 15] = goal2 // 100
                new_df_row[secondTeamStartIndex + 16] = (goal2 % 100) // 10
                new_df_row[secondTeamStartIndex + 17] = (goal2 % 10)
            # L 3
            df_row[-4] = int(df_row[-4])
            if df_row[-4] != 0:
                new_df_row[secondTeamStartIndex + 9] = df_row[-4] // 100
                new_df_row[secondTeamStartIndex + 10] = (df_row[-4] % 100) // 10
                new_df_row[secondTeamStartIndex + 11] = (df_row[-4] % 10)
            # D 3
            df_row[-5] = int(df_row[-5])
            if df_row[-5] != 0:
                new_df_row[secondTeamStartIndex + 6] = df_row[-5] // 100
                new_df_row[secondTeamStartIndex + 7] = (df_row[-5] % 100) // 10
                new_df_row[secondTeamStartIndex + 8] = (df_row[-5] % 10)
            # W 3
            df_row[-6] = int(df_row[-6])
            if df_row[-6] != 0:
                new_df_row[secondTeamStartIndex + 3] = df_row[-6] // 100
                new_df_row[secondTeamStartIndex + 4] = (df_row[-6] % 100) // 10
                new_df_row[secondTeamStartIndex + 5] = (df_row[-6] % 10)
            # M. 3
            df_row[-7] = int(df_row[-7])
            if df_row[-7] != 0:
                new_df_row[secondTeamStartIndex] = df_row[-7] // 100
                new_df_row[secondTeamStartIndex + 1] = (df_row[-7] % 100) // 10
                new_df_row[secondTeamStartIndex + 2] = (df_row[-7] % 10)

        df.append(new_df_row)

    return df

# 라운드가 있는 게임에서 게임별로 그룹화
# @return [[...rounds, dunkel], [...rounds, dunkel], ...]
def splitGame(rows):
    result = []
    current_chunk = []

    for row in rows:
        current_chunk.append(row)

        if('dunkel' in row.get('class', [])):
            result.append(current_chunk)
            current_chunk = []

    if current_chunk:
        result.append(current_chunk)

    return result;

# 요구사항: 점수 표현이 4개 이거나 끝자리에 aet 혹은 pso 라는 영문이 적혀있다면
#         괄호 안에 표시되는 점수들 중 0번째, 1번째만 파싱
# 점수 표현 경우의 수
# 0:1 (0:1)
# 3:5 (0:1, 0:1, 0:1) pso
def parseScore(score):
    isMultipleScore = score.count(':') == 4 or score.count(':') == 3

    au:str = None
    av:str = None
    aw:str = None
    ax:str = None

    if isMultipleScore:
        inBracket = score[score.find('(') + 1:score.find(')')]
        splittedScores = inBracket.split(', ')
        auav = splittedScores[1].split(':')
        au = auav[0]
        av = auav[1]
        awax = splittedScores[0].split(':')
        aw = awax[0]
        ax = awax[1]

    else:
        mainScore = score.split(' ')[0]
        auav = mainScore.split(':')
        au = auav[0]
        av = auav[1]

        braketStartIndex = score.find('(')
        braketEndIndex = score.find(')')

        hasBraket = braketStartIndex != -1 and braketEndIndex != -1
        if(hasBraket):
            braketScore = score[score.find('(') + 1:score.find(')')]
            awax = braketScore.split(',')[0].split(':')
            aw = awax[0]
            ax = awax[1]

    return au, av, aw, ax

def parseGame(game):
    column = game.select('td')
    time = column[1].text.strip().split(':')

    try:
        hour = time[0]
    except IndexError:
        hour = '00'
    try:
        minute = time[1]
    except IndexError:
        minute = '00'

    homeTeamName = column[2].text.strip()
    awayTeamName = column[4].text.strip()

    score = column[5].text.strip()

    if score == 'dnp' or score == 'abor.':
        return None

    au, av, aw, ax = parseScore(score)

    return {
        'hour': hour,
        'minute': minute,
        'homeTeam': homeTeamName,
        'awayTeam': awayTeamName,
        'au': au,
        'av': av,
        'aw': aw,
        'ax': ax
    }

def parseRound(round, dunkel, index):
    column = round.select('td')
    # 1st(1st leg): 20/02/2024 12:30
    dateText = dunkel.select('td[title$="leg"]')[index].text

    # date에 시간이 없는 경우 (1st(1st leg): 20/02/2024)
    if(len(dateText.split(' ')[-1]) != 5):
        datePart = dateText.split(' ')[-1]
        day = datePart.split('/')[0]
        month = datePart.split('/')[1]
        year = datePart.split('/')[2]
        hour = '00'
        minute = '00'
    else:
        datePart = dateText[-16:-6]
        timePart = dateText[-5:]
        day = datePart.split('/')[0]
        month = datePart.split('/')[1]
        year = datePart.split('/')[2]
        hour = timePart.split(':')[0]
        minute = timePart.split(':')[1]

    homeTeamName = column[1].text.strip()
    awayTeamName = column[3].text.strip()

    score = column[-2].text.strip()

    if score == 'dnp':
        return None

    au, av, aw, ax = parseScore(score)

    return {
        'day': day,
        'month': month,
        'year': year,
        'hour': hour,
        'minute': minute,
        'homeTeam': homeTeamName,
        'awayTeam': awayTeamName,
        'au': au,
        'av': av,
        'aw': aw,
        'ax': ax
    }

def scrapResults(Round: str, url: str):
    games = []

    bs = BeautifulSoup(requests.get(url).text, "html.parser")
    table = bs.select_one('.box .data .standard_tabelle')
    # row: [일자(DD/MM/YYYY), 시각(HH:mm), A팀명, -, B팀명, 스코어, 공백]
    rows = table.find_all('tr')
    # dunkel이 있는 경우 라운드가 있는 게임이라고 판단
    hasRound = len([row for row in rows if 'dunkel' in row.get('class', [])]) > 0

    if(hasRound):
        # class에 dunkel과 hell을 포함한 row만 셀렉 (라운드가 있는경우 게임 사이에 divider가 있음)
        filtered_rows = [row for row in rows if 'dunkel' in row.get('class', []) or 'hell' in row.get('class', [])]

        splitted = splitGame(filtered_rows)
        for game in splitted:
            for index, round in enumerate(game[:-1]):
                parsedData = parseRound(round, game[-1], index);
                if(parsedData):
                    games.append(parsedData)
    else:
        date = rows[0].text.strip().split('/')

        for row in rows:
            if row.select('td')[0].text.strip():
                date = row.select('td')[0].text.strip().split('/')

            game = parseGame(row)
            if(game):
                game['day'] = date[0]
                game['month'] = date[1]
                game['year'] = date[2]
                games.append(game)

    return games

def crawling(datas: list[tuple[int, int, int, int]]):
    base_url = "https://www.worldfootball.net"
    countryData = getCountryURL()

    crwaling_datas = []
    for data in datas:
        country = data[0]; what = data[1]; fromYear = data[2]; toYear = data[3]
        whatData = getWhatURL(base_url + countryData[country][1])
        country_datas = []

        generatedUrlInfo = makeYearsURL((fromYear, toYear, base_url + whatData[what][1]))
        for year, url in generatedUrlInfo["urls"]:
            year_datas = []
            for round, roundUrl in makeRoundsURL(url):
                if generatedUrlInfo["type"] == 'Results & Tables':
                    round_datas = []
                    try:
                        round_datas = selectTable((countryData[country][0], whatData[what][0], year, round, roundUrl))
                    except Exception as e:
                        logger.exception(f"{countryData[country][0]}_{whatData[what][0]}_{year}_{round}_{roundUrl}")

                    for round_data in round_datas:
                        year_datas.append([round] + round_data)

                elif generatedUrlInfo["type"] == 'Results':
                    results = []
                    try:
                        results = scrapResults(countryData[country][0], roundUrl)
                    except Exception as e:
                        logger.exception(f"{countryData[country][0]}_{whatData[what][0]}_{year}_{round}_{roundUrl}")

                    for result in results:
                        year_datas.append([
                            round,
                            # 날짜정보
                            result["year"],result["month"],result["day"],result["hour"],result["minute"],
                            # Pt.
                            None,None,None,
                            # Dif.
                            None,None,None,None,
                            # goals
                            None,None,None,None,None,None,
                            # L
                            None,None,None,
                            # D
                            None,None,None,
                            # W
                            None,None,None,
                            # M.
                            None,None,None,
                            result["homeTeam"],
                            # rank
                            None,None,None,
                            # score
                            result["au"],result["av"],result["aw"],result["ax"],
                            # rank
                            None,None,None,
                            result["awayTeam"],
                            None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,None,
                        ])

                crwalingScreen((countryData[country][0], whatData[what][0], year, round))
            for year_data in year_datas:
                country_datas.append([year] + year_data)
        for country_data in country_datas:
            crwaling_datas.append([countryData[country][0], whatData[what][0]] + country_data)

    sorted_datas = sorted(crwaling_datas, key=lambda x: (x[4], x[5], x[6], x[7], x[8]))

    # 38, 50번째 행에 빈 데이터를 추가해 간격 띄우기
    parsed_datas = []
    for row in sorted_datas:
        mEndIndex = 34
        awayTeamEndIndex = mEndIndex + 12
        addEmptyColumnCount = 8
        row = addToList(row, [None]*addEmptyColumnCount, mEndIndex)
        row = addToList(row, [None]*addEmptyColumnCount, awayTeamEndIndex + addEmptyColumnCount)
        parsed_datas.append(row)

    countryCode = data[0]
    countryName = countryData[country][0]

    excelDirPath = os.path.join('./Data')
    if not os.path.exists(excelDirPath):
        os.makedirs(excelDirPath)

    fileName = f"{countryCode}_{countryName}"
    invalid_chars = r'[\/:*?"<>|]'
    sanitized_filename = re.sub(invalid_chars, '-', fileName)

    writeExcel(parsed_datas, os.path.join(excelDirPath, sanitized_filename))
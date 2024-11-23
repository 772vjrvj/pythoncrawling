import time
import random
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from sites.ruliweb import extract_links_ruriweb, extract_contents_ruriweb
from sites.arcalive import extract_links_arcalive, extract_contents_arcalive
from sites.inven import extract_links_inven, extract_contents_inven
from sites.fmkorea import extract_links_fmkorea, extract_contents_fmkorea


# 드라이버 세팅
def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--incognito")
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
        driver.maximize_window()

        return driver
    except Exception as e:
        print(f"Error setting up the WebDriver: {e}")
        return None


# 링크 가져오기
def get_links(driver, site, keyword, start_page=1):
    all_links = set()  # 중복 제거를 위해 set 사용
    page = start_page
    while True:
        print(f"site : {site}, keyword : {keyword}, link_page : {page}")
        if site == "fmkorea":
            links = extract_links_fmkorea(driver, keyword, page)
        elif site == "ruliweb":
            links = extract_links_ruriweb(driver, keyword, page)
        elif site == "inven":
            links = extract_links_inven(driver, keyword, page)
        elif site == "arcalive":
            links = extract_links_arcalive(driver, keyword, page)
        else:
            raise ValueError(f"Unknown site: {site}")

        # 검색 결과가 없는 경우 종료
        if not links:
            print("No more results for keyword:", keyword)
            break
        all_links.update(links)
        page += 1

        # 페이지 요청 간 간격
        time.sleep(random.uniform(2, 3))

    all_links = list(all_links)
    print(f'all_links len : {len(all_links)}')
    return all_links


# 페이지에서 데이터 추출
def extract_contents(driver, site, keyword, link, forbidden_keywords):
    if site == "fmkorea":
        new_link = f"https://www.fmkorea.com/{link}"
        return extract_contents_fmkorea(driver, site, keyword, new_link, forbidden_keywords)
    elif site == "inven":
        return extract_contents_inven(driver, keyword, site, keyword, link, forbidden_keywords)
    elif site == "ruliweb":
        return extract_contents_ruriweb(driver, keyword, site, keyword, link, forbidden_keywords)
    elif site == "arcalive":
        return extract_contents_arcalive(driver, keyword, site, keyword, link, forbidden_keywords)
    else:
        raise ValueError(f"Unknown site: {site}")


# 엑셀 저장
def save_or_append_to_excel(data, site, keyword):
    df = pd.DataFrame(data)
    filename = f'{site}_results.xlsx'
    try:
        # 기존 파일이 있을 경우 데이터를 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(filename)
            sheet_name = workbook.sheetnames[0]  # 첫 번째 시트 이름 가져오기
            # 기존 데이터의 마지막 행 번호 계산
            if writer.sheets.get(sheet_name):
                startrow = writer.sheets[sheet_name].max_row
            else:
                startrow = 0
            # 데이터 추가
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
            print(f"site : {site}, keyword : {keyword}, excel append save")
    except FileNotFoundError:
        # 파일이 없을 경우 새 파일 생성
        print(f"site : {site}, keyword : {keyword}, excel new save")
        df.to_excel(filename, index=False)


# main
if __name__ == "__main__":
    sites = [
        "fmkorea",
        # "ruliweb",
        # "inven",
        # "arcalive"
    ]
    keywords = [
        "마공스시",
        "읍읍스시",
        "마공읍읍",
        "ㅁㄱㅅㅅ",
        "ㅁㄱ스시",
        "신지수",
        # "ㅅㅈㅅ",
        "보일러집 아들",
        "대열보일러",
        "project02",
        "버블트리"
    ]
    forbidden_keywords = ["병신지수", "혁신지수", "여신지수"]
    driver = setup_driver()

    # 중복 제거를 위해 set 사용
    all_result_links = set()
    if not driver:
        print("Driver setup failed!")
        exit()
    try:
        for index, site in enumerate(sites, start=1):
            print(f"site : {site} ({index}/{len(sites)})")
            for idx, keyword in enumerate(keywords, start=1):
                print(f"site : {site} ({index}/{len(sites)}), keyword : {keyword} ({idx}/{len(keywords)})")
                result_links = get_links(driver, site, keyword)
                if not result_links:
                    continue
                # result_links에서 all_result_links와 중복된 것 제거
                unique_links = [link for link in result_links if link not in all_result_links]
                print(f"unique_links len : {len(unique_links)}")

                # all_result_links에 고유 링크 추가
                all_result_links.update(unique_links)
                results = []
                for ix, link in enumerate(unique_links, start=1):  # 중복 제거된 unique_links 사용
                    print(f'site : {site} ({index}/{len(sites)}), keyword : {keyword} ({idx}/{len(keywords)}), links ({ix}/{len(unique_links)})')
                    data = extract_contents(driver, site, keyword, link, forbidden_keywords)
                    if data:
                        results.extend(data)

                # 엑셀 저장 또는 추가
                print(f'results : {len(results)}')
                if results:
                    save_or_append_to_excel(results, site, keyword)

    finally:
        driver.quit()
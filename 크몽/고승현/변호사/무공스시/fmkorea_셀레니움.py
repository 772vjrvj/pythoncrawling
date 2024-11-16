import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from PIL import Image
from openpyxl import load_workbook


# 이미지 저장 폴더 설정
IMAGE_FOLDER = "fmkorea_image_list"
os.makedirs(IMAGE_FOLDER, exist_ok=True)


# 드라이버 세팅
def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--incognito")
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
        driver.maximize_window()

        return driver
    except Exception as e:
        print(f"Error setting up the WebDriver: {e}")
        return None


# 페이지 전체 스크린샷 함수
def capture_full_page_screenshot(driver, file_path):
    try:
        # 페이지 전체 크기 가져오기
        total_width = driver.execute_script("return document.body.scrollWidth")
        total_height = driver.execute_script("return document.body.scrollHeight")
        viewport_height = driver.execute_script("return window.innerHeight")

        # 스크롤 단계와 캡처된 이미지를 저장할 리스트
        scroll_steps = range(0, total_height, viewport_height)
        screenshot_parts = []

        for step in scroll_steps:
            # 스크롤 위치 이동
            driver.execute_script(f"window.scrollTo(0, {step});")
            time.sleep(0.3)  # 스크롤 대기

            # 현재 뷰포트 캡처
            screenshot_part_path = f"{file_path}_part_{step}.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 마지막 스크롤에서 남은 높이 처리
        if total_height % viewport_height > 0:
            driver.execute_script(f"window.scrollTo(0, {total_height - viewport_height});")
            time.sleep(0.3)
            screenshot_part_path = f"{file_path}_part_final.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 이미지 결합
        stitched_image = Image.new("RGB", (total_width, total_height))
        current_height = 0

        for idx, part_path in enumerate(screenshot_parts):
            with Image.open(part_path) as part_image:
                # 현재 캡처된 이미지 크기 가져오기
                part_width, part_height = part_image.size

                # 마지막 스크롤 조정
                if idx == len(screenshot_parts) - 1 and total_height % viewport_height > 0:
                    part_image = part_image.crop((0, part_height - (total_height % viewport_height), part_width, part_height))

                stitched_image.paste(part_image, (0, current_height))
                current_height += part_image.size[1]

            os.remove(part_path)  # 임시 파일 삭제

        # 최종 스크린샷 저장
        final_path = f"{file_path}_full.png"
        stitched_image.save(final_path)
        print(f"Full page screenshot saved: {final_path}")
        return final_path

    except Exception as e:
        print(f"Error capturing full page screenshot: {e}")
        return None


# 페이지에서 데이터 추출
def extract_page_data(driver, url, category, link):
    try:
        driver.get(url)
        time.sleep(3)  # 페이지 로딩 대기

        # 스크린샷 저장
        screenshot_path = os.path.join(IMAGE_FOLDER, f"fmkorea_{url.split('/')[-1]}")
        full_screenshot_path = capture_full_page_screenshot(driver, screenshot_path)

        # 공통 데이터 추출
        base_data = {
            "글 번호": link,
            "제목": "",
            "내용": "",
            "날짜": "",
            "아이디": "",
            "키워드": category,
            "url": url,
            "스크린샷": full_screenshot_path,
        }

        # 제목 추출
        try:
            title = driver.find_element(By.CSS_SELECTOR, "span.np_18px_span").text
            base_data["제목"] = title
        except NoSuchElementException:
            base_data["제목"] = ""

        # 내용 추출
        try:
            content = driver.find_element(By.TAG_NAME, "article").text
            base_data["내용"] = content
        except NoSuchElementException:
            base_data["내용"] = ""

        # 날짜 추출
        try:
            date = driver.find_element(By.CSS_SELECTOR, "span.date.m_no").text
            base_data["날짜"] = date
        except NoSuchElementException:
            base_data["날짜"] = ""

        # 아이디 추출
        try:
            user_id = driver.find_element(By.CSS_SELECTOR, "a[href='#popup_menu_area']").text
            base_data["아이디"] = user_id
        except NoSuchElementException:
            base_data["아이디"] = ""

        # 리플 데이터 추출
        comments_section = driver.find_elements(By.CSS_SELECTOR, "ul.fdb_lst_ul > li")
        comment_list = []

        if not comments_section:  # 리플이 없다면
            # 리플 데이터가 없으면 빈 데이터로 하나의 배열을 리턴
            base_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 날짜": "",
                "리플 내용": ""
            })
            comment_list.append(base_data)
        else:
            # 리플이 있다면 기존 로직대로 처리
            for idx, comment in enumerate(comments_section, start=1):
                try:
                    # 리플 번호
                    reply_data = {"리플 번호": idx}

                    # 리플 아이디
                    try:
                        reply_user_id = comment.find_element(By.CSS_SELECTOR, "div.meta a").text
                        reply_data["리플 아이디"] = reply_user_id
                    except NoSuchElementException:
                        reply_data["리플 아이디"] = ""

                    # 리플 날짜
                    try:
                        reply_date = comment.find_element(By.CSS_SELECTOR, "div.meta .date").text
                        reply_data["리플 날짜"] = reply_date
                    except NoSuchElementException:
                        reply_data["리플 날짜"] = ""

                    # 리플 내용
                    try:
                        # 댓글 내용 전체 추출
                        comment_content = comment.find_element(By.CSS_SELECTOR, "div.comment-content .xe_content")

                        # findParent 안의 내용 추출
                        try:
                            find_parent_text = comment_content.find_element(By.CSS_SELECTOR, "a.findParent").text
                            reply_content = f"@{find_parent_text}<-"

                            # 댓글 본문에서 findParent 텍스트 제거
                            full_comment_text = comment_content.text
                            full_comment_text = full_comment_text.replace(find_parent_text, "", 1).strip()
                        except NoSuchElementException:
                            reply_content = ""
                            full_comment_text = comment_content.text.strip()

                        # 댓글 내용 추가
                        reply_content += full_comment_text
                        reply_data["리플 내용"] = reply_content
                    except NoSuchElementException:
                        reply_data["리플 내용"] = ""

                    obj = {**base_data, **reply_data}
                    print(f"obj : {obj}")
                    # 공통 데이터 병합
                    comment_list.append(obj)
                except Exception as e:
                    print(f"Error processing comment {idx}: {e}")

        return comment_list

    except Exception as e:
        print(f"Error processing {url}: {e}")
        return []


#엑셀
def save_or_append_to_excel(data, filename="fmkorea_results.xlsx"):
    df = pd.DataFrame(data)

    try:
        # 기존 파일이 있을 경우 데이터를 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(filename)
            sheet_name = workbook.sheetnames[0]  # 첫 번째 시트 이름 가져오기
            # 기존 데이터의 마지막 행 번호 계산
            if writer.sheets.get(sheet_name):
                startrow = writer.sheets[sheet_name].max_row
            else:
                startrow = 0
            # 데이터 추가
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
            print(f"Data successfully appended to {filename}")
    except FileNotFoundError:
        # 파일이 없을 경우 새 파일 생성
        df.to_excel(filename, index=False)
        print(f"New file created: {filename}")


# 함수: href 값 추출 (Selenium으로 변경)
def extract_links_selenium(driver, keyword, page):
    try:
        base_url = "https://www.fmkorea.com/search.php"
        params = f"?act=IS&is_keyword={keyword}&mid=home&where=document&page={page}"
        url = base_url + params
        driver.get(url)
        time.sleep(3)  # 페이지 로딩 대기

        # 검색 결과 링크 추출
        search_results = driver.find_elements(By.CSS_SELECTOR, "ul.searchResult > li > dl > dt > a")
        if not search_results:
            return []

        # 링크 추출
        links = [link.get_attribute("href").split("/")[-1] for link in search_results]
        return links

    except Exception as e:
        print(f"Error extracting links: {e}")
        return []


# 메인 함수 (Selenium 기반으로 변경)
def main_selenium(driver, keyword, start_page=1):
    all_links = []
    page = start_page
    while True:
        print(f"Fetching page {page} for keyword: {keyword}...")
        links = extract_links_selenium(driver, keyword, page)
        if not links:  # 검색 결과가 없는 경우 종료
            print("No more results for keyword:", keyword)
            break
        all_links.extend(links)
        page += 1
        time.sleep(1)  # 페이지 요청 간 간격
    print(f"Extraction complete for keyword: {keyword}")
    return all_links


# 실행
if __name__ == "__main__":
    keywords = [
        "마공스시",
        "읍읍스시",
        "마공읍읍",
        "ㅁㄱㅅㅅ",
        "ㅁㄱ스시",
        # "신지수",
        # "ㅅㅈㅅ",
        "보일러집 아들",
        "대열보일러",
        "project02",
        "버블트리"
    ]  # 검색어 배열

    driver = setup_driver()
    if not driver:
        print("Driver setup failed!")
        exit()

    try:
        for keyword in keywords:
            print(f"Processing keyword: {keyword}")
            result_links = main_selenium(driver, keyword)

            results = []
            for link in result_links:
                url = f"https://www.fmkorea.com/{link}"
                print(f"Processing URL: {url}")
                data = extract_page_data(driver, url, keyword, link)
                if data:
                    results.extend(data)

            # 엑셀 저장 또는 추가
            print(f'results len : {len(results)}')
            print(f'results : {results}')
            if results:
                save_or_append_to_excel(results)

    finally:
        driver.quit()

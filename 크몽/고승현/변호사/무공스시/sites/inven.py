import requests
from bs4 import BeautifulSoup
import time
import random
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

# 페이지 데이터 추출 함수
def extract_contents_inven(driver, url, keyword):
    try:
        driver.get(url)
        time.sleep(3)
        # 댓글 데이터 추출
        comments = []

        # 스크린샷 저장
        screenshot_path = os.path.join(IMAGE_FOLDER, f"inven_{url.split('/')[-1]}.png")
        full_screenshot_path = capture_full_page_screenshot(driver, screenshot_path)

        # 공통 데이터 추출
        page_data = {
            "글 번호": url.split("/")[-1],
            "제목": "",
            "내용": "",
            "아이디": "",
            "작성일": "",
            "키워드": keyword,
            "url": url,
            "스크린샷": full_screenshot_path,
        }

        # 제목, 내용, 사용자 정보 추출
        try:
            page_data["제목"] = driver.find_element(By.CLASS_NAME, "articleTitle").text
            page_data["내용"] = driver.find_element(By.ID, "powerbbsContent").text

            articleInfo = driver.find_element(By.CLASS_NAME, "articleInfo")
            page_data["아이디"] = articleInfo.find_element(By.CLASS_NAME, "articleWriter").text
            page_data["작성일"] = articleInfo.find_element(By.CLASS_NAME, "articleDate").text
        except NoSuchElementException as e:
            print(f"Error extracting main data: {e}")

            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments


        try:
            # 두 번째 class="commentList1"을 찾기
            comment_lists = driver.find_elements(By.CLASS_NAME, "commentList1")

            if len(comment_lists) < 2:
                print("두 번째 commentList1이 없습니다.")

                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                print(f"obj : {page_data}")
                comments.append(page_data)

                return comments

            # 두 번째 commentList1에서 ul -> li를 찾아서 댓글 추출
            second_comment_list = comment_lists[1]

            comment_items = []

            # ul 태그가 존재하는지 확인
            ul_elements = second_comment_list.find_elements(By.TAG_NAME, "ul")
            if len(ul_elements) > 0:
                # ul이 존재하면 그 안에서 li 태그 찾기
                li_elements = ul_elements[0].find_elements(By.TAG_NAME, "li")
                if len(li_elements) > 0:
                    comment_items = li_elements


            if not comment_items:  # 댓글이 없을 경우
                # 리플 데이터가 없으면 빈 값으로 하나의 배열을 리턴
                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                comments.append(page_data)
                print(f"obj : {page_data}")
                return comments
            else:
                for idx, item in enumerate(comment_items, start=1):
                    comment_data = {"리플 번호": idx}

                    try:
                        # 리플 아이디 추출
                        comment_data["리플 아이디"] = item.find_element(By.CLASS_NAME, "nickname").text
                    except NoSuchElementException:
                        comment_data["리플 아이디"] = ""

                    try:
                        # 리플 내용 추출
                        comment_data["리플 내용"] = item.find_element(By.CLASS_NAME, "content.cmtContentOne").text
                    except NoSuchElementException:
                        comment_data["리플 내용"] = ""

                    try:
                        # 리플 날짜 추출
                        comment_data["리플 날짜"] = item.find_element(By.CLASS_NAME, "date").text
                    except NoSuchElementException:
                        comment_data["리플 날짜"] = ""

                    # 공통 데이터 병합
                    obj = {**page_data, **comment_data}
                    print(f"obj : {obj}")
                    comments.append(obj)
                return comments

        except NoSuchElementException as e:
            print(f"Error extracting comments: {e}")
            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

    except Exception as e:
        print(f"Error processing {url}: {e}")
        return []

# 페이지에서 링크 추출
def extract_links_inven(query, page):
    try:
        url = f"https://bbs.ruliweb.com/search?q={query}&op=&page={page}#board_search&gsc.tab=0&gsc.q={query}&gsc.page=1"
        print(f"Fetching URL: {url}")
        headers = {
            "authority": "bbs.ruliweb.com",
            "method": "GET",
            "scheme": "https",
            "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
            "cache-control": "max-age=0",
            "cookie": "ruli_board_font=; _ga=GA1.1.879995571.1731767001; __igaw__adid=NzAwPWRiOTkxOWY1ZTkyNTRiYzRmMmNmYmIzNGVhNjM3YTRhOzAwMD0wNjJiOTc4NC04OTMxLTExZWYtOTM3Ny0wMjQyYWMxMTAwMDI=; dable_uid=81234842.1725213932429; __dbl_v=1; session_key=3AmOQmAiIIQBsmm6; ad4989ad4989=1; _pubcid=ed3d4e1a-f369-4551-a6b5-190fcc7bba2d; _pubcid_cst=zix7LPQsHA%3D%3D; panoramaId_expiry=1732374265083; _cc_id=fc8c9804394c04519f4f0ea0a620cc34; panoramaId=3dcdfedecd3cdc6bf59d0e2bd03616d539385c1dede6a341753a5d252be3ae13; __gsas=ID=51e797045032c6fa:T=1731769477:RT=1731769477:S=ALNI_MYZ59MUOI5OQ4P1RvqVIcFAvsL9pQ; pbjs-unifiedid=%7B%22TDID%22%3A%22af80919c-ef84-4080-8e16-ce1f830c7fc1%22%2C%22TDID_LOOKUP%22%3A%22TRUE%22%2C%22TDID_CREATED_AT%22%3A%222024-10-16T15%3A04%3A40%22%7D; pbjs-unifiedid_cst=zix7LPQsHA%3D%3D; adfit_sdk_id=3158cb60-701b-4edb-92de-eded4235a340; ra_l=3600250; login_redirect_url=https%3A%2F%2Fbbs.ruliweb.com%2Fsearch%3Fq%3D%25EB%25A7%2588%25EA%25B3%25B5%25EC%258A%25A4%25EC%258B%259C%26op%3D%26page%3D1; cfa5=60; __gads=ID=17e51649880101e1:T=1731767001:RT=1731776455:S=ALNI_MbU62cFyl7oV3BAG-vbwIftunYotQ; __gpi=UID=00000f968f0d02e3:T=1731767001:RT=1731776455:S=ALNI_MYyFpCkneX3yk7V-HkempGapQIyRw; __eoi=ID=726e175340070e65:T=1731767001:RT=1731776455:S=AA-AfjaEb4mwbsRplKRWYie5FOA5; push_id=201565; _ga_1MX4XJ9Y2Q=GS1.1.1731776456.4.1.1731776749.58.0.0; cto_bidid=ue3vEF9RVCUyRkdmVUJyM1duS3B0aWhrMjRhQUowSndEd1dJam5LJTJCc2N3RVJRUUZ6UHI4NkFxd3BTZ2R2bEwyd2VrZGs4ZVdtTSUyRkM2bDdHSWtqdTNiJTJCcmE1dEo5bVVvczRMMm80dDFGdGhvVjhOYVprJTNE; cto_bundle=AD3JGV9BT3BYZUQ4YzRwSGxoUmxqWCUyRjdPV0hJMURDMHRKZyUyQjJ0T2RodklGc0dSdGtBUEhmTCUyQk5BQzdieWpLQ2tNRzh2U05XUktsVk5ITTdWMHRBWEZJNDA1bVhSY2VhZ3AwMDI0RVdGRHlxR1olMkZzc3dmT1Q4ZlM1M01EeVhEb1RLJTJCTm9NRHJDM1lzbWxESDJLeiUyRnlzdTFyJTJCQmZOQVdDM3I0T0ZXTFBwT1FoVWppb0pjd1BQdUlGQjdlTFpzbzNRb05odDYxWFZSekQlMkZIbVBMUHhRbzVZNGtXdyUzRCUzRA",
            "sec-ch-ua": "\"Chromium\";v=\"130\", \"Google Chrome\";v=\"130\", \"Not?A_Brand\";v=\"99\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "none",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()

        soup = BeautifulSoup(response.text, "html.parser")

        # id="board_search" 안의 ul.search_result_list 찾기
        search_result_list = soup.find("div", {"id": "board_search"}).find("ul", {"class": "search_result_list"})
        if not search_result_list:
            print(f"No search results found on page {page}")
            return []

        # li 태그 내 a 태그의 href 값 추출
        links = []
        for item in search_result_list.find_all("li", {"class": "search_result_item"}):
            link_element = item.find("a", {"class": "title text_over"})
            if link_element and link_element.get("href"):
                links.append(link_element["href"])

        return links

    except Exception as e:
        print(f"Error processing page {page}: {e}")
        return []




def get_page(url, headers):
    """
    주어진 URL로 GET 요청을 보내고 응답을 반환합니다.
    """
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response
    else:
        print(f"Failed to retrieve {url}. Status code: {response.status_code}")
        return None


def get_links(keyword):
    base_url = f'https://www.inven.co.kr/search/webzine/article/{keyword}/'
    headers = {
        'authority': 'www.inven.co.kr',
        'method': 'GET',
        'scheme': 'https',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-encoding': 'gzip, deflate, br, zstd',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        'cookie': 'OAX=2pOE7Gc48vgADVAa; _fwb=139MSl6E6DfQ2lfW1GBv8KU.1731785466460; a1_gid=2pOE7GcG/NMAC9Ch; a1_sgid=2pOE7GcG/NMAC9Ch1731785466907; _ga=GA1.1.2050071318.1731785467; _clck=1k33mp8%7C2%7Cfqx%7C0%7C1781; inven_link=stream; VISIT_SITE=webzine%7Cgirlsfrontline; invenrchk=%7B%225078%7C290350%22%3A%7B%22d%22%3A%222024-11-17%2004%3A33%3A22%22%2C%22s%22%3Anull%7D%2C%222097%7C2027852%22%3A%7B%22d%22%3A%222024-11-17%2004%3A53%3A46%22%2C%22s%22%3Anull%7D%7D; topskyCnt=36; _ga_K8NJQS78X7=GS1.1.1731785467.1.1.1731788734.0.0.0; _ga_P0BBTC1DR3=GS1.1.1731785466.1.1.1731788734.59.0.0; wcs_bt=105811aa1e895d:1731788734; _clsk=1ha1idj%7C1731788734991%7C33%7C0%7Cx.clarity.ms%2Fcollect; cto_bundle=5g4suV9iemRkZFpEVTcwQnlCbWczJTJGSWNNS3pMa0V6eXVuTjM2UWdhTlJONHpZMkMxdEtQTm84Y29ocUIwMkFMWDJLdk9yZ2ZBMGdwbGQ5UFN3JTJCdmo5emlDdkFkdFVsTFJWRVhCb210MkppJTJCVGdEOFhqUnN1N1RuaG5xZnJXbFFKUTJZSGx1cXJXVXBuUjNDNFRuako5JTJGYXZpaEZkNlBDdldwZVVKZVB2OVBZRGhldTl0MGNycmpQNVpBRVBBcnRKZDhBaWVWZ1NlU2VkT29WZkZJenhvMWxKWWclM0QlM0Q',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'document',
        'sec-fetch-mode': 'navigate',
        'sec-fetch-site': 'none',
        'sec-fetch-user': '?1',
        'upgrade-insecure-requests': '1'
    }
    article_links = []
    page = 1

    while True:
        url = f"{base_url}{page}"
        print(f"Fetching: {url}")

        # 페이지 요청
        response = get_page(url, headers)
        if not response:
            break

        soup = BeautifulSoup(response.content, 'html.parser')
        news_list = soup.find('ul', class_='news_list')

        if not news_list:
            print("No news list found. Exiting.")
            break

        # Extract article links from the list
        li_elements = news_list.find_all('li')
        current_page_links = []
        for li in li_elements:
            h1_tag = li.find('h1')
            if h1_tag:
                a_tag = h1_tag.find('a')
                if a_tag:
                    href = a_tag.get('href')
                    current_page_links.append(href)

        # If there are no new links, exit the loop
        if len(current_page_links) == 0:
            print(f"No articles found on page {page}. Exiting.")
            break

        # Check if the current and next href are the same
        if len(article_links) > 0 and article_links[-1] == current_page_links[-1]:
            break

        print(f'current_page_links : {current_page_links}')
        article_links.extend(current_page_links)
        page += 1
        time.sleep(random.uniform(2, 3))

    return article_links

# 스크린샷 폴더 설정
IMAGE_FOLDER = "inven_image_list"
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# 드라이버 세팅
def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--headless")
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
        driver.maximize_window()

        return driver
    except Exception as e:
        print(f"Error setting up the WebDriver: {e}")
        return None


# 전체 페이지 스크린샷 캡처
def capture_full_page_screenshot(driver, file_path):
    try:
        # 페이지 전체 크기 가져오기
        total_width = driver.execute_script("return document.body.scrollWidth")
        total_height = driver.execute_script("return document.body.scrollHeight")
        viewport_height = driver.execute_script("return window.innerHeight")

        # 스크롤 단계와 캡처된 이미지를 저장할 리스트
        scroll_steps = range(0, total_height, viewport_height)
        screenshot_parts = []

        for step in scroll_steps:
            # 스크롤 위치 이동
            driver.execute_script(f"window.scrollTo(0, {step});")
            time.sleep(0.3)  # 스크롤 대기

            # 현재 뷰포트 캡처
            screenshot_part_path = f"{file_path}_part_{step}.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 마지막 스크롤에서 남은 높이 처리
        if total_height % viewport_height > 0:
            driver.execute_script(f"window.scrollTo(0, {total_height - viewport_height});")
            time.sleep(0.3)
            screenshot_part_path = f"{file_path}_part_final.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 이미지 결합
        stitched_image = Image.new("RGB", (total_width, total_height))
        current_height = 0

        for idx, part_path in enumerate(screenshot_parts):
            with Image.open(part_path) as part_image:
                # 현재 캡처된 이미지 크기 가져오기
                part_width, part_height = part_image.size

                # 마지막 스크롤 조정
                if idx == len(screenshot_parts) - 1 and total_height % viewport_height > 0:
                    part_image = part_image.crop((0, part_height - (total_height % viewport_height), part_width, part_height))

                stitched_image.paste(part_image, (0, current_height))
                current_height += part_image.size[1]

            os.remove(part_path)  # 임시 파일 삭제

        # 최종 스크린샷 저장
        final_path = f"{file_path}_full.png"
        stitched_image.save(final_path)
        print(f"Full page screenshot saved: {final_path}")
        return final_path

    except Exception as e:
        print(f"Error capturing full page screenshot: {e}")
        return None


# 페이지 데이터 추출 함수
def extract_page_data(driver, url, keyword):
    try:
        driver.get(url)
        time.sleep(3)
        # 댓글 데이터 추출
        comments = []

        # 스크린샷 저장
        screenshot_path = os.path.join(IMAGE_FOLDER, f"inven_{url.split('/')[-1]}.png")
        full_screenshot_path = capture_full_page_screenshot(driver, screenshot_path)

        # 공통 데이터 추출
        page_data = {
            "글 번호": url.split("/")[-1],
            "제목": "",
            "내용": "",
            "아이디": "",
            "작성일": "",
            "키워드": keyword,
            "url": url,
            "스크린샷": full_screenshot_path,
        }

        # 제목, 내용, 사용자 정보 추출
        try:
            page_data["제목"] = driver.find_element(By.CLASS_NAME, "articleTitle").text
            page_data["내용"] = driver.find_element(By.ID, "powerbbsContent").text

            articleInfo = driver.find_element(By.CLASS_NAME, "articleInfo")
            page_data["아이디"] = articleInfo.find_element(By.CLASS_NAME, "articleWriter").text
            page_data["작성일"] = articleInfo.find_element(By.CLASS_NAME, "articleDate").text
        except NoSuchElementException as e:
            print(f"Error extracting main data: {e}")

            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments


        try:
            # 두 번째 class="commentList1"을 찾기
            comment_lists = driver.find_elements(By.CLASS_NAME, "commentList1")

            if len(comment_lists) < 2:
                print("두 번째 commentList1이 없습니다.")

                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                print(f"obj : {page_data}")
                comments.append(page_data)

                return comments

            # 두 번째 commentList1에서 ul -> li를 찾아서 댓글 추출
            second_comment_list = comment_lists[1]

            comment_items = []

            # ul 태그가 존재하는지 확인
            ul_elements = second_comment_list.find_elements(By.TAG_NAME, "ul")
            if len(ul_elements) > 0:
                # ul이 존재하면 그 안에서 li 태그 찾기
                li_elements = ul_elements[0].find_elements(By.TAG_NAME, "li")
                if len(li_elements) > 0:
                    comment_items = li_elements


            if not comment_items:  # 댓글이 없을 경우
                # 리플 데이터가 없으면 빈 값으로 하나의 배열을 리턴
                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                comments.append(page_data)
                print(f"obj : {page_data}")
                return comments
            else:
                for idx, item in enumerate(comment_items, start=1):
                    comment_data = {"리플 번호": idx}

                    try:
                        # 리플 아이디 추출
                        comment_data["리플 아이디"] = item.find_element(By.CLASS_NAME, "nickname").text
                    except NoSuchElementException:
                        comment_data["리플 아이디"] = ""

                    try:
                        # 리플 내용 추출
                        comment_data["리플 내용"] = item.find_element(By.CLASS_NAME, "content.cmtContentOne").text
                    except NoSuchElementException:
                        comment_data["리플 내용"] = ""

                    try:
                        # 리플 날짜 추출
                        comment_data["리플 날짜"] = item.find_element(By.CLASS_NAME, "date").text
                    except NoSuchElementException:
                        comment_data["리플 날짜"] = ""

                    # 공통 데이터 병합
                    obj = {**page_data, **comment_data}
                    print(f"obj : {obj}")
                    comments.append(obj)
                return comments

        except NoSuchElementException as e:
            print(f"Error extracting comments: {e}")
            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

    except Exception as e:
        print(f"Error processing {url}: {e}")
        return []


def save_or_append_to_excel(data, filename="inven_results.xlsx"):
    df = pd.DataFrame(data)

    try:
        # 기존 파일이 있을 경우 데이터를 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(filename)
            sheet_name = workbook.sheetnames[0]  # 첫 번째 시트 이름 가져오기
            # 기존 데이터의 마지막 행 번호 계산
            if writer.sheets.get(sheet_name):
                startrow = writer.sheets[sheet_name].max_row
            else:
                startrow = 0
            # 데이터 추가
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
            print(f"Data successfully appended to {filename}")
    except FileNotFoundError:
        # 파일이 없을 경우 새 파일 생성
        df.to_excel(filename, index=False)
        print(f"New file created: {filename}")


# 실행
if __name__ == "__main__":
    keywords = [
        # "마공스시",
        # "읍읍스시",
        # "마공읍읍",
        # "ㅁㄱㅅㅅ",
        "ㅁㄱ스시",
        # "신지수",
        # "ㅅㅈㅅ",
        "보일러집 아들",
        "대열보일러",
        "project02",
        "버블트리"
    ]
    driver = setup_driver()
    if not driver:
        print("Driver setup failed!")
        exit()
    try:
        for keyword in keywords:
            print(f"Processing keyword: {keyword}")
            result_links = get_links(keyword)
            if not result_links:
                continue
            results = []
            for index, link in enumerate(result_links):
                data = extract_page_data(driver, link, keyword)
                if data:
                    results.extend(data)

            # 엑셀 저장 또는 추가
            print(f'results len : {len(results)}')
            if results:
                save_or_append_to_excel(results)

    finally:
        driver.quit()

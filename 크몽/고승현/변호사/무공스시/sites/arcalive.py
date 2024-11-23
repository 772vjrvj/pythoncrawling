import requests
from bs4 import BeautifulSoup
import random
import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
from urllib.parse import urlparse

# 페이지 데이터 추출 함수
def extract_contents_arcalive(driver, link, keyword):
    try:
        # 댓글을 포함한 전체 글 목록
        comments = []

        driver.get(link)
        time.sleep(3)

        # URL 파싱
        parsed_url = urlparse(link)
        path_parts = parsed_url.path.split('/')  # 경로를 '/' 기준으로 분리, 경로에서 마지막 부분 추출
        last_part = path_parts[-1]  # 마지막 부분이 '115901698'

        # 스크린샷 저장
        screenshot_path = os.path.join(IMAGE_FOLDER, f"arcalive_{last_part}.png")
        full_screenshot_path = capture_full_page_screenshot(driver, screenshot_path)

        # 공통 데이터 추출
        page_data = {
            "사이트": '아카라이브',
            "글 번호": last_part,
            "제목": "",
            "내용": "",
            "아이디": "",
            "작성일": "",
            "키워드": keyword,
            "url": link,
            "스크린샷": full_screenshot_path,
        }

        # 제목, 내용, 사용자 정보 추출
        try:
            article_head = driver.find_element(By.CLASS_NAME, "article-head")

            # 제목
            title_row = article_head.find_element(By.CLASS_NAME, "title-row")
            title_div = title_row.find_element(By.CLASS_NAME, "title")

            # span 태그를 제외한 텍스트만 추출
            spans = title_div.find_elements(By.TAG_NAME, "span")

            # span 태그 내부 텍스트는 제외하고 나머지 텍스트 추출
            text = title_div.text.strip()

            # span 태그 텍스트를 제외한 부분을 반환
            for span in spans:
                text = text.replace(span.text, "").strip()

            page_data["제목"] = text

            # 아이디
            info_row = article_head.find_element(By.CLASS_NAME, "info-row")
            user_info = info_row.find_element(By.CLASS_NAME, "user-info")
            page_data["아이디"] = user_info.text

            # 작성일
            date_element = info_row.find_element(By.CLASS_NAME, "date")
            page_data["작성일"] = date_element.find_element(By.TAG_NAME, 'time').text

            # 내용
            page_data["내용"] = driver.find_element(By.CLASS_NAME, "article-body").text

        except NoSuchElementException as e:
            print(f"Error extracting main data: {e}")

            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

        try:
            # 두 번째 class="commentList1"을 찾기
            list_area = driver.find_element(By.CLASS_NAME, "list-area")
            comment_item_list = list_area.find_elements(By.CLASS_NAME, "comment-item")

            if not comment_item_list:  # 댓글이 없을 경우
                # 리플 데이터가 없으면 빈 값으로 하나의 배열을 리턴
                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                comments.append(page_data)
                print(f"obj : {page_data}")
                return comments
            else:
                for idx, item in enumerate(comment_item_list, start=1):
                    comment_data = {"리플 번호": idx}

                    try:
                        # 리플 아이디 추출
                        reply_user_info = item.find_element(By.CLASS_NAME, "user-info")
                        comment_data["리플 아이디"] = reply_user_info.find_element(By.TAG_NAME, 'a').text
                    except NoSuchElementException:
                        comment_data["리플 아이디"] = ""

                    try:
                        # 리플 내용 추출
                        comment_data["리플 내용"] = item.find_element(By.CLASS_NAME, "text").text
                    except NoSuchElementException:
                        comment_data["리플 내용"] = ""

                    try:
                        # 리플 날짜 추출
                        comment_data["리플 날짜"] = item.find_element(By.TAG_NAME, 'time').text
                    except NoSuchElementException:
                        comment_data["리플 날짜"] = ""

                    # 공통 데이터 병합
                    obj = {**page_data, **comment_data}
                    print(f"obj : {obj}")
                    comments.append(obj)

                return comments

        except NoSuchElementException as e:
            print(f"Error extracting comments: {e}")
            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

    except Exception as e:
        print(f"Error processing {link}: {e}")
        return []

def extract_links_arcalive(keyword):
    base_url = f'https://arca.live/b/breaking?keyword={keyword}'
    headers = {
        "authority": "arca.live",
        "method": "GET",
        "scheme": "https",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "max-age=0",
        "cookie": "_sharedid=e4b3d565-87b3-4f57-8d09-cd9ff8392b93; _sharedid_cst=zix7LPQsHA%3D%3D; _lr_env_src_ats=false; _ga=GA1.1.698925934.1731785623; _au_1d=AU1D-0100-001731785623-UEJCV9QC-QMB9; arca.nick=%E3%85%87%E3%85%87; arca.nick.sig=ETcGz1yS4GbrdOJZZg18BIpnFrg; arca.password=KinjoEgL; arca.password.sig=uy3DCL6W45baJO-eI2UNfLA3UvQ; visited-channel=[{%22name%22:%22%EC%A2%85%ED%95%A9%20%EC%86%8D%EB%B3%B4%22%2C%22slug%22:%22breaking%22}]; campaign.session=s%3ArjAoc2WMdIxHfaRRrDORTLNX9zc0czIz.fkf4Z%2FFMjln%2FLLz1SiZo3m5QqWyQm90jFxACLCUVMaY; _lr_retry_request=true; arca.csrf=-QznQ9id3ajsLFv0AlN_k4jE; arca.csrf.sig=HIvL_2GHENskn7NoSadIsLXrlpM; _ga_EVNC8JD9DJ=GS1.1.1731820998.2.1.1731821128.0.0.0; cto_bundle=V8A0pF9pRXA0c21DeUxuMHR0alF5OUhXd28zWVR0WkJ1UFMxR29rMDNydFNYRVhsTTBCcWZtY2pCNEFwZmc1S0FwODBDU3hqVktjdXVQMVpTR25uVzhqRmxEbFhQTVclMkJLOXV6Q3FMRE8yN0k1T3BRVmdpTHlNRkslMkIyTnRwd0pyYzhjQUtJak00dERLUXNvaWU3UGV6eEhNUHh3JTNEJTNE; cto_bidid=VEn6nl9lJTJGRWc1Wm90QkJobWtJbmYlMkJLd2R0aXJtSEY2TWpjd0FHUEdHTzFvaERtQlR3Mm9CWlBTb1hhQzY4TDczTjIzTjlRdHZQb2pvZVd5OVpFMFVGVzlGa3c1TXJWWVlMVDVxSmh0WU9OSFNSMTAlM0Q",
        "priority": "u=0, i",
        "sec-ch-ua": '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        "sec-ch-ua-arch": "x86",
        "sec-ch-ua-bitness": "64",
        "sec-ch-ua-full-version": "130.0.6723.119",
        "sec-ch-ua-full-version-list": '"Chromium";v="130.0.6723.119", "Google Chrome";v="130.0.6723.119", "Not?A_Brand";v="99.0.0.0"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-model": "",
        "sec-ch-ua-platform": "Windows",
        "sec-ch-ua-platform-version": "10.0.0",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
    }
    article_links = []
    page = 1

    while True:
        url = f"{base_url}&p={page}"
        print(f"Fetching: {url}")

        # 페이지 요청
        response = get_page(url, headers)
        if not response:
            break

        soup = BeautifulSoup(response.content, 'html.parser')
        # 조건에 맞는 a 태그 찾기
        # 사실 아래 조건만으로도 이미 notice와 notice-service는 제외됨
        a_tags = soup.find_all("a", class_="vrow column")
        hrefs = []

        if a_tags:
            for tag in a_tags:
                # class 속성 가져오기
                classes = tag.get("class", [])

                # 'notice'와 'notice-service'가 포함되지 않은 경우 href 추출
                if "notice" not in classes and "notice-service" not in classes:
                    href = tag.get("href")
                    if href:
                        hrefs.append(f"https://arca.live{href}")

        if not hrefs:
            print("No list found. Exiting.")
            break

        article_links.extend(hrefs)

        page += 1
        time.sleep(random.uniform(2, 3))

    return article_links






def get_page(url, headers):
    """
    주어진 URL로 GET 요청을 보내고 응답을 반환합니다.
    """
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response
    else:
        print(f"Failed to retrieve {url}. Status code: {response.status_code}")
        return None


def get_links(keyword):
    base_url = f'https://arca.live/b/breaking?keyword={keyword}'
    headers = {
        "authority": "arca.live",
        "method": "GET",
        "scheme": "https",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "max-age=0",
        "cookie": "_sharedid=e4b3d565-87b3-4f57-8d09-cd9ff8392b93; _sharedid_cst=zix7LPQsHA%3D%3D; _lr_env_src_ats=false; _ga=GA1.1.698925934.1731785623; _au_1d=AU1D-0100-001731785623-UEJCV9QC-QMB9; arca.nick=%E3%85%87%E3%85%87; arca.nick.sig=ETcGz1yS4GbrdOJZZg18BIpnFrg; arca.password=KinjoEgL; arca.password.sig=uy3DCL6W45baJO-eI2UNfLA3UvQ; visited-channel=[{%22name%22:%22%EC%A2%85%ED%95%A9%20%EC%86%8D%EB%B3%B4%22%2C%22slug%22:%22breaking%22}]; campaign.session=s%3ArjAoc2WMdIxHfaRRrDORTLNX9zc0czIz.fkf4Z%2FFMjln%2FLLz1SiZo3m5QqWyQm90jFxACLCUVMaY; _lr_retry_request=true; arca.csrf=-QznQ9id3ajsLFv0AlN_k4jE; arca.csrf.sig=HIvL_2GHENskn7NoSadIsLXrlpM; _ga_EVNC8JD9DJ=GS1.1.1731820998.2.1.1731821128.0.0.0; cto_bundle=V8A0pF9pRXA0c21DeUxuMHR0alF5OUhXd28zWVR0WkJ1UFMxR29rMDNydFNYRVhsTTBCcWZtY2pCNEFwZmc1S0FwODBDU3hqVktjdXVQMVpTR25uVzhqRmxEbFhQTVclMkJLOXV6Q3FMRE8yN0k1T3BRVmdpTHlNRkslMkIyTnRwd0pyYzhjQUtJak00dERLUXNvaWU3UGV6eEhNUHh3JTNEJTNE; cto_bidid=VEn6nl9lJTJGRWc1Wm90QkJobWtJbmYlMkJLd2R0aXJtSEY2TWpjd0FHUEdHTzFvaERtQlR3Mm9CWlBTb1hhQzY4TDczTjIzTjlRdHZQb2pvZVd5OVpFMFVGVzlGa3c1TXJWWVlMVDVxSmh0WU9OSFNSMTAlM0Q",
        "priority": "u=0, i",
        "sec-ch-ua": '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        "sec-ch-ua-arch": "x86",
        "sec-ch-ua-bitness": "64",
        "sec-ch-ua-full-version": "130.0.6723.119",
        "sec-ch-ua-full-version-list": '"Chromium";v="130.0.6723.119", "Google Chrome";v="130.0.6723.119", "Not?A_Brand";v="99.0.0.0"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-model": "",
        "sec-ch-ua-platform": "Windows",
        "sec-ch-ua-platform-version": "10.0.0",
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
    }
    article_links = []
    page = 1

    while True:
        url = f"{base_url}&p={page}"
        print(f"Fetching: {url}")

        # 페이지 요청
        response = get_page(url, headers)
        if not response:
            break

        soup = BeautifulSoup(response.content, 'html.parser')
        # 조건에 맞는 a 태그 찾기
        # 사실 아래 조건만으로도 이미 notice와 notice-service는 제외됨
        a_tags = soup.find_all("a", class_="vrow column")
        hrefs = []

        if a_tags:
            for tag in a_tags:
                # class 속성 가져오기
                classes = tag.get("class", [])

                # 'notice'와 'notice-service'가 포함되지 않은 경우 href 추출
                if "notice" not in classes and "notice-service" not in classes:
                    href = tag.get("href")
                    if href:
                        hrefs.append(f"https://arca.live{href}")

        if not hrefs:
            print("No list found. Exiting.")
            break

        article_links.extend(hrefs)

        page += 1
        time.sleep(random.uniform(2, 3))

    return article_links

# 스크린샷 폴더 설정
IMAGE_FOLDER = "arcalive_image_list"
os.makedirs(IMAGE_FOLDER, exist_ok=True)

# 드라이버 세팅
def setup_driver():
    try:
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--incognito")
        chrome_options.add_argument("--headless")
        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
        driver.maximize_window()

        return driver
    except Exception as e:
        print(f"Error setting up the WebDriver: {e}")
        return None


# 전체 페이지 스크린샷 캡처
def capture_full_page_screenshot(driver, file_path):
    try:
        # 페이지 전체 크기 가져오기
        total_width = driver.execute_script("return document.body.scrollWidth")
        total_height = driver.execute_script("return document.body.scrollHeight")
        viewport_height = driver.execute_script("return window.innerHeight")

        # 스크롤 단계와 캡처된 이미지를 저장할 리스트
        scroll_steps = range(0, total_height, viewport_height)
        screenshot_parts = []

        for step in scroll_steps:
            # 스크롤 위치 이동
            driver.execute_script(f"window.scrollTo(0, {step});")
            time.sleep(0.3)  # 스크롤 대기

            # 현재 뷰포트 캡처
            screenshot_part_path = f"{file_path}_part_{step}.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 마지막 스크롤에서 남은 높이 처리
        if total_height % viewport_height > 0:
            driver.execute_script(f"window.scrollTo(0, {total_height - viewport_height});")
            time.sleep(0.3)
            screenshot_part_path = f"{file_path}_part_final.png"
            driver.save_screenshot(screenshot_part_path)
            screenshot_parts.append(screenshot_part_path)

        # 이미지 결합
        stitched_image = Image.new("RGB", (total_width, total_height))
        current_height = 0

        for idx, part_path in enumerate(screenshot_parts):
            with Image.open(part_path) as part_image:
                # 현재 캡처된 이미지 크기 가져오기
                part_width, part_height = part_image.size

                # 마지막 스크롤 조정
                if idx == len(screenshot_parts) - 1 and total_height % viewport_height > 0:
                    part_image = part_image.crop((0, part_height - (total_height % viewport_height), part_width, part_height))

                stitched_image.paste(part_image, (0, current_height))
                current_height += part_image.size[1]

            os.remove(part_path)  # 임시 파일 삭제

        # 최종 스크린샷 저장
        final_path = f"{file_path}_full.png"
        stitched_image.save(final_path)
        print(f"Full page screenshot saved: {final_path}")
        return final_path

    except Exception as e:
        print(f"Error capturing full page screenshot: {e}")
        return ""


# 페이지 데이터 추출 함수
def extract_page_data(driver, link, keyword):
    try:
        # 댓글을 포함한 전체 글 목록
        comments = []

        driver.get(link)
        time.sleep(3)

        # URL 파싱
        parsed_url = urlparse(link)
        path_parts = parsed_url.path.split('/')  # 경로를 '/' 기준으로 분리, 경로에서 마지막 부분 추출
        last_part = path_parts[-1]  # 마지막 부분이 '115901698'

        # 스크린샷 저장
        screenshot_path = os.path.join(IMAGE_FOLDER, f"arcalive_{last_part}.png")
        full_screenshot_path = capture_full_page_screenshot(driver, screenshot_path)

        # 공통 데이터 추출
        page_data = {
            "사이트": '아카라이브',
            "글 번호": last_part,
            "제목": "",
            "내용": "",
            "아이디": "",
            "작성일": "",
            "키워드": keyword,
            "url": link,
            "스크린샷": full_screenshot_path,
        }

        # 제목, 내용, 사용자 정보 추출
        try:
            article_head = driver.find_element(By.CLASS_NAME, "article-head")

            # 제목
            title_row = article_head.find_element(By.CLASS_NAME, "title-row")
            title_div = title_row.find_element(By.CLASS_NAME, "title")

            # span 태그를 제외한 텍스트만 추출
            spans = title_div.find_elements(By.TAG_NAME, "span")

            # span 태그 내부 텍스트는 제외하고 나머지 텍스트 추출
            text = title_div.text.strip()

            # span 태그 텍스트를 제외한 부분을 반환
            for span in spans:
                text = text.replace(span.text, "").strip()

            page_data["제목"] = text

            # 아이디
            info_row = article_head.find_element(By.CLASS_NAME, "info-row")
            user_info = info_row.find_element(By.CLASS_NAME, "user-info")
            page_data["아이디"] = user_info.text

            # 작성일
            date_element = info_row.find_element(By.CLASS_NAME, "date")
            page_data["작성일"] = date_element.find_element(By.TAG_NAME, 'time').text

            # 내용
            page_data["내용"] = driver.find_element(By.CLASS_NAME, "article-body").text

        except NoSuchElementException as e:
            print(f"Error extracting main data: {e}")

            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

        try:
            # 두 번째 class="commentList1"을 찾기
            list_area = driver.find_element(By.CLASS_NAME, "list-area")
            comment_item_list = list_area.find_elements(By.CLASS_NAME, "comment-item")

            if not comment_item_list:  # 댓글이 없을 경우
                # 리플 데이터가 없으면 빈 값으로 하나의 배열을 리턴
                page_data.update({
                    "리플 번호": "",
                    "리플 아이디": "",
                    "리플 내용": "",
                    "리플 날짜": ""
                })
                comments.append(page_data)
                print(f"obj : {page_data}")
                return comments
            else:
                for idx, item in enumerate(comment_item_list, start=1):
                    comment_data = {"리플 번호": idx}

                    try:
                        # 리플 아이디 추출
                        reply_user_info = item.find_element(By.CLASS_NAME, "user-info")
                        comment_data["리플 아이디"] = reply_user_info.find_element(By.TAG_NAME, 'a').text
                    except NoSuchElementException:
                        comment_data["리플 아이디"] = ""

                    try:
                        # 리플 내용 추출
                        comment_data["리플 내용"] = item.find_element(By.CLASS_NAME, "text").text
                    except NoSuchElementException:
                        comment_data["리플 내용"] = ""

                    try:
                        # 리플 날짜 추출
                        comment_data["리플 날짜"] = item.find_element(By.TAG_NAME, 'time').text
                    except NoSuchElementException:
                        comment_data["리플 날짜"] = ""

                    # 공통 데이터 병합
                    obj = {**page_data, **comment_data}
                    print(f"obj : {obj}")
                    comments.append(obj)

                return comments

        except NoSuchElementException as e:
            print(f"Error extracting comments: {e}")
            page_data.update({
                "리플 번호": "",
                "리플 아이디": "",
                "리플 내용": "",
                "리플 날짜": ""
            })
            print(f"obj : {page_data}")
            comments.append(page_data)
            return comments

    except Exception as e:
        print(f"Error processing {link}: {e}")
        return []


def save_or_append_to_excel(data, filename="arcalive_results.xlsx"):
    df = pd.DataFrame(data)

    try:
        # 기존 파일이 있을 경우 데이터를 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(filename)
            sheet_name = workbook.sheetnames[0]  # 첫 번째 시트 이름 가져오기
            # 기존 데이터의 마지막 행 번호 계산
            if writer.sheets.get(sheet_name):
                startrow = writer.sheets[sheet_name].max_row
            else:
                startrow = 0
            # 데이터 추가
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
            print(f"Data successfully appended to {filename}")
    except FileNotFoundError:
        # 파일이 없을 경우 새 파일 생성
        df.to_excel(filename, index=False)
        print(f"New file created: {filename}")


# 실행
if __name__ == "__main__":
    keywords = [
        "마공스시",
        "읍읍스시",
        "마공읍읍",
        "ㅁㄱㅅㅅ",
        "ㅁㄱ스시",
        "신지수",
        # "ㅅㅈㅅ", # 검색량이 너무 많은 ㅈㅅㅈㅅ(지성지성)
        "보일러집 아들",
        "대열보일러",
        "project02",
        "버블트리"
    ]
    driver = setup_driver()
    if not driver:
        print("Driver setup failed!")
        exit()
    try:
        for keyword in keywords:
            print(f"Processing keyword: {keyword}")
            result_links = get_links(keyword)
            if not result_links:
                continue
            results = []
            for index, link in enumerate(result_links):
                data = extract_page_data(driver, link, keyword)
                if data:
                    results.extend(data)

            # 엑셀 저장 또는 추가
            print(f'results len : {len(results)}')
            if results:
                save_or_append_to_excel(results)
    finally:
        driver.quit()

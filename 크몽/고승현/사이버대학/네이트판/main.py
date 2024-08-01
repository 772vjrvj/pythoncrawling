import requests
from bs4 import BeautifulSoup
import time
import random
import os
from openpyxl import Workbook, load_workbook

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'
}

base_url = "https://pann.nate.com"

def get_search_results(q, page):
    try:
        url = f"https://pann.nate.com/search/talk?q={q}&page={page}"
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        if not response.text.strip():
            print(f"No content received for search results page {page}")
            return []
    except requests.RequestException as e:
        print(f"Error fetching search results for page {page}: {e}")
        return []

    soup = BeautifulSoup(response.text, 'html.parser')
    results = []
    try:
        for li in soup.select('.s_list li'):
            a_tag = li.select_one('.tit a')
            if a_tag and 'href' in a_tag.attrs:
                full_url = base_url + a_tag['href']
                results.append(full_url)
    except Exception as e:
        print(f"Error parsing search results: {e}")
    return results

def get_post_details(url, index):
    print(f"[{index}] Fetching details from {url}")
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        if not response.text.strip():
            print(f"No content received from {url}")
            return None
    except requests.RequestException as e:
        print(f"Error fetching post details from {url}: {e}")
        return None

    obj = {}
    soup = BeautifulSoup(response.text, 'html.parser')
    try:
        obj['사이트'] = "네이트판"
        obj['URL'] = url
        obj['제목'] = soup.select_one('.post-tit-info h1').get_text(strip=True) if soup.select_one('.post-tit-info h1') else "No title"
        obj['원문'] = soup.select_one('.posting #contentArea').get_text(strip=True) if soup.select_one('.posting #contentArea') else "No content"
        obj['날짜'] = soup.select_one('.post-tit-info .date').get_text(strip=True) if soup.select_one('.post-tit-info .date') else "No date"
    except Exception as e:
        print(f"Error parsing post details from {url}: {e}")
        return None

    print(f"[{index}] obj: {obj}")
    return obj

def save_to_excel(data, file_name='pann_results.xlsx'):
    try:
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['사이트', 'URL', '제목', '원문', '날짜'])  # Header 추가

        for entry in data:
            ws.append([entry['사이트'], entry['URL'], entry['제목'], entry['원문'], entry['날짜']])

        wb.save(file_name)
        print(f"Data successfully saved to {file_name}")

    except Exception as e:
        print(f"Error saving to Excel: {e}")

def main(q):
    all_results = []
    post_index = 1
    page = 1
    while True:
        search_results = get_search_results(q, page)
        if not search_results:
            break
        for url in search_results:
            time.sleep(random.uniform(1, 2))  # 1~2초 사이의 랜덤 간격
            post_details = get_post_details(url, post_index)
            if post_details:
                all_results.append(post_details)
                post_index += 1
                if len(all_results) >= 100:
                    save_to_excel(all_results)
                    all_results = []
        page += 1


    if all_results:
        save_to_excel(all_results)

if __name__ == "__main__":
    query = "사이버대학교"  # 원하는 검색어로 바꿔주세요
    main(query)

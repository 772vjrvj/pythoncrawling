import requests
from bs4 import BeautifulSoup
import json
import re
import time
import random
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime


def fetch_search_results(query, page):
    try:
        url = f"https://map.naver.com/p/api/search/allSearch?query={query}&type=all&searchCoord=&boundary=&page={page}"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36',
            'Referer': ''
        }
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Failed to fetch search results: {e}")
    return None

def fetch_place_info(place_id):
    try:
        url = f"https://m.place.naver.com/place/{place_id}"
        headers = {
            'authority': 'm.place.naver.com',
            'method': 'GET',
            'scheme': 'https',
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'accept-encoding': 'gzip, deflate, br, zstd',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            'priority': 'u=0, i',
            'sec-ch-ua': '"Not/A)Brand";v="99", "Google Chrome";v="127", "Chromium";v="127"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'referer': '',
            'sec-fetch-user': '?1',
            'upgrade-insecure-requests': '1',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/127.0.0.0 Safari/537.36'
        }
        response = requests.get(url, headers=headers)
        response.encoding = 'utf-8'

        if response.status_code == 200:
            print(f'response.status_code : {response.status_code}')
            soup = BeautifulSoup(response.text, 'html.parser')
            script_tag = soup.find('script', string=re.compile('window.__APOLLO_STATE__'))

            if script_tag:
                json_text = re.search(r'window\.__APOLLO_STATE__\s*=\s*(\{.*\});', script_tag.string)
                if json_text:
                    data = json.loads(json_text.group(1))

                    name = data.get(f"PlaceDetailBase:{place_id}", {}).get("name", "")
                    address = data.get(f"PlaceDetailBase:{place_id}", {}).get("address", "")
                    roadAddress = data.get(f"PlaceDetailBase:{place_id}", {}).get("roadAddress", "")
                    category = data.get(f"PlaceDetailBase:{place_id}", {}).get("category", "")
                    conveniences = data.get(f"PlaceDetailBase:{place_id}", {}).get("conveniences", [])
                    visitorReviewsScore = data.get(f"PlaceDetailBase:{place_id}", {}).get("visitorReviewsScore", "")
                    visitorReviewsTotal = data.get(f"PlaceDetailBase:{place_id}", {}).get("visitorReviewsTotal", "")

                    root_query = data.get("ROOT_QUERY", {})
                    place_detail_key = f'placeDetail({{"input":{{"deviceType":"pc","id":"{place_id}","isNx":false}}}})'

                    # 기본 place_detail_key 값이 없으면 checkRedirect 포함된 key로 재시도
                    if place_detail_key not in root_query:
                        place_detail_key = f'placeDetail({{"input":{{"checkRedirect":true,"deviceType":"pc","id":"{place_id}","isNx":false}}}})'

                    fsasReviewsTotal = root_query.get(place_detail_key, {}).get('fsasReviews', {}).get("total", "")
                    if not fsasReviewsTotal:
                        fsasReviewsTotal = root_query.get(place_detail_key, {}).get("fsasReviews({\"fsasReviewsType\":\"restaurant\"})", {}).get("total", "")

                    # business_hours 초기 시도
                    business_hours = root_query.get(place_detail_key, {}).get("businessHours({\"source\":[\"tpirates\",\"shopWindow\"]})", [])

                    # business_hours 값이 없으면 다른 source를 시도
                    if not business_hours:
                        business_hours = root_query.get(place_detail_key, {}).get("businessHours({\"source\":[\"tpirates\",\"jto\",\"shopWindow\"]})", [])

                    new_business_hours_json = root_query.get(place_detail_key, {}).get('newBusinessHours', [])

                    if not new_business_hours_json:
                        new_business_hours_json = root_query.get(place_detail_key, {}).get("newBusinessHours({\"format\":\"restaurant\"})", [])

                    # 별점, 방문자 리뷰 수, 블로그 리뷰 수가 0이거나 없으면 공백 처리
                    visitorReviewsScore = visitorReviewsScore if visitorReviewsScore and visitorReviewsScore != "0" else ""
                    visitorReviewsTotal = visitorReviewsTotal if visitorReviewsTotal and visitorReviewsTotal != "0" else ""
                    fsasReviewsTotal = fsasReviewsTotal if fsasReviewsTotal and fsasReviewsTotal != "0" else ""

                    # category를 대분류와 소분류로 나누기
                    category_list = category.split(',') if category else ["", ""]
                    main_category = category_list[0] if len(category_list) > 0 else ""
                    sub_category = category_list[1] if len(category_list) > 1 else ""

                    url = f"https://m.place.naver.com/place/{place_id}/home"
                    map_url = f"https://map.naver.com/p/entry/place/{place_id}"

                    result = {
                        "이름": name,
                        "주소(지번)": address,
                        "주소(도로명)": roadAddress,
                        "대분류": main_category,
                        "소분류": sub_category,
                        "별점": visitorReviewsScore,
                        "방문자리뷰수": visitorReviewsTotal,
                        "블로그리뷰수": fsasReviewsTotal,
                        "이용시간1": format_business_hours(business_hours),
                        "이용시간2": format_new_business_hours(new_business_hours_json),
                        "카테고리": category,
                        "URL": url,
                        "지도": map_url,
                        "편의시설": ', '.join(conveniences) if conveniences else ''
                    }

                    print(f'name : {name}')

                    return result
    except requests.exceptions.RequestException as e:
        new_print(f"Failed to fetch data for Place ID: {place_id}. Error: {e}")
    except Exception as e:
        new_print(f"Error processing data for Place ID: {place_id}: {e}")

    print(f'None : {None}')
    return None


def format_business_hours(business_hours):
    formatted_hours = []
    try:
        if business_hours:
            for hour in business_hours:
                day = hour.get('day', '') or ''
                start_time = hour.get('startTime', '') or ''
                end_time = hour.get('endTime', '') or ''
                if day and start_time and end_time:
                    formatted_hours.append(f"{day} {start_time} - {end_time}")
    except Exception as e:
        new_print(f"Unexpected error: {e}")
        return ""
    return '\n'.join(formatted_hours).strip() if formatted_hours else ""


def format_new_business_hours(new_business_hours):
    formatted_hours = []
    try:
        if new_business_hours:
            for item in new_business_hours:
                status_description = item.get('businessStatusDescription', {}) or {}
                status = status_description.get('status', '') or ''
                description = status_description.get('description', '') or ''

                if status:
                    formatted_hours.append(status)
                if description:
                    formatted_hours.append(description)

                for info in item.get('businessHours', []) or []:
                    day = info.get('day', '') or ''
                    business_hours = info.get('businessHours', {}) or {}
                    start_time = business_hours.get('start', '') or ''
                    end_time = business_hours.get('end', '') or ''

                    break_hours = info.get('breakHours', []) or []
                    break_times = [f"{bh.get('start', '') or ''} - {bh.get('end', '') or ''}" for bh in break_hours]
                    break_times_str = ', '.join(break_times) + ' 브레이크타임' if break_times else ''

                    last_order_times = info.get('lastOrderTimes', []) or []
                    last_order_times_str = ', '.join([f"{lo.get('type', '')}: {lo.get('time', '')}" for lo in last_order_times]) + ' 라스트오더' if last_order_times else ''

                    if day:
                        formatted_hours.append(day)
                    if start_time and end_time:
                        formatted_hours.append(f"{start_time} - {end_time}")
                    if break_times_str:
                        formatted_hours.append(break_times_str)
                    if last_order_times_str:
                        formatted_hours.append(last_order_times_str)
    except Exception as e:
        print(f"Unexpected error: {e}")
        return ""
    return '\n'.join(formatted_hours).strip() if formatted_hours else ""


def new_print(text, level="INFO"):
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    formatted_text = f"[{timestamp}] [{level}] {text}"
    print(formatted_text)


def append_to_excel(data, filename="서울시 맛집1.xlsx"):
    df = pd.DataFrame(data)

    try:
        # 기존 파일이 있을 경우 파일을 열고 데이터를 추가
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            workbook = load_workbook(filename)
            sheet_name = workbook.sheetnames[0]  # 첫 번째 시트 이름 가져오기
            startrow = writer.sheets[sheet_name].max_row  # 기존 데이터의 마지막 행 번호
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)
    except FileNotFoundError:
        # 파일이 없을 경우 새로 생성
        df.to_excel(filename, index=False)


def main(query):
    try:
        page = 1
        results = []
        all_ids = set()

        new_print(f"크롤링 시작")
        while True:
            result = fetch_search_results(query, page)
            if not result:
                break

            place_list = result.get("result", {}).get("place", {}).get("list", [])
            ids_this_page = [place.get("id") for place in place_list if place.get("id")]

            new_print(f"페이지 : {page}, 목록 : {ids_this_page}")

            if not ids_this_page:
                break

            all_ids.update(ids_this_page)
            page += 1
            time.sleep(random.uniform(2, 3))

        all_ids_list = list(all_ids)
        total_count = len(all_ids_list)
        new_print(f"전체 매물 수 : {total_count}")
        new_print(f"전체 매물 : {all_ids_list}")

        for idx, place_id in enumerate(all_ids_list, start=1):
            print(f'place_id : {place_id}')
            place_info = fetch_place_info(place_id)
            if place_info:
                place_info["가게번호"] = place_id
                place_info["검색어"] = query
                place_info["식당이름"] = query.split()[-1]
                new_print(place_info)
                results.append(place_info)
                time.sleep(random.uniform(2, 3))

        return results

    except Exception as e:
        print(f"Unexpected error: {e}")


if __name__ == "__main__":
    names = [
        "인천송도 Pet Shop",
        "인천송도 키즈카페",
        "인천송도 산후조리원",
        "인천송도 학원",
        "인천송도 스터디카페",
        "인천송도 헬스장",
        "인천송도 필라테스, 요가",
        "인천송도 골프 아카데미",
        "인천송도 스크린골프",
        "인천송도 당구장",
        "인천송도 병원",
        "인천송도 안경점",
        "인천송도 식당",
        "인천송도 카페",
        "인천송도 정육점",
        "인천송도 반찬가게",
        "인천송도 자동차 정비소",
        "인천송도 문구점",
        "인천송도 전자용품 매장",
        "인천송도 세탁소",
        "인천송도 스튜디오",
        "인천송도 꽃집",
        "인천송도 미용실",
        "인천송도 네일샵",
        "인천송도 속눈썹펌",
        "인천송도 피부관리샵",
        "인천송도 마사지",
        "인천송도 화장품 가게",
        "인천송도 주얼리샵",
        "인천송도 옷가게",
        "인천송도 신발가게",
        "인천송도 동물병원",
        "인천송도 동물미용실",
    ]

    # 현재 시간 가져오기
    current_time = datetime.now().strftime("%Y%m%d%H%M")

    # 파일 이름에 시간 추가
    filename = f"상호명_{current_time}.xlsx"
    for index, name in enumerate(names, start=1):
        new_print(f"Total : {len(names)}, index : {index}, name : {name}==============================================")
        query = f"{name}"
        rs = main(query)
        if rs:
            append_to_excel(rs, filename=filename)

        new_print(f"============================================================================")

import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
import time
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# 드라이버 설정 함수
def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1080,750")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
    except Exception as e:
        print(f"Error setting up the driver: {e}")
        driver = None
    return driver


# 네이버 크롤링 함수
def scrape_naver(driver, name, naver_url):
    try:
        print("============================== 네이버 시작 ==============================")
        if not naver_url:
            return []
        driver.get(naver_url)
        print(naver_url)
        naver_temp_list = []
        time.sleep(3)

        # 카드할인 토클 클릭 ON으로 둘다 변경 (위 아래 있음)
        driver.find_elements(By.CSS_SELECTOR, '[data-shp-contents-type="카드할인가 정렬"]')[0].click()  # 카드할인
        time.sleep(0.5)

        # 옵션 이름 ex)수량, 개수, 상품구성 등등
        opt_name = driver.execute_script('return document.querySelector("#section_price em").closest("div");').text.split(" : ")[0].split(',')[-1]
        print(opt_name)

        #상품구성: 1개 아래 ['1개', '2개', '3개', '4개', '5개'] ...
        if len(driver.find_elements(By.CSS_SELECTOR, f'.condition_area a[data-shp-contents-type="{opt_name}"] .info')) != 0:
            qtys = driver.find_elements(By.CSS_SELECTOR, f'.condition_area a[data-shp-contents-type="{opt_name}"] .info')
        elif len(driver.find_elements(By.CSS_SELECTOR, '.condition_area a .info')) != 0:
            qtys = driver.find_elements(By.CSS_SELECTOR, '.condition_area a .info')  # 수량, 개수 옵션이 없다면 2번째 옵션으로 지정

        #['1개', '2개', '3개', '4개', '5개']
        qlist = [q.text for q in qtys]
        print(qlist)

        for p in range(len(qtys)):
            driver.find_element(By.CSS_SELECTOR, f'[data-shp-contents-id="{qlist[p]}"]').click()
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.execute_script("window.scrollTo(0, 0);")

            uls = driver.find_elements(By.CSS_SELECTOR, '#section_price ul')  # ul 목록
            ul_class = ""

            for e in uls:  # ul 안에 li class 이름 가져오기
                if 'productList_list_seller' in e.get_attribute('class'):
                    ul_class = e.get_attribute('class').replace(' ', '.')

            prod_list = driver.find_elements(By.CSS_SELECTOR, f'#section_price .{ul_class} li')

            action = ActionChains(driver)
            action.move_to_element(prod_list[0]).perform()
            time.sleep(1)

            for idx, e in enumerate(prod_list):
                try:
                    try:
                        mall_name = e.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.mall"] img').get_attribute('alt')
                    except:
                        mall_name = e.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.mall"]').text.split('\n')[0]

                    prod_name = e.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.pd"]').text
                    price = e.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.price"]').text.replace('최저', '').strip()

                    temp_list = [name, '네이버', qlist[p], mall_name, '', prod_name, price]

                    print(f"네이버 {idx} : {temp_list}")

                    naver_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping 네이버 (free shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

        print("============================== 네이버 끝 ==============================")
        return naver_temp_list

    except Exception as e:
        print(f"Error in Naver scraping: {e}")
        return []


# 다나와 크롤링 함수
def scrape_danawa(driver, name, danawa_url, limit_count, on_on_off):
    try:
        print("============================== 다나와 시작 ==============================")
        if not danawa_url:
            return []
        driver.get(danawa_url)
        print(danawa_url)
        danawa_temp_list = []

        if driver.find_elements(By.XPATH, '//*[@id="bundleProductMoreOpen"]') != []:   #구성 상품열기 #다른 구성상품5개 (아래화살표)
            driver.find_element(By.XPATH, '//*[@id="bundleProductMoreOpen"]').click()

        danawa_opt_url_list = [e.get_attribute('href') for e in driver.find_elements(By.CSS_SELECTOR, '[class="othr_list"] li .chk a')]
        danawa_opt_text_list = [e.text for e in driver.find_elements(By.CSS_SELECTOR, '[class="othr_list"] li .chk a')]
        # ['1개', '2개', '3개', '4개', '5개'] #['https://prod.danawa.com/info/?pcode=5970722', 'https://prod.danawa.com/info/?pcode=5970724', 'https://prod.danawa.com/info/?pcode=5970731', 'https://prod.danawa.com/info/?pcode=5970748', 'https://prod.danawa.com/info/?pcode=5970738']

        print(danawa_opt_url_list)
        print(danawa_opt_text_list)

        for ii in range(len(danawa_opt_url_list)):
            print(danawa_opt_url_list[ii])
            print(danawa_opt_text_list[ii])

            if on_on_off == 0 and danawa_opt_text_list[ii] == '1개':  # 1개는 스킵 # 복수 구성이 없는 상품은 에러처리안나게 건너 뛰기
                continue

            driver.get(danawa_opt_url_list[ii])
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.execute_script("window.scrollTo(0, 0);")
            driver.find_elements(By.CSS_SELECTOR, '.cardSaleChkbox')[0].click()  # 카드할인가 클릭
            time.sleep(1)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 좌측: 무료배송
            free_dil_prod_e_list = soup.select('.columm.left_col .diff_item')
            time.sleep(1)

            for ei in range(min(len(free_dil_prod_e_list), limit_count)):
                try:
                    try:
                        mall_name = soup.select('.columm.left_col .diff_item')[ei].select('img')[0].get('alt')
                    except:
                        mall_name = soup.select('.columm.left_col .diff_item')[ei].select('a .txt_logo')[0].text

                    prod_name = soup.select('.columm.left_col .diff_item')[ei].select('.info_line')[0].text.strip()
                    price = soup.select('.columm.left_col .diff_item')[ei].select('.prc_c')[0].text

                    temp_list = [name, '다나와', danawa_opt_text_list[ii], mall_name, '무료배송', prod_name, price]
                    print(f"다나와 {ei} : {temp_list}")
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (free shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

            # 우측: 유/무료 배송
            pay_dil_prod_e_list = soup.select('.columm.rgt_col .diff_item')
            time.sleep(1)

            for ei in range(min(len(pay_dil_prod_e_list), limit_count)):
                try:
                    try:
                        mall_name = soup.select('.columm.rgt_col .diff_item')[ei].select('img')[0].get('alt')
                    except:
                        mall_name = soup.select('.columm.rgt_col .diff_item')[ei].select('a .txt_logo')[0].text

                    prod_name = soup.select('.columm.rgt_col .diff_item')[ei].select('.info_line')[0].text.strip()
                    price = soup.select('.columm.rgt_col .diff_item')[ei].select('.prc_c')[0].text

                    temp_list = [name, '다나와', danawa_opt_text_list[ii], mall_name, '유/무료배송', prod_name, price]
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (paid shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

        print("============================== 다나와 끝 ==============================")
        return danawa_temp_list

    except Exception as e:
        print(f"Error in Danawa scraping: {e}")
        return []


# 에누리 크롤링 함수
def scrape_enuri(driver, name, enuri_url, limit_count, on_on_off):
    try:
        # 테스트를 위한 에러 발생
        # raise Exception("테스트를 위한 에러 발생")
        print("============================== 에누리 시작 ==============================")
        merge_list = []
        if not enuri_url:
            return []
        print(f"enuri_url : {enuri_url}")

        driver.get(enuri_url)
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(0.5)

        # 1개 2개 3개 4개 5개 라디오 아래 더보기 버튼
        if len(driver.find_elements(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more')) != 0:
            driver.find_element(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more').click()

        time.sleep(1)

        radio_opts = driver.find_elements(By.CSS_SELECTOR, '[name="radioOPTION"]')
        xpath_list = ['//*[@for="' + e.get_attribute('id') + '"]' for e in radio_opts]
        time.sleep(1)

        for eei, e in enumerate(radio_opts):
            time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);")
            elem = driver.find_element(By.XPATH, xpath_list[eei])  # 갯수 엘리먼트

            how_many = elem.text.split()[0]  # 갯수 텍스트
            if on_on_off == 0 and how_many == '1개':  # 1개는 스킵
                continue

            elem.click()  # 갯수 클릭
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.execute_script("window.scrollTo(0, 0);")
            time.sleep(1)
            # 카드 할인 토글 클릭
            try:
                # 두 개의 클래스를 가진 label 요소를 바로 찾습니다.
                label = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--card.inp-switch'))
                )

                # 요소가 화면에 보이도록 스크롤
                driver.execute_script("arguments[0].scrollIntoView(true);", label)
                time.sleep(1)  # 스크롤 후 잠시 대기

                # JavaScript로 강제로 클릭
                driver.execute_script("arguments[0].click();", label)
                print("카드할인 클릭")

            except Exception as e:
                print(f"Error clicking the label with JavaScript: {e}")

            max_wait_time = 10  # 최대 대기 시간 (초)
            start_time = time.time()

            while True:
                try:
                    # 로딩 상태가 아직 보이는지 확인
                    if driver.find_element(By.XPATH, '//*[@class="comm-loader"]').get_attribute('style') != "display: none;":
                        time.sleep(0.5)

                        # 최대 대기 시간이 넘으면 루프 탈출
                        if time.time() - start_time > max_wait_time:
                            print("Loader is taking too long. Skipping this step.")
                            break
                    else:
                        # 로딩이 끝났으면 루프 탈출
                        break
                except Exception as e:
                    # 에러가 발생했을 경우 에러 메시지 출력 후 루프 탈출
                    print(f"Error while waiting for loader: {e}")
                    print(traceback.format_exc())
                    break

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 'tb-compare__list' 클래스 내부의 tbody 태그에서 모든 tr 태그를 찾음
            free_dil_prod_e_list = soup.select('.tb-compare__list tbody tr')

            for ei in range(min(len(free_dil_prod_e_list), limit_count)):
                try:
                    # 판매처 추출
                    mall_name = ""
                    shop_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--shop')
                    if shop_td:
                        img_tag = shop_td.find('img')
                        mall_name = img_tag['alt'] if img_tag and 'alt' in img_tag.attrs else shop_td.get_text(strip=True)

                    # 제품명 추출
                    prod_name = ""
                    name_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--name')
                    if name_td:
                        name_div = name_td.find('div', class_='tb-col__tx--name')
                        if name_div:
                            # <i> 태그 제거 후 텍스트만 추출
                            [i_tag.extract() for i_tag in name_div.find_all('i')]
                            prod_name = name_div.get_text(strip=True)

                    # 가격 추출
                    price = ""
                    price_div = free_dil_prod_e_list[ei].find('div', class_='tx-price')
                    if price_div:
                        em_tag = price_div.find('em')
                        price = em_tag.get_text(strip=True) if em_tag else ""

                    # 리스트에 데이터 추가
                    temp_list = [name, '에누리', how_many, mall_name, '무료배송', prod_name, price]
                    print(f"에누리 {ei} : {temp_list}")
                    merge_list.append(temp_list)

                except Exception as e:
                    print(f"Error in scraping Enuri free shipping list (index {ei}): {e}")
                    continue

        print("============================== 에누리 끝 ==============================")
        return merge_list
    except Exception as e:
        print(f"Error in Enuri scraping: {e}")
        return []


# 행별로 데이터를 처리하고 엑셀에 업데이트하는 함수
def save_row_to_excel(ws, merge_list, row_index, err_list, five_per_mall_name):
    try:
        name = ws[f'E{row_index}'].value
        except_list = ws[f'A{row_index}'].value.split(',') if ws[f'A{row_index}'].value else []

        # 데이터를 필터링하고 가격/갯수 계산 후 병합 리스트 업데이트
        filtered_list = [entry for entry in merge_list if entry[0] == name]

        # 5% 할인을 적용할 상점 이름과 비교
        for entry in filtered_list:
            mall_name = entry[3]  # entry[3]은 mall_name을 의미
            if mall_name == five_per_mall_name:
                # entry[6]의 가격 부분에서 앞뒤의 글자를 분리하고, 중간 숫자만 추출
                price_str = entry[6]

                # 앞쪽의 글자 추출
                prefix = ''.join([c for c in price_str if not c.isdigit() and c != ','])

                # 숫자 추출
                numeric_part = ''.join([c for c in price_str if c.isdigit()])

                # 뒤쪽의 글자 추출
                suffix = ''.join([c for c in price_str[len(prefix)+len(numeric_part):] if not c.isdigit() and c != ','])

                # 숫자 부분만 5% 할인 적용
                price_value = int(numeric_part) * 0.95

                # 할인된 가격에 콤마 추가 후 앞뒤 글자와 함께 저장
                entry[6] = f"{prefix}{int(price_value):,}{suffix}"

        # 기준가격(네이버) 계산
        filtered_list_naver = [entry for entry in filtered_list if entry[1] == '네이버' and entry[3] not in except_list]
        if filtered_list_naver:
            naver_price = int(''.join(filter(str.isdigit, filtered_list_naver[0][6]))) / int(''.join(filter(str.isdigit, filtered_list_naver[0][2])))
            ws[f'G{row_index}'] = naver_price  # 기준가격(네이버) 셀

        # 나머지 판매처 및 상품 업데이트 (정렬 후)
        filtered_list.sort(key=lambda x: int(''.join(filter(str.isdigit, x[6]))) / int(''.join(filter(str.isdigit, x[2]))))

        if len(filtered_list) > 0:
            ws[f'H{row_index}'] = f"{filtered_list[0][1]}-{filtered_list[0][2]}-{filtered_list[0][3]}"  # 판매처1
            ws[f'I{row_index}'] = filtered_list[0][5]  # 상품명1
            ws[f'J{row_index}'] = int(''.join(filter(str.isdigit, filtered_list[0][6]))) / int(''.join(filter(str.isdigit, filtered_list[0][2])))  # 가격1
        if len(filtered_list) > 1:
            ws[f'K{row_index}'] = f"{filtered_list[1][1]}-{filtered_list[1][2]}-{filtered_list[1][3]}"  # 판매처2
            ws[f'L{row_index}'] = filtered_list[1][5]  # 상품명2
            ws[f'M{row_index}'] = int(''.join(filter(str.isdigit, filtered_list[1][6]))) / int(''.join(filter(str.isdigit, filtered_list[1][2])))  # 가격2
        if len(filtered_list) > 2:
            ws[f'N{row_index}'] = f"{filtered_list[2][1]}-{filtered_list[2][2]}-{filtered_list[2][3]}"  # 판매처3
            ws[f'O{row_index}'] = filtered_list[2][5]  # 상품명3
            ws[f'P{row_index}'] = int(''.join(filter(str.isdigit, filtered_list[2][6]))) / int(''.join(filter(str.isdigit, filtered_list[2][2])))  # 가격3

        # 배경색 설정 (빨간색)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # err_list 값에 따라 배경색 변경
        if err_list[0] == 1:
            ws[f'B{row_index}'].fill = red_fill  # "네이버 URL" 셀
        if err_list[1] == 1:
            ws[f'C{row_index}'].fill = red_fill  # "다나와 URL" 셀
        if err_list[2] == 1:
            ws[f'D{row_index}'].fill = red_fill  # "에누리 URL" 셀

        print(f"Row {row_index} 엑셀에 업데이트됨.")

    except Exception as e:
        print(f"Error saving row {row_index} to Excel: {e}")


# 중복 체크 함수
def is_duplicate(entry, naver_entries):
    # 동일한 판매처, 상품명, 구성 개수인 경우 중복으로 판단
    return any(
        naver_entry[3] == entry[3] and  # 판매처
        naver_entry[4] == entry[4] and  # 구성 개수
        naver_entry[5] == entry[5]      # 상품명
        for naver_entry in naver_entries
    )

# 크롤링 결과를 확인하고 에러 리스트에 반영하는 함수
def handle_scraping_result(result_list, index, err_list, merge_list):
    if len(result_list) == 0:
        err_list[index] = 1  # 크롤링 실패 시 해당 인덱스에 에러 표시
    else:
        add_non_duplicates(merge_list, result_list)

# 중복 체크 후 리스트에 추가하는 함수
def add_non_duplicates(merge_list, new_entries):
    for entry in new_entries:
        if not is_duplicate(entry, merge_list):
            merge_list.append(entry)

# 메인 함수
def main(excel_path, limit_count, on_and_off, five_per_mall_name):
    # 엑셀 파일 열기 (openpyxl로 읽기)
    wb = load_workbook(excel_path)
    ws = wb.active

    driver = setup_driver()
    if not driver:
        print("Failed to initialize the web driver.")
        return

    # 엑셀의 각 행을 처리
    for i in range(2, ws.max_row + 1):
        name = ws[f'E{i}'].value
        naver_url = ws[f'B{i}'].value
        danawa_url = ws[f'C{i}'].value
        enuri_url = ws[f'D{i}'].value

        # 에러 리스트 초기화 (각 크롤링 사이트별 에러 체크)
        err_list = [0, 0, 0]

        # 전체 병합 리스트 초기화
        merge_list = []

        # 1. 네이버 크롤링 처리
        naver_result = scrape_naver(driver, name, naver_url)
        handle_scraping_result(naver_result, 0, err_list, merge_list)

        # 2. 다나와 크롤링 처리
        danawa_result = scrape_danawa(driver, name, danawa_url, limit_count, on_and_off)
        handle_scraping_result(danawa_result, 1, err_list, merge_list)

        # 3. 에누리 크롤링 처리
        enuri_result = scrape_enuri(driver, name, enuri_url, limit_count, on_and_off)
        handle_scraping_result(enuri_result, 2, err_list, merge_list)

        # 엑셀로 저장 (각 행별로 저장)
        save_row_to_excel(ws, merge_list, i, err_list, five_per_mall_name)

        # 엑셀 파일을 즉시 저장
        wb.save(excel_path)

    driver.quit()

if __name__ == "__main__":

    # 원하는 값을 여기에 입력하세요.

    excel_path = "프로그램.xlsx"    # 파일 이름 (프로그램이 실행되는 경로에 파일이 있어야 합니다.)
    limit_count = 3                    # 수집 갯수
    on_and_off = 0                     # 1개 수집 : on (수집 변수 1) / off 미수집 변수 0 (기본 미수집 0)
    five_per_mall_name = 'G마켓'        # 5% 할인 적용 판매처 (옥션, G마켓, 11번가...)

    main(excel_path, limit_count, on_and_off, five_per_mall_name)
import requests
from bs4 import BeautifulSoup
import openpyxl
import os
from urllib.request import urlretrieve
from urllib.parse import urljoin

HEADERS = {
    "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    "accept-encoding": "gzip, deflate, br, zstd",
    "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "connection": "keep-alive",
    "cookie": "...",  # ✅ 유효한 쿠키 넣기
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36"
}

def read_it_ids_from_excel(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    return [str(row[0].value) for row in ws.iter_rows(min_row=2) if row[0].value]

def download_image(url, folder, filename):
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, filename)
    try:
        urlretrieve(url, path)
        return path
    except Exception as e:
        print(f"[⚠️ 이미지 실패] {url} → {e}")
        return ""

def parse_item(it_id):
    url = f"https://www.luxurycelebrity2.kr/shop/item.php?it_id={it_id}"
    print(f"🔎 요청 중: {url}")
    res = requests.get(url, headers=HEADERS)
    soup = BeautifulSoup(res.text, "html.parser")
    data = {"it_id": it_id}
    folder = os.path.join("images", it_id)


    # ✅ 메인 이미지 (sit_pvi 안에서 가장 큰 이미지 1개)
    main_img = soup.select_one("#sit_pvi #sit_pvi_big a.popup_item_image img")
    if main_img:
        raw_src = main_img.get("src")
        if raw_src:
            img_url = urljoin(url, raw_src)
            path = download_image(img_url, folder, f"{it_id}_0.jpg")
            data["메인이미지"] = path
        else:
            print(f"[⚠️ 메인 이미지 src 없음] it_id={it_id}")
    else:
        print(f"[⚠️ 메인 이미지 태그 없음] it_id={it_id}")



    # ✅ 상품명
    sit_top = soup.select_one("#sit_ov .sit_top")
    if sit_top:
        h6 = sit_top.select_one("h6")
        h2 = sit_top.select_one("h2")
        if h6 and h2:
            key = h6.text.strip()
            value = h2.get_text(separator=" ", strip=True).replace("요약정보 및 구매", "").strip()
            data[key] = value

    # ✅ 상품 속성들
    for block in soup.select("#sit_ov .sit_text"):
        h5 = block.select_one("h5")
        p = block.select_one("p")
        if h5 and p:
            key = h5.text.strip()
            if p.select("option"):
                options = [opt.text.strip() for opt in p.select("option")]
                data[key] = options
            else:
                data[key] = p.get_text(strip=True)

    # ✅ 상세 이미지
    for idx, img in enumerate(soup.select("#sit_inf_explan img"), start=1):
        src = img.get("src")
        if src:
            full_url = urljoin(url, src)
            download_image(full_url, folder, f"{it_id}_{idx}.jpg")

    return data

def save_to_excel(data_list, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "상품정보"

    all_keys = set()
    for item in data_list:
        all_keys.update(item.keys())
    keys = sorted(all_keys)
    ws.append(keys)

    for item in data_list:
        row = [", ".join(item[k]) if isinstance(item.get(k), list) else item.get(k, "") for k in keys]
        ws.append(row)

    wb.save(filename)

# ✅ 실행 부분
if __name__ == "__main__":
    it_ids = read_it_ids_from_excel("it_ids.xlsx")
    total = len(it_ids)
    results = []

    for i, it_id in enumerate(it_ids, start=1):
        print(f"\n📦 [{i}/{total}] 처리 중: {it_id}")
        try:
            obj = parse_item(it_id)
            results.append(obj)
            print(f"✔ 완료: {it_id} ({len(results)}개 완료)")
        except Exception as e:
            print(f"❌ 에러: {it_id} → {e}")

    save_to_excel(results, "상품정보.xlsx")
    print("\n✅ 전체 완료! 엑셀 저장됨 → 상품정보.xlsx")

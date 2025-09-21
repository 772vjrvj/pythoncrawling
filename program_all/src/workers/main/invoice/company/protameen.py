# -*- coding: utf-8 -*-
import os, re
from datetime import datetime
from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Side

import pdfplumber, pytesseract, fitz
from PIL import Image
from io import BytesIO


class ProtameenInvoiceParser:

    def __init__(self, folder="수입 선적서류/Protameen", log_func=print):
        self.folder, self.log_func = folder, log_func

        # --- Tesseract 경로 고정(설치·배포 모두 커버) ---
        _tess = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_tess):
            pytesseract.pytesseract.tesseract_cmd = _tess
        else:
            _vend = os.path.join(os.getcwd(), "vendor", "bin", "tesseract", "tesseract.exe")
            if os.path.isfile(_vend):
                pytesseract.pytesseract.tesseract_cmd = _vend




    # region 날짜 변환2 : 06/26/2025 → YYYY-MM-DD
    def _to_iso_date_numeric(self, s: str) -> str:
        if not s:
            return ""
        m = re.search(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b', s.strip())
        if not m:
            return ""
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000 if yy < 70 else 1900
        try:
            return datetime(yy, mm, dd).strftime('%Y-%m-%d')
        except ValueError:
            return ""

    def _to_iso_date_any(self, s: str) -> str:
        return self._to_iso_date_named(s) or self._to_iso_date_numeric(s)
    # endregion


    # region pdf안에 전체 text읽기
    def read_text(self, path):
        result = None
        r = PdfReader(path)
        for i, p in enumerate(getattr(r, "pages", []), 1):
            t = p.extract_text() or ""
            # 첫 줄 체크: Invoice VKR이 있는 페이지만 사용
            if "INVOICE DATE" in (t.upper() if t else ""):
                result =  t
        if result:
            return result
        else:
            doc = fitz.open(path)
            texts = []
            for i in range(len(doc)):
                page = doc.load_page(i)
                pix = page.get_pixmap(dpi=300)
                img = Image.open(BytesIO(pix.tobytes("png")))
                try:
                    txt = pytesseract.image_to_string(img, lang="eng+kor") or ""
                except Exception:
                    txt = pytesseract.image_to_string(img, lang="eng") or ""
                texts.append(txt)

            return "\n".join(texts)

    # endregion


    # region 전용 파서: "Invoice details" 블록
    def _parse_invoice_details_block(self, page_text: str, res: dict) -> dict:
        if not page_text or not re.search(r'(?i)\bInvoice\s+details\b', page_text):
            return res

        m_no = re.search(r'(?i)\bInvoice\s*no\.?\s*[:#-]?\s*([A-Za-z0-9\-_/ ]+)', page_text)
        if m_no:
            res.setdefault("Invoice no.", m_no.group(1).strip())
            if not res["Invoice no."]:
                res["Invoice no."] = m_no.group(1).strip()

        m_dt_num = re.search(r'(?i)\bInvoice\s*date\b[:#-]?\s*([0-9]{1,2}[/-][0-9]{1,2}[/-][0-9]{2,4})', page_text)
        if m_dt_num:
            res_date = self._to_iso_date_numeric(m_dt_num.group(1))
            if res_date:
                res["Invoice date"] = res_date

        if res.get("Amount") is None:
            totals = []
            for m in re.finditer(r'(?is)(?<!sub)\btotal\b[^0-9$]{0,16}\$?\s*([\d,]+(?:\.\d{2})?)', page_text, flags=re.I):
                val = m.group(1)
                if val:
                    try:
                        totals.append(float(val.replace(",", "")))
                    except ValueError:
                        pass
            if totals:
                res["Amount"] = totals[-1]
        return res
    # endregion

    # region 유틸: 금액 문자열 -> float
    def _money_to_float(self, s: str):
        if not s:
            return None
        # "$1,016.80" / "1,016.80" → 1016.80
        m = re.search(r'[\$€]?\s*([\d,]+(?:\.\d{2})?)', s)
        if not m:
            return None
        try:
            return float(m.group(1).replace(',', ''))
        except ValueError:
            return None
    # endregion


    # region field 추출 (Protameen 전용)
    def parse_fields(self, page_text: str):
        """
        Protameen 문서 전용 파싱:
        - Invoice no. : "HANACARE" 라인 기준 위로 3번째 비어있지 않은 라인
        - Invoice date: "INVOICE DATE" 바로 위 줄 (m/d/yyyy → yyyy-mm-dd)
        - Amount      : "PLEASE REMIT TO: TOTAL:" 앞에 붙은 금액(같은 줄/이전 줄 모두 지원)
        """
        res = {"Invoice date": None, "Invoice no.": None, "Amount": None}
        if not page_text:
            return res

        # 줄 단위 전처리
        raw_lines = page_text.splitlines()
        lines = [ln.strip() for ln in raw_lines]

        # -------------------------
        # 1) Invoice date (INVOICE DATE 바로 위 줄)
        # -------------------------
        inv_date = None
        for idx, ln in enumerate(lines):
            if ln.strip().upper() == "INVOICE DATE":
                # 위쪽에서 가장 가까운 non-empty 라인 찾기
                j = idx - 1
                while j >= 0 and (not lines[j].strip()):
                    j -= 1
                if j >= 0:
                    # 예: 9/14/2025
                    cand = lines[j].strip()
                    iso = self._to_iso_date_numeric(cand)
                    if iso:
                        inv_date = iso
                break
        res["Invoice date"] = inv_date

        # -------------------------
        # 2) Invoice no. (HANACARE 기준 위로 3번째 non-empty 라인)
        # -------------------------
        inv_no = None
        idx_hana = -1
        for idx, ln in enumerate(lines):
            if "Page" in ln:
                idx_hana = idx
                break
        if idx_hana >= 0:
            found = []
            j = idx_hana - 1
            while j >= 0 and len(found) < 2:
                if lines[j].strip():
                    found.append(lines[j].strip())
                j -= 1
            if len(found) >= 2:
                cand = found[1]  # 두 번째
                # 숫자/영문 조합 허용 (예: 143556, C03551 등)
                m_no = re.search(r'([A-Za-z0-9\-_/]+)', cand)
                if m_no:
                    inv_no = m_no.group(1)
        res["Invoice no."] = inv_no

        # -------------------------
        # 3) Amount ("PLEASE REMIT TO: TOTAL:" 앞에 있는 금액)
        #    - 같은 줄에 붙어있는 "$1,016.80PLEASE REMIT TO: TOTAL:" 패턴 우선 처리
        #    - 없으면 해당 라인의 직전 non-empty 라인에서 금액 추출
        # -------------------------
        amount_val = None

        # 3-1) 같은 줄 케이스 먼저 (멀티라인 포함 검색)
        m_inline = re.search(
            r'(?is)([\$€]?\s*[\d,]+(?:\.\d{2})?)\s*PLEASE\s*REMIT\s*TO:\s*TOTAL\s*:',
            page_text,
            flags=re.I
        )
        if m_inline:
            amount_val = self._money_to_float(m_inline.group(1))

        # 3-2) 라인 분리 케이스: "PLEASE REMIT TO: TOTAL:" 라인의 바로 위 non-empty 라인
        if amount_val is None:
            idx_total = -1
            for idx, ln in enumerate(lines):
                if re.search(r'(?i)PLEASE\s*REMIT\s*TO:\s*TOTAL\s*:', ln):
                    idx_total = idx
                    break
            if idx_total >= 0:
                j = idx_total - 1
                while j >= 0 and (not lines[j].strip()):
                    j -= 1
                if j >= 0:
                    amount_val = self._money_to_float(lines[j].strip())

        res["Amount"] = amount_val

        # 유틸: 대문자 기준 사이 텍스트 추출
        def _between_upper(s: str, left: str, right: str) -> str:
            if not s:
                return ""
            u = s.upper()
            L, R = u.find(left.upper()), u.find(right.upper())
            if L == -1 or R == -1 or R <= L:
                return ""
            return s[L + len(left):R].strip()

        # 예외 1) Invoice no. : 첫 줄이 "C035511427827/11/2025Page" 형태인 경우
        # 요구사항: 맨 첫 줄에서 1-based 6~11 글자(= 파이썬 0-based slice [5:11])를 Invoice no.로
        if not res.get("Invoice no."):
            # 첫 non-empty 라인
            first_nonempty = next((ln for ln in lines if ln.strip()), "")
            if first_nonempty and len(first_nonempty) >= 11:
                invno_candidate = first_nonempty[5:11]  # 1-based 6~11 → 0-based [5:11]
                # 숫자 6자리이면 그대로 사용 (예: 142782)
                if re.fullmatch(r'\d{6}', invno_candidate):
                    res["Invoice no."] = invno_candidate

        # 예외 2) Invoice date : "USA6/11/2025INVOICE DATEPRO" 형태인 경우
        # 요구사항: "USA"와 "INVOICE" 사이를 잘라서 날짜(m/d/yyyy) → yyyy-mm-dd
        if not res.get("Invoice date"):
            for ln in lines:
                if "USA" in ln.upper() and "INVOICE" in ln.upper():
                    mid = _between_upper(ln, "USA", "INVOICE")
                    # 공백 제거된 중간문자열에서 날짜 패턴 추출
                    m = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', mid.replace(" ", ""))
                    if m:
                        iso = self._to_iso_date_numeric(m.group(1))
                        if iso:
                            res["Invoice date"] = iso
                            break


        return res
    # endregion


    # region 엑셀 동기화
    def sync_excel(self, objs: list):
        """
        실행경로/수입 선적서류/해외 공급사 Payment 관리대장.xlsx
        시트명: 폴더명(BioRenuva)
        A4:H4 = ["Order no.","OC","Invoice no.","Amount","Invoice date","Due date","Payment","Note"]
        - 시트에는 있는데 폴더엔 없는 Order no. → A열 셀에 메모 추가
        - 시트에 없는 신규 Order no. → 다음 행에 추가
          * Amount: 회계 서식
          * Invoice date: 사용자 지정 mm"월" dd"일"
        """
        # 경로/시트 결정
        xlsx_path = os.path.join(os.getcwd(), "수입 선적서류", "해외 공급사 Payment 관리대장.xlsx")
        sheet_name = os.path.basename(self.folder).strip() or "Sheet1" # 수입 선적서류/BioRenuva -> BioRenuva

        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # 헤더 위치 및 컬럼 인덱스
        header_row = 4
        headers = ["Order no.", "OC", "Invoice no.", "Amount", "Invoice date", "Due date", "Payment", "Note"]
        header_to_col = {h: idx+1 for idx, h in enumerate(headers)}  # A=1 ... H=8

        # 시트에 있는 기존 Order no. 수집 (A열, 5행부터)
        existing_map = {}  # order_no -> row
        max_row = ws.max_row if ws.max_row >= header_row else header_row
        for r in range(header_row + 1, max_row + 1):
            v = ws.cell(row=r, column=header_to_col["Order no."]).value
            if v is not None and str(v).strip() != "":
                existing_map[str(v).strip()] = r

        # 폴더에서 파싱한 Order no. 목록
        parsed_map = {str(o.get("Order no.", "")).strip(): o for o in objs if str(o.get("Order no.", "")).strip()}

        # 1) 정합성 갱신: 폴더에도 있고 시트에도 있는 경우 → 기존 메모 제거
        for order_no, row in existing_map.items():
            cell = ws.cell(row=row, column=header_to_col["Order no."])
            if order_no in parsed_map:
                if cell.comment is not None:
                    cell.comment = None  # 메모 제거
                # 배경색 원복 (기본값) ===
                cell.fill = PatternFill(fill_type=None)


        # 2) 시트에는 있으나 폴더엔 없는 항목 → 메모 추가 (삭제 표시)  # === 기존+보완 ===
        for order_no, row in existing_map.items():
            if order_no not in parsed_map:
                cell = ws.cell(row=row, column=header_to_col["Order no."])
                cell.comment = Comment("해당 서류가 폴더에서 삭제되었습니다.", "메모")
                # 배경색 빨강 ===
                cell.fill = PatternFill(start_color="FFD4D5", end_color="FFD4D5", fill_type="solid")


        # 2) 신규 추가 (시트에 없는 폴더 항목)
        append_row = max_row + 1
        # 얇은 테두리/중앙정렬 공통 스타일  === 신규 ===
        thin = Side(style='thin')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(horizontal='center', vertical='center')

        for order_no, obj in parsed_map.items():
            if order_no in existing_map:
                continue  # 이미 존재

            ws.cell(row=append_row, column=header_to_col["Order no."]).value = order_no
            ws.cell(row=append_row, column=header_to_col["OC"]).value = obj.get("OC")
            ws.cell(row=append_row, column=header_to_col["Invoice no."]).value = obj.get("Invoice no.")

            # Amount: 회계(USD, 소수 2자리)  === 변경 ===
            c_amount = ws.cell(row=append_row, column=header_to_col["Amount"])
            amount = obj.get("Amount")
            if isinstance(amount, (int, float)):
                c_amount.value = float(amount)
                # 회계 서식 + USD 심볼
                c_amount.number_format = '_-[$USD]* #,##0.00_-;[Red]-[$USD]* #,##0.00_-;_-[$USD]* "-"??_-;_-@_-'

            # Invoice date: 사용자 지정 mm"월" dd"일"  (값은 date로)
            c_date = ws.cell(row=append_row, column=header_to_col["Invoice date"])
            inv_date_iso = obj.get("Invoice date")
            if inv_date_iso:
                try:
                    dt = datetime.strptime(inv_date_iso, "%Y-%m-%d").date()
                    c_date.value = dt
                    c_date.number_format = 'mm"월" dd"일"'
                except Exception:
                    c_date.value = inv_date_iso  # 파싱 실패 시 문자열로 기록
            else:
                c_date.number_format = 'mm"월" dd"일"'  # === 신규 === 값 없어도 서식 적용

            # === 신규: Due date, Payment 는 값 없이 서식만
            c_due = ws.cell(row=append_row, column=header_to_col["Due date"])
            c_due.value = None
            c_due.number_format = 'mm"월" dd"일"'

            c_pay = ws.cell(row=append_row, column=header_to_col["Payment"])
            c_pay.value = None
            c_pay.number_format = 'mm"월" dd"일"'


            # === 신규: 해당 행 A~H 정렬/테두리 적용
            for col in range(1, len(header_to_col) + 1):  # 1..8
                cell = ws.cell(row=append_row, column=col)
                cell.alignment = align_center
                cell.border = border_all

            append_row += 1
        wb.save(xlsx_path)
        self.log_func(f"엑셀 동기화 완료: {xlsx_path} / 시트: {sheet_name}")
    # endregion


    # region run
    def run(self):
        # .pdf 로 끝나는 파일들을 가져옴.
        files = [f for f in sorted(os.listdir(self.folder)) if f.lower().endswith(".pdf")]
        total = len(files)

        objs = []  # === 신규 === 결과 객체 배열

        for idx, fname in enumerate(files, 1):
            self.log_func(f"[{idx}/{total}] : {fname}")
            fpath = os.path.join(self.folder, fname)
            txt = self.read_text(fpath)
            fields = self.parse_fields(txt)

            # === 신규 === 결과 객체 생성 (Order no. = fname)
            order_no = os.path.splitext(fname)[0]  # === 신규: .pdf 제거

            obj = {
                "Order no.": order_no,
                "OC": None,
                "Invoice no.": fields.get("Invoice no."),
                "Amount": fields.get("Amount"),
                "Invoice date": fields.get("Invoice date"),
                "Due date": None,
                "Payment": None,
                "Note": None,
            }
            objs.append(obj)

            self.log_func(
                f"[RESULT] {fname} -> "
                f"Invoice date={fields.get('Invoice date')}, "
                f"Invoice no.={fields.get('Invoice no.')}, "
                f"Amount={fields.get('Amount')}"
            )

        # === 신규 === 엑셀 반영
        if objs:
            self.sync_excel(objs)
    # endregion
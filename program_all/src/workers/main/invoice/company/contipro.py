# -*- coding: utf-8 -*-
"""
CONTIPRO 인보이스 파서 (3필드 고정)
- Invoice No. / Date of issue(YYYY-MM-DD) / Total price(EUR → 숫자)
- 기본 폴더: 수입 선적서류/contipro
- 동작: run()은 내부 로그 출력만 수행(리턴 없음)
"""
from datetime import datetime
import os, re
import pdfplumber, pytesseract, fitz
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Side, PatternFill



class ContiproInvoiceParser:

    def __init__(self, folder="수입 선적서류/CONTIPRO", log_func=print):
        self.folder, self.log_func = folder, log_func

        # --- Tesseract 경로 고정(설치·배포 모두 커버) ---
        _tess = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_tess):
            pytesseract.pytesseract.tesseract_cmd = _tess
        else:
            _vend = os.path.join(os.getcwd(), "vendor", "bin", "tesseract", "tesseract.exe")
            if os.path.isfile(_vend):
                pytesseract.pytesseract.tesseract_cmd = _vend


    def parse_fields(self, page_text: str) -> dict:
        """
        TAX CERTIFICATE 문서 전용 파서
        - Invoice no.: 'INVOICE - TAX CERTIFICATE <번호>'
        - Amount: 'Total price EUR <금액>'
        - Invoice date: 'Date of issue <dd.mm.yyyy>'
        """
        res = {"Invoice no.": None, "Invoice date": None, "Amount": None}
        if not page_text:
            return res

        lines = [ln.strip() for ln in page_text.splitlines() if ln.strip()]

        for ln in lines:
            # 1) Invoice no.
            if "INVOICE - TAX CERTIFICATE" in ln.upper():
                m = re.search(r'INVOICE\s*-\s*TAX\s*CERTIFICATE\s+(\d+)', ln, re.I)
                if m:
                    res["Invoice no."] = m.group(1)

            # 2) Amount
            # "Total price EUR 10 025,00" 라인에서 금액만 잡기
            if re.search(r'(?i)Total\s+price\s+EUR', ln):
                # EUR 뒤에 최소 한 자리 숫자부터 캡처하도록 수정
                m_amt = re.search(r'(?i)Total\s+price\s+EUR\s+([0-9][\d\s.,]*)', ln)
                if m_amt:
                    raw_amt = m_amt.group(1)
                    # 비표준 공백(예: NBSP) 정리
                    raw_amt = raw_amt.replace('\u00A0', ' ').strip()
                    # 유럽식 "10 025,00" → "10025.00"
                    cleaned = raw_amt.replace(' ', '').replace('.', '').replace(',', '.')
                    try:
                        res["Amount"] = float(cleaned)
                    except ValueError:
                        res["Amount"] = None
            # 3) Date of issue
            if re.search(r'(?i)Date\s+of\s+issue', ln):
                m_dt = re.search(r'Date\s+of\s+issue\s+(\d{2}\.\d{2}\.\d{4})', ln, re.I)
                if m_dt:
                    raw_date = m_dt.group(1)
                    try:
                        dt = datetime.strptime(raw_date, "%d.%m.%Y")
                        res["Invoice date"] = dt.strftime("%Y-%m-%d")
                    except ValueError:
                        res["Invoice date"] = None
        return res


    def read_text(self, path):
        try:
            with pdfplumber.open(path) as pdf:
                t = "\n".join((p.extract_text() or "") for p in pdf.pages)
                if len((t or "").strip()) > 40:
                    return t
        except Exception:
            pass

        doc = fitz.open(path)
        texts = []
        for i in range(len(doc)):
            page = doc.load_page(i)
            pix = page.get_pixmap(dpi=300)
            img = Image.open(BytesIO(pix.tobytes("png")))
            try:
                txt = pytesseract.image_to_string(img, lang="eng+kor") or ""
            except Exception:
                txt = pytesseract.image_to_string(img, lang="eng") or ""
            texts.append(txt)

        return "\n".join(texts)


    # region 엑셀 동기화
    def sync_excel(self, objs: list):
        """
        실행경로/수입 선적서류/해외 공급사 Payment 관리대장.xlsx
        시트명: 폴더명(BioRenuva)
        A4:H4 = ["Order no.","OC","Invoice no.","Amount","Invoice date","Due date","Payment","Note"]
        - 시트에는 있는데 폴더엔 없는 Order no. → A열 셀에 메모 추가
        - 시트에 없는 신규 Order no. → 다음 행에 추가
          * Amount: 회계 서식
          * Invoice date: 사용자 지정 mm"월" dd"일"
        """
        # 경로/시트 결정
        xlsx_path = os.path.join(os.getcwd(), "수입 선적서류", "해외 공급사 Payment 관리대장.xlsx")
        sheet_name = os.path.basename(self.folder).strip() or "Sheet1" # 수입 선적서류/BioRenuva -> BioRenuva

        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # 헤더 위치 및 컬럼 인덱스
        header_row = 4
        headers = ["Order no.", "OC", "Invoice no.", "Amount", "Invoice date", "Due date", "Payment", "Note"]
        header_to_col = {h: idx+1 for idx, h in enumerate(headers)}  # A=1 ... H=8

        # 시트에 있는 기존 Order no. 수집 (A열, 5행부터)
        existing_map = {}  # order_no -> row
        max_row = ws.max_row if ws.max_row >= header_row else header_row
        for r in range(header_row + 1, max_row + 1):
            v = ws.cell(row=r, column=header_to_col["Order no."]).value
            if v is not None and str(v).strip() != "":
                existing_map[str(v).strip()] = r

        # 폴더에서 파싱한 Order no. 목록
        parsed_map = {str(o.get("Order no.", "")).strip(): o for o in objs if str(o.get("Order no.", "")).strip()}

        # 1) 정합성 갱신: 폴더에도 있고 시트에도 있는 경우 → 기존 메모 제거
        for order_no, row in existing_map.items():
            cell = ws.cell(row=row, column=header_to_col["Order no."])
            if order_no in parsed_map:
                if cell.comment is not None:
                    cell.comment = None  # 메모 제거
                # 배경색 원복 (기본값) ===
                cell.fill = PatternFill(fill_type=None)

        # 2) 시트에는 있으나 폴더엔 없는 항목 → 메모 추가 (삭제 표시)  # === 기존+보완 ===
        for order_no, row in existing_map.items():
            if order_no not in parsed_map:
                cell = ws.cell(row=row, column=header_to_col["Order no."])
                cell.comment = Comment("해당 서류가 폴더에서 삭제되었습니다.", "메모")
                # 배경색 빨강 ===
                cell.fill = PatternFill(start_color="FFD4D5", end_color="FFD4D5", fill_type="solid")


        # 2) 신규 추가 (시트에 없는 폴더 항목)
        append_row = max_row + 1
        # 얇은 테두리/중앙정렬 공통 스타일  === 신규 ===
        thin = Side(style='thin')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(horizontal='center', vertical='center')

        for order_no, obj in parsed_map.items():
            if order_no in existing_map:
                continue  # 이미 존재

            ws.cell(row=append_row, column=header_to_col["Order no."]).value = order_no
            ws.cell(row=append_row, column=header_to_col["OC"]).value = obj.get("OC")
            ws.cell(row=append_row, column=header_to_col["Invoice no."]).value = obj.get("Invoice no.")

            # Amount: 회계(USD, 소수 2자리)  === 변경 ===
            c_amount = ws.cell(row=append_row, column=header_to_col["Amount"])
            amount = obj.get("Amount")
            if isinstance(amount, (int, float)):
                c_amount.value = float(amount)
                # 회계 서식 + USD 심볼
                c_amount.number_format = '_-[$EUR]* #,##0.00_-;[Red]-[$EUR]* #,##0.00_-;_-[$EUR]* "-"??_-;_-@_-'

            # Invoice date: 사용자 지정 mm"월" dd"일"  (값은 date로)
            c_date = ws.cell(row=append_row, column=header_to_col["Invoice date"])
            inv_date_iso = obj.get("Invoice date")
            if inv_date_iso:
                try:
                    dt = datetime.strptime(inv_date_iso, "%Y-%m-%d").date()
                    c_date.value = dt
                    c_date.number_format = 'mm"월" dd"일"'
                except Exception:
                    c_date.value = inv_date_iso  # 파싱 실패 시 문자열로 기록
            else:
                c_date.number_format = 'mm"월" dd"일"'  # === 신규 === 값 없어도 서식 적용

            # === 신규: Due date, Payment 는 값 없이 서식만
            c_due = ws.cell(row=append_row, column=header_to_col["Due date"])
            c_due.value = None
            c_due.number_format = 'mm"월" dd"일"'

            c_pay = ws.cell(row=append_row, column=header_to_col["Payment"])
            c_pay.value = None
            c_pay.number_format = 'mm"월" dd"일"'


            # === 신규: 해당 행 A~H 정렬/테두리 적용
            for col in range(1, len(header_to_col) + 1):  # 1..8
                cell = ws.cell(row=append_row, column=col)
                cell.alignment = align_center
                cell.border = border_all

            append_row += 1
        wb.save(xlsx_path)
        self.log_func(f"엑셀 동기화 완료: {xlsx_path} / 시트: {sheet_name}")
    # endregion


    def run(self):
        files = [f for f in sorted(os.listdir(self.folder)) if f.lower().endswith(".pdf")]
        total = len(files)
        objs = []
        for idx, fname in enumerate(files, 1):
            self.log_func(f"[{idx}/{total}] : {fname}")
            fpath = os.path.join(self.folder, fname)
            txt = self.read_text(fpath)
            fields = self.parse_fields(txt)
            order_no = os.path.splitext(fname)[0]
            obj = {
                "Order no.": order_no,
                "OC": None,
                "Invoice no.": fields.get("Invoice no."),
                "Amount": fields.get("Amount"),
                "Invoice date": fields.get("Invoice date"),
                "Due date": None,
                "Payment": None,
                "Note": None,
            }
            objs.append(obj)
            self.log_func(
                f"[RESULT] {fname} -> "
                f"Invoice date={fields.get('Invoice date')}, "
                f"Invoice no.={fields.get('Invoice no.')}, "
                f"Amount={fields.get('Amount')}"
            )
        if objs:
            self.sync_excel(objs)
    # endregion
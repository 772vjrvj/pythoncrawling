# -*- coding: utf-8 -*-
import os, re
from datetime import datetime
from pypdf import PdfReader
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Side

import pdfplumber, pytesseract, fitz
from PIL import Image
from io import BytesIO

class LignopureInvoiceParser:

    def __init__(self, folder="수입 선적서류/Lignopure", log_func=print):
        self.folder, self.log_func = folder, log_func

        # --- Tesseract 경로 고정(설치·배포 모두 커버) ---
        _tess = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.isfile(_tess):
            pytesseract.pytesseract.tesseract_cmd = _tess
        else:
            _vend = os.path.join(os.getcwd(), "vendor", "bin", "tesseract", "tesseract.exe")
            if os.path.isfile(_vend):
                pytesseract.pytesseract.tesseract_cmd = _vend




    # region 날짜 변환2 : 06/26/2025 → YYYY-MM-DD
    def _to_iso_date_numeric(self, s: str) -> str:
        if not s:
            return ""
        m = re.search(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b', s.strip())
        if not m:
            return ""
        mm, dd, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if yy < 100:
            yy += 2000 if yy < 70 else 1900
        try:
            return datetime(yy, mm, dd).strftime('%Y-%m-%d')
        except ValueError:
            return ""

    def _to_iso_date_any(self, s: str) -> str:
        return self._to_iso_date_named(s) or self._to_iso_date_numeric(s)
    # endregion


    # region pdf안에 전체 text읽기
    def read_text(self, path):
        result = None
        r = PdfReader(path)
        for i, p in enumerate(getattr(r, "pages", []), 1):
            t = p.extract_text() or ""
            if not t.strip():
                continue
            # 모든 라인 확인
            for ln in t.splitlines():
                if "INVOICE HANA" in ln.upper():
                    result = t
        if result:
            return result
        else:
            doc = fitz.open(path)
            texts = []
            for i in range(len(doc)):
                page = doc.load_page(i)
                pix = page.get_pixmap(dpi=300)
                img = Image.open(BytesIO(pix.tobytes("png")))
                try:
                    txt = pytesseract.image_to_string(img, lang="eng+kor") or ""
                except Exception:
                    txt = pytesseract.image_to_string(img, lang="eng") or ""
                texts.append(txt)
            return "\n".join(texts)
    # endregion


    # region 유틸: 금액 문자열 -> float
    # === 신규: 유럽/미국 혼합 금액 파서 (€, 천단위/소수점 기호 자동 판별)
    def _money_to_float_any(self, s: str):
        if not s:
            return None
        s = s.strip()
        # 통화기호, 공백 제거
        s = re.sub(r'[^\d,.\-]', '', s)  # 숫자/쉼표/점/마이너스만 남김

        # 패턴1: 유럽식 "1.234,56" / "600,00"
        if re.fullmatch(r'\d{1,3}(?:\.\d{3})*,\d{2}', s) or re.fullmatch(r'\d+,\d{2}', s):
            s2 = s.replace('.', '').replace(',', '.')
            try:
                return float(s2)
            except ValueError:
                return None

        # 패턴2: 미국식 "1,234.56" / "600.00"
        if re.fullmatch(r'\d{1,3}(?:,\d{3})*\.\d{2}', s) or re.fullmatch(r'\d+\.\d{2}', s):
            try:
                return float(s.replace(',', ''))
            except ValueError:
                return None

        # 패턴3: 소수점 없는 정수(천단위 구분 가능)
        if re.fullmatch(r'\d{1,3}(?:[.,]\d{3})+', s) or re.fullmatch(r'\d+', s):
            try:
                return float(s.replace(',', '').replace('.', ''))
            except ValueError:
                return None

        # 최후: 쉼표를 점으로 바꾸고 시도
        try:
            return float(s.replace(',', '.'))
        except ValueError:
            return None
    # endregion


    # region field 추출 (Protameen 전용)
    def parse_fields(self, page_text: str) -> dict:
        """
        HANA 문서 전용 필드 추출
        - Invoice no.: "Invoice HANA..." 라인에서 HANA로 시작하는 토큰 (예: HANA20250401)
        - Invoice date: "Hamburg, dd.mm.yyyy" → yyyy-mm-dd
        - Amount: "TOTAL (inkl. MwSt.)  <금액> €" → float (유럽/미국 표기 모두 지원)
        """
        res = {"Invoice no.": None, "Invoice date": None, "Amount": None}
        if not page_text:
            return res

        lines = [ln.strip() for ln in page_text.splitlines() if ln.strip()]

        # 1) Invoice no. : "Invoice HANA..." 라인에서 HANA로 시작하는 연속 토큰 추출
        #    예: "Invoice HANA20250401" → HANA20250401
        inv_no = None
        for ln in lines:
            if re.search(r'(?i)\bInvoice\s+HANA', ln):
                m = re.search(r'\b(HANA[A-Za-z0-9\-_]+)\b', ln, re.I)
                if m:
                    inv_no = m.group(1).strip()
                    break
        res["Invoice no."] = inv_no

        # 2) Invoice date : "Hamburg, dd.mm.yyyy" → yyyy-mm-dd
        inv_date = None
        # 우선 "Hamburg," 바로 뒤 날짜 우선
        m_city_date = re.search(r'(?i)Hamburg,\s*(\d{1,2}\.\d{1,2}\.\d{4})', page_text, re.I)
        cand_date = None
        if m_city_date:
            cand_date = m_city_date.group(1).strip()
        else:
            # 백업: 본문 내 dd.mm.yyyy 아무거나 (가장 먼저 나오는 것)
            m_any_date = re.search(r'\b(\d{1,2}\.\d{1,2}\.\d{4})\b', page_text)
            if m_any_date:
                cand_date = m_any_date.group(1).strip()

        if cand_date:
            # dd.mm.yyyy -> yyyy-mm-dd
            m = re.match(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', cand_date)
            if m:
                dd, mm, yy = map(int, m.groups())
                try:
                    inv_date = datetime(yy, mm, dd).strftime('%Y-%m-%d')
                except ValueError:
                    inv_date = None
        res["Invoice date"] = inv_date

        # 3) Amount : "TOTAL (inkl. MwSt.)  600.00 €" 또는 "TOTAL (inkl. MwSt.)  600,00 €"
        amt = None
        # 한 줄에 함께 있는 경우 우선
        m_amt_inline = re.search(
            r'(?i)TOTAL\s*\(inkl\.\s*MwSt\.\)\s*([€$]?\s*[\d\.,]+)\s*€',
            page_text
        )
        if m_amt_inline:
            amt = self._money_to_float_any(m_amt_inline.group(1))

        # 라인 분리 케이스 백업 처리: "TOTAL (inkl. MwSt.)" 라인 다음/같은 줄의 금액+€
        if amt is None:
            for i, ln in enumerate(lines):
                if re.search(r'(?i)^TOTAL\s*\(inkl\.\s*MwSt\.\)\b', ln):
                    # 같은 줄에서 금액 찾기
                    m_same = re.search(r'([€$]?\s*[\d\.,]+)\s*€', ln)
                    if m_same:
                        amt = self._money_to_float_any(m_same.group(1))
                        break
                    # 다음 줄에서 금액 찾기
                    if i + 1 < len(lines):
                        m_next = re.search(r'([€$]?\s*[\d\.,]+)\s*€', lines[i + 1])
                        if m_next:
                            amt = self._money_to_float_any(m_next.group(1))
                            break

        res["Amount"] = amt
        return res
    # endregion


    # region 엑셀 동기화
    def sync_excel(self, objs: list):
        """
        실행경로/수입 선적서류/해외 공급사 Payment 관리대장.xlsx
        시트명: 폴더명(BioRenuva)
        A4:H4 = ["Order no.","OC","Invoice no.","Amount","Invoice date","Due date","Payment","Note"]
        - 시트에는 있는데 폴더엔 없는 Order no. → A열 셀에 메모 추가
        - 시트에 없는 신규 Order no. → 다음 행에 추가
          * Amount: 회계 서식
          * Invoice date: 사용자 지정 mm"월" dd"일"
        """
        # 경로/시트 결정
        xlsx_path = os.path.join(os.getcwd(), "수입 선적서류", "해외 공급사 Payment 관리대장.xlsx")
        sheet_name = os.path.basename(self.folder).strip() or "Sheet1" # 수입 선적서류/BioRenuva -> BioRenuva

        wb = load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

        # 헤더 위치 및 컬럼 인덱스
        header_row = 4
        headers = ["Order no.", "OC", "Invoice no.", "Amount", "Invoice date", "Due date", "Payment", "Note"]
        header_to_col = {h: idx+1 for idx, h in enumerate(headers)}  # A=1 ... H=8

        # 시트에 있는 기존 Order no. 수집 (A열, 5행부터)
        existing_map = {}  # order_no -> row
        max_row = ws.max_row if ws.max_row >= header_row else header_row
        for r in range(header_row + 1, max_row + 1):
            v = ws.cell(row=r, column=header_to_col["Order no."]).value
            if v is not None and str(v).strip() != "":
                existing_map[str(v).strip()] = r

        # 폴더에서 파싱한 Order no. 목록
        parsed_map = {str(o.get("Order no.", "")).strip(): o for o in objs if str(o.get("Order no.", "")).strip()}

        # 1) 정합성 갱신: 폴더에도 있고 시트에도 있는 경우 → 기존 메모 제거
        for order_no, row in existing_map.items():
            cell = ws.cell(row=row, column=header_to_col["Order no."])
            if order_no in parsed_map:
                if cell.comment is not None:
                    cell.comment = None  # 메모 제거
                # 배경색 원복 (기본값) ===
                cell.fill = PatternFill(fill_type=None)

        # 2) 시트에는 있으나 폴더엔 없는 항목 → 메모 추가 (삭제 표시)  # === 기존+보완 ===
        for order_no, row in existing_map.items():
            if order_no not in parsed_map:
                cell = ws.cell(row=row, column=header_to_col["Order no."])
                cell.comment = Comment("해당 서류가 폴더에서 삭제되었습니다.", "메모")
                # 배경색 빨강 ===
                cell.fill = PatternFill(start_color="FFD4D5", end_color="FFD4D5", fill_type="solid")



        # 2) 신규 추가 (시트에 없는 폴더 항목)
        append_row = max_row + 1
        # 얇은 테두리/중앙정렬 공통 스타일  === 신규 ===
        thin = Side(style='thin')
        border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
        align_center = Alignment(horizontal='center', vertical='center')

        for order_no, obj in parsed_map.items():
            if order_no in existing_map:
                continue  # 이미 존재

            ws.cell(row=append_row, column=header_to_col["Order no."]).value = order_no
            ws.cell(row=append_row, column=header_to_col["OC"]).value = obj.get("OC")
            ws.cell(row=append_row, column=header_to_col["Invoice no."]).value = obj.get("Invoice no.")

            # Amount: 회계(USD, 소수 2자리)  === 변경 ===
            c_amount = ws.cell(row=append_row, column=header_to_col["Amount"])
            amount = obj.get("Amount")
            if isinstance(amount, (int, float)):
                c_amount.value = float(amount)
                # 회계 서식 + USD 심볼
                c_amount.number_format = '_-[$USD]* #,##0.00_-;[Red]-[$USD]* #,##0.00_-;_-[$USD]* "-"??_-;_-@_-'

            # Invoice date: 사용자 지정 mm"월" dd"일"  (값은 date로)
            c_date = ws.cell(row=append_row, column=header_to_col["Invoice date"])
            inv_date_iso = obj.get("Invoice date")
            if inv_date_iso:
                try:
                    dt = datetime.strptime(inv_date_iso, "%Y-%m-%d").date()
                    c_date.value = dt
                    c_date.number_format = 'mm"월" dd"일"'
                except Exception:
                    c_date.value = inv_date_iso  # 파싱 실패 시 문자열로 기록
            else:
                c_date.number_format = 'mm"월" dd"일"'  # === 신규 === 값 없어도 서식 적용

            # === 신규: Due date, Payment 는 값 없이 서식만
            c_due = ws.cell(row=append_row, column=header_to_col["Due date"])
            # row 번호에 맞춰 E열 참조
            c_due.value = f"=E{append_row}+60"
            c_due.number_format = 'mm"월" dd"일"'

            c_pay = ws.cell(row=append_row, column=header_to_col["Payment"])
            c_pay.value = None
            c_pay.number_format = 'mm"월" dd"일"'


            # === 신규: 해당 행 A~H 정렬/테두리 적용
            for col in range(1, len(header_to_col) + 1):  # 1..8
                cell = ws.cell(row=append_row, column=col)
                cell.alignment = align_center
                cell.border = border_all

            append_row += 1
        wb.save(xlsx_path)
        self.log_func(f"엑셀 동기화 완료: {xlsx_path} / 시트: {sheet_name}")
    # endregion


    # region run
    def run(self):
        # .pdf 로 끝나는 파일들을 가져옴.
        files = [f for f in sorted(os.listdir(self.folder)) if f.lower().endswith(".pdf")]
        total = len(files)

        objs = []  # === 신규 === 결과 객체 배열

        for idx, fname in enumerate(files, 1):
            self.log_func(f"[{idx}/{total}] : {fname}")
            fpath = os.path.join(self.folder, fname)
            txt = self.read_text(fpath)
            fields = self.parse_fields(txt)

            # === 신규 === 결과 객체 생성 (Order no. = fname)
            order_no = os.path.splitext(fname)[0]  # === 신규: .pdf 제거

            obj = {
                "Order no.": order_no,
                "OC": None,
                "Invoice no.": fields.get("Invoice no."),
                "Amount": fields.get("Amount"),
                "Invoice date": fields.get("Invoice date"),
                "Due date": None,
                "Payment": None,
                "Note": None,
            }
            objs.append(obj)

            self.log_func(
                f"[RESULT] {fname} -> "
                f"Invoice date={fields.get('Invoice date')}, "
                f"Invoice no.={fields.get('Invoice no.')}, "
                f"Amount={fields.get('Amount')}"
            )

        # === 신규 === 엑셀 반영
        if objs:
            self.sync_excel(objs)
    # endregion
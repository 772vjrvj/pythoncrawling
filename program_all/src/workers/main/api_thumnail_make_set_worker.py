# /src/workers/api_thumnail_make_set_load_worker.py
import os
import time
import shutil
from datetime import datetime
from urllib.parse import urlparse  # (기존 유지)

from PIL import Image, ImageEnhance
from openpyxl import load_workbook

from src.utils.api_utils import APIClient
from src.utils.excel_utils import ExcelUtils
from src.utils.file_utils import FileUtils
from src.workers.api_base_worker import BaseApiWorker
from src.utils.str_utils import to_str
from src.utils.number_utils import to_int


class ApiThumnailMakeSetLoadWorker(BaseApiWorker):

    def __init__(self):
        super().__init__()
        self.driver = None
        self.file_driver = None
        self.excel_driver = None

        self.total_cnt = 0
        self.current_cnt = 0
        self.before_pro_value = 0

        self.api_client = APIClient(use_cache=False)

        # === 신규(DB.xlsx 직접 업데이트용) ===
        self.db_dir = os.path.join(os.getcwd(), "DB")
        self.db_path = os.path.join(self.db_dir, "DB.xlsx")
        self.db_wb = None
        self.db_ws = None
        self.db_cols = {}           # {"이미지 URL": col_idx, ...}
        self.db_sheet_name = ""     # active 시트명
        self.dirty_cnt = 0
        self.save_every = 200       # ✅ 200건마다 저장
        self._stopped = False

    def init(self):
        self.excel_driver = ExcelUtils(self.log_signal_func)
        self.file_driver = FileUtils(self.log_signal_func, api_client=self.api_client)  # === 변경 ===
        return True

    def main(self):
        try:
            self.log_signal_func("작업 시작(DB.xlsx 직접 업데이트)")
            self.log_signal_func(f"세팅 항목: {self.setting}")
            self.log_signal_func(f"컬럼 항목: {self.columns}")

            # 0) DB 오픈 + rows 로드
            self._open_db()
            db_rows = self._read_db_rows()

            if not db_rows:
                self.log_signal_func("❌ DB.xlsx 데이터 없음(2행~)")
                return False

            self.total_cnt = len(db_rows)
            self.current_cnt = 0

            # 폴더 준비
            origin_dir = self.file_driver.create_folder("이미지 저장")
            edit_dir = self.file_driver.create_folder("이미지 수정")

            # setting
            tw = to_int(self.get_setting_value(self.setting, "thumb_width"), 1000)
            th = to_int(self.get_setting_value(self.setting, "thumb_height"), 1000)
            rotate_deg = to_int(self.get_setting_value(self.setting, "thumb_rotate_deg"), 0)
            scale_pct = to_int(self.get_setting_value(self.setting, "thumb_scale_pct"), 100)
            ext = to_str(self.get_setting_value(self.setting, "thumb_ext"), "jpg").lower().strip(".")
            delay_sec = to_int(self.get_setting_value(self.setting, "thumb_delay_sec"), 0)

            wm_enabled = bool(self.get_setting_value(self.setting, "wm_enabled"))
            wm_file = to_str(self.get_setting_value(self.setting, "wm_file"), "watermark.png")
            wm_width = to_int(self.get_setting_value(self.setting, "wm_width"), 35)
            wm_height = to_int(self.get_setting_value(self.setting, "wm_height"), 35)
            # wm_opacity = to_int(self.get_setting_value(self.setting, "wm_opacity_pct"), 15)
            wm_opacity = to_int(self.get_setting_value(self.setting, "wm_opacity_pct"), 100)
            wm_anchor = to_str(self.get_setting_value(self.setting, "wm_anchor"), "br")
            wm_padding = to_int(self.get_setting_value(self.setting, "wm_padding"), 20)
            wm_x_offset = to_int(self.get_setting_value(self.setting, "wm_x_offset"), 0)
            wm_y_offset = to_int(self.get_setting_value(self.setting, "wm_y_offset"), 0)

            wm_path = self._resolve_wm_path(wm_file)
            if wm_enabled:
                if wm_path and os.path.exists(wm_path):
                    self.log_signal_func(f"워터마크 사용: {wm_path}")
                else:
                    self.log_signal_func(f"⚠️ 워터마크 ON 이지만 파일 없음: {wm_path} (워터마크 스킵)")

            # 1) 루프
            for idx, (excel_row_idx, row) in enumerate(db_rows, start=1):
                if not self.running:
                    self._stopped = True
                    self.log_signal_func("⛔ 사용자 중단 요청 감지(저장 후 종료)")
                    break

                self.current_cnt += 1

                try:
                    # 성공이면 스킵
                    status = to_str(row.get("상태"), "").strip()
                    if status == "성공":
                        self.log_signal_func(f"↩️ 스킵(성공): {idx}/{self.total_cnt}")
                        continue

                    # URL
                    url = to_str(row.get("이미지 URL"), "").strip()
                    if not url:
                        raise Exception("이미지 URL 없음")

                    # 결과 파일명
                    result_filename = to_str(row.get("결과 파일명"), "").strip()
                    result_filename = self._safe_filename(result_filename)
                    if not result_filename:
                        result_filename = f"{idx}.{ext}"
                    else:
                        result_filename = self._ensure_ext(result_filename, ext)

                    # 수정 파일명
                    edit_filename = to_str(row.get("수정 파일명"), "").strip()
                    edit_filename = self._safe_filename(edit_filename)
                    if not edit_filename:
                        base, _ = os.path.splitext(result_filename)
                        edit_filename = f"{base}_edit.{ext}"
                    else:
                        edit_filename = self._ensure_ext(edit_filename, ext)

                    # 1) 원본 다운로드
                    origin_path = self.file_driver.save_image(origin_dir, result_filename, url)
                    if not origin_path:
                        raise Exception("원본 이미지 저장 실패")

                    # 2) 수정본 생성 (cover+crop)
                    edit_path = os.path.join(edit_dir, edit_filename)
                    edit_path = self._make_edit_image(
                        src_path=origin_path,
                        dst_path=edit_path,
                        tw=tw,
                        th=th,
                        rotate_deg=rotate_deg,
                        scale_pct=scale_pct,
                        ext=ext,
                    )

                    # 3) 워터마크 (수정본에 합성)
                    if wm_enabled and wm_path and os.path.exists(wm_path):
                        edit_path = self._apply_watermark(
                            base_path=edit_path,
                            wm_path=wm_path,
                            wm_w=wm_width,
                            wm_h=wm_height,
                            opacity_pct=wm_opacity,
                            anchor=wm_anchor,
                            padding=wm_padding,
                            x_off=wm_x_offset,
                            y_off=wm_y_offset,
                        )

                    # row 업데이트 (dict)
                    row["이미지 URL"] = url
                    row["결과 파일명"] = result_filename
                    row["수정 파일명"] = edit_filename
                    row["결과 파일 경로"] = origin_path
                    row["수정 파일 경로"] = edit_path
                    row["상태"] = "성공"
                    row["메모"] = ""

                    # DB.xlsx(메모리) 반영
                    self._write_db_row(excel_row_idx, row)

                    self.log_signal_func(f"✅ 완료: {idx}/{self.total_cnt}  {result_filename}")

                except Exception as e:
                    row["상태"] = "실패"
                    row["메모"] = str(e)
                    # 실패도 DB.xlsx에 반영(재시도/로그 목적)
                    self._write_db_row(excel_row_idx, row)
                    self.log_signal_func(f"❌ 실패: {idx}/{self.total_cnt}  {e}")

                # progress
                pro_value = (self.current_cnt / self.total_cnt) * 1000000
                self.progress_signal.emit(self.before_pro_value, pro_value)
                self.before_pro_value = pro_value

                if delay_sec > 0:
                    time.sleep(delay_sec)

            # 2) 마지막 저장 (중간에 멈췄든, 끝났든 무조건 flush)
            self._flush_db()

            if self._stopped:
                self.log_signal_func("🧾 사용자 중단 처리 완료(저장 완료)")
            else:
                self.log_signal_func("🧾 전체 처리 완료(저장 완료)")

            return True

        except Exception as e:
            # 예외 시에도 혹시 변경된 게 있으면 저장 시도
            try:
                self._flush_db()
            except Exception:
                pass
            self.log_signal_func(f"❌ 전체 실행 중 예외 발생: {e}")
            return False

    def destroy(self):
        # 종료 시점 flush 저장
        try:
            self._flush_db()
        except Exception:
            pass

        self.progress_signal.emit(self.before_pro_value, 1000000)
        self.log_signal_func("=============== 작업 종료")
        self.progress_end_signal.emit()

    def stop(self):
        # UI에서 중지 눌렀을 때: 즉시 저장하고 종료
        self.running = False
        self._stopped = True
        try:
            self._flush_db()
        except Exception:
            pass

        if self.driver:
            self.driver.quit()

    def _ensure_db_exists(self):
        os.makedirs(self.db_dir, exist_ok=True)
        if not os.path.exists(self.db_path):
            raise Exception(f"DB.xlsx 없음: {self.db_path}")

    def _backup_db(self):
        """
        DB/DB.xlsx -> DB/bak/DB_YYYYMMDD_HHMMSS.xlsx
        """
        try:
            bak_dir = os.path.join(self.db_dir, "bak")
            os.makedirs(bak_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak_path = os.path.join(bak_dir, f"DB_{ts}.xlsx")
            shutil.copy2(self.db_path, bak_path)
            self.log_signal_func(f"[DB] 백업 생성: {bak_path}")
        except Exception as e:
            # 백업 실패는 치명적이진 않지만 알림은 필요
            self.log_signal_func(f"[DB] 백업 실패: {e}")

    def _open_db(self):
        self._ensure_db_exists()
        self._backup_db()

        self.db_wb = load_workbook(self.db_path)
        self.db_ws = self.db_wb.active
        self.db_sheet_name = self.db_ws.title

        # 1행 헤더 매핑
        header = [str(c.value or "").strip() for c in self.db_ws[1]]
        self.db_cols = {name: i + 1 for i, name in enumerate(header) if name}

        for k in (self.columns or []):
            if k not in self.db_cols:
                raise Exception(f"DB.xlsx 헤더에 컬럼 없음: {k}")

        self.log_signal_func(f"[DB] 로드 완료: {self.db_path} (sheet={self.db_sheet_name})")

    def _read_db_rows(self):
        """
        DB.xlsx의 데이터(2행~)를 dict로 읽어서 (excel_row_idx, row_dict) 리스트로 반환
        """
        out = []
        ws = self.db_ws

        max_row = ws.max_row or 1
        if max_row <= 1:
            return out

        col = self.db_cols

        for r in range(2, max_row + 1):
            row = {}
            for k in (self.columns or []):
                row[k] = to_str(ws.cell(r, col[k]).value, "").strip()
            out.append((r, row))
        return out

    def _write_db_row(self, excel_row_idx: int, row: dict):
        """
        메모리 워크북에만 반영 (저장은 _flush_db에서)
        """
        ws = self.db_ws
        col = self.db_cols
        for k in (self.columns or []):
            ws.cell(excel_row_idx, col[k]).value = row.get(k, "")

        self.dirty_cnt += 1
        if self.dirty_cnt >= self.save_every:
            self._flush_db()

    def _flush_db(self):
        """
        파일 저장. stop/destroy에서도 호출됨.
        """
        if not self.db_wb:
            return
        if self.dirty_cnt <= 0:
            return
        try:
            self.db_wb.save(self.db_path)
            self.log_signal_func(f"[DB] 저장 완료 (+{self.dirty_cnt}건)")
            self.dirty_cnt = 0
        except Exception as e:
            self.log_signal_func(f"[DB] 저장 실패: {e}")

    def _safe_filename(self, name: str) -> str:
        name = (name or "").strip()
        if not name:
            return ""
        bad = r'<>:"/\|?*'
        for ch in bad:
            name = name.replace(ch, "_")
        name = name.replace("\n", " ").replace("\r", " ").strip()
        return name

    def _ensure_ext(self, filename: str, ext: str) -> str:
        filename = self._safe_filename(filename)
        ext = (ext or "jpg").lower().strip(".")
        base, cur = os.path.splitext(filename)
        if not filename:
            return ""
        if cur:
            return base + cur
        return f"{base}.{ext}"

    def _center_crop(self, img: Image.Image, tw: int, th: int) -> Image.Image:
        #
        # +---------------------------+
        # |   잘림   |               |   잘림   |
        # |----------|   남는 영역   |----------|
        # |          |   (tw x th)   |          |
        # |----------|               |----------|
        # |   잘림   |               |   잘림   |
        # +---------------------------+
        # 중앙 빼고 잘림
    
        w, h = img.size
        left = (w - tw) // 2
        top = (h - th) // 2
        return img.crop((left, top, left + tw, top + th))

    def _make_edit_image(self, src_path, dst_path, tw, th, rotate_deg, scale_pct, ext):
        """
        ✅ 요구사항 버전: 덮는 형태 (cover + center crop)
        - 최종은 tw x th 고정
        - 먼저 cover로 꽉 채움
        - scale_pct(150 등)는 더 크게 만들어 overflow 유도
        - rotate 후에도 overflow 가능
        - 마지막은 center crop으로 tw x th로 자름
        """
        img = Image.open(src_path).convert("RGBA")

        # 0) cover로 tw x th를 꽉 채우는 크기로 resize
        w, h = img.size
        cover = max(tw / w, th / h)
        rw = max(1, int(w * cover))
        rh = max(1, int(h * cover))
        img = img.resize((rw, rh), Image.LANCZOS)

        # 1) scale_pct 적용 (150 => 1.5배)
        if scale_pct and scale_pct != 100:
            s = scale_pct / 100.0
            w, h = img.size
            img = img.resize(
                (max(1, int(w * s)), max(1, int(h * s))),
                Image.LANCZOS
            )

        # 2) 회전 (expand=True)
        if rotate_deg:
            img = img.rotate(-rotate_deg, expand=True)

        # 3) center crop
        # “최종 결과 이미지는 무조건 tw × th 크기여야 한다.”
        # (비율 유지 + 잘라내기 OK)
        # 늘려지는게 아니고 빈 부분의 픽셀이 채워짐 흰색으로
        w, h = img.size
        if w < tw or h < th:
            cover2 = max(tw / w, th / h)
            img = img.resize(
                (max(1, int(w * cover2)), max(1, int(h * cover2))),
                Image.LANCZOS
            )

        out_img = self._center_crop(img, tw, th)

        # 4) 저장
        ext_l = (ext or "jpg").lower().strip(".")
        os.makedirs(os.path.dirname(dst_path), exist_ok=True)

        if ext_l in ("jpg", "jpeg"):
            out_img.convert("RGB").save(dst_path, format="JPEG", quality=95)
        elif ext_l == "png":
            out_img.save(dst_path, format="PNG")
        elif ext_l == "webp":
            out_img.convert("RGB").save(dst_path, format="WEBP", quality=90)
        else:
            dst_path = os.path.splitext(dst_path)[0] + ".jpg"
            out_img.convert("RGB").save(dst_path, format="JPEG", quality=95)

        return dst_path

    def _apply_watermark(self, base_path, wm_path, wm_w, wm_h, opacity_pct, anchor, padding, x_off, y_off):
        if not wm_path or not os.path.exists(wm_path):
            return base_path

        base = Image.open(base_path).convert("RGBA")
        wm = Image.open(wm_path).convert("RGBA")

        wm = wm.resize((max(1, wm_w), max(1, wm_h)), Image.LANCZOS)

        opacity = max(0, min(100, int(opacity_pct)))
        if opacity < 100:
            alpha = wm.split()[-1]
            alpha = ImageEnhance.Brightness(alpha).enhance(opacity / 100.0)
            wm.putalpha(alpha)

        W, H = base.size
        w, h = wm.size
        p = max(0, int(padding))
        xo = int(x_off)
        yo = int(y_off)

        anchor = (anchor or "br").lower().strip()
        if anchor == "tl":
            x = p + xo
            y = p + yo
        elif anchor == "tr":
            x = (W - w - p) + xo
            y = p + yo
        elif anchor == "bl":
            x = p + xo
            y = (H - h - p) + yo
        else:
            x = (W - w - p) + xo
            y = (H - h - p) + yo

        x = max(0, min(W - w, x))
        y = max(0, min(H - h, y))

        # 덮어쓰기(합성)
        base.paste(wm, (x, y), wm)

        ext = os.path.splitext(base_path)[1].lower()
        if ext in (".jpg", ".jpeg", ".webp"):
            out = base.convert("RGB")
            if ext in (".jpg", ".jpeg"):
                out.save(base_path, format="JPEG", quality=95)
            else:
                out.save(base_path, format="WEBP", quality=90)
        else:
            base.save(base_path, format="PNG")

        return base_path

    def _resolve_wm_path(self, wm_file: str) -> str:
        wm_file = (wm_file or "").strip()
        if not wm_file:
            return ""
        if os.path.isabs(wm_file):
            return wm_file
        return os.path.join(os.getcwd(), wm_file)

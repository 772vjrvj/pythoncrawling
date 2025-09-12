import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class ExcelUtils:
    def __init__(self, log_func=None):
        self.log_func = log_func


    def init_csv(self, filename, columns):
        df = pd.DataFrame(columns=columns)
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        if self.log_func:
            self.log_func(f"CSV 초기화 완료: {filename}")


    def append_to_csv(self, filename, data_list, columns):

        if not data_list:
            return

        df = pd.DataFrame(data_list, columns=columns)
        df.to_csv(filename, mode='a', header=False, index=False, encoding="utf-8-sig")
        data_list.clear()
        self.log_func("csv 저장완료")


    def append_to_excel(self, filename, data_list, columns, sheet_name="Sheet1"):
        if not data_list:
            return

        df = pd.DataFrame(data_list, columns=columns)

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                start_row = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        data_list.clear()
        self.log_func("excel 저장완료")


    def convert_csv_to_excel_and_delete(self, csv_filename, sheet_name="Sheet1"):
        """
        CSV 파일을 엑셀 파일로 변환한 후, 원본 CSV 파일을 삭제합니다.
        엑셀 파일명은 CSV와 동일하고 확장자만 .xlsx 로 바뀝니다.
        """
        if not os.path.exists(csv_filename):
            if self.log_func:
                self.log_func(f"❌ CSV 파일이 존재하지 않습니다: {csv_filename}")
            return

        try:
            df = pd.read_csv(csv_filename, encoding="utf-8-sig")

            if df.empty:
                if self.log_func:
                    self.log_func(f"⚠️ CSV에 데이터가 없습니다: {csv_filename}")
                return

            # 엑셀 파일 이름은 CSV와 동일하게 (확장자만 .xlsx)
            excel_filename = os.path.splitext(csv_filename)[0] + ".xlsx"

            with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)

            os.remove(csv_filename)

            if self.log_func:
                self.log_func(f"✅ 엑셀 파일 저장 완료: {excel_filename}")
                self.log_func(f"🗑️ CSV 파일 삭제 완료: {csv_filename}")

        except Exception as e:
            if self.log_func:
                self.log_func(f"❌ 변환 중 오류 발생: {e}")


    # 일반 객체 리스트인 경우
    def obj_to_row(self, o, cols):
        if isinstance(o, dict):
            return {c: o.get(c) for c in cols}
        # 객체 속성에서 추출
        return {c: getattr(o, c, None) for c in cols}


    def obj_list_to_dataframe(self, obj_list, columns=None):
        """
        obj_list 를 pandas.DataFrame 으로 변환
        - obj_list 가 dict 리스트면 그대로 사용
        - 일반 객체 리스트면 __dict__ 또는 지정 columns 기준으로 추출
        """
        if not obj_list:
            return None

        # dict 리스트인 경우
        if isinstance(obj_list[0], dict):
            if columns:
                rows = [{col: obj.get(col) for col in columns} for obj in obj_list]
                return pd.DataFrame(rows, columns=columns)
            return pd.DataFrame(obj_list)

        if columns:
            rows = [self.obj_to_row(o, columns) for o in obj_list]
            return pd.DataFrame(rows, columns=columns)

        # columns 미지정이면 첫 객체의 __dict__ 키 사용
        first = obj_list[0]
        if hasattr(first, "__dict__") and first.__dict__:
            cols = list(first.__dict__.keys())
            rows = [self.obj_to_row(o, cols) for o in obj_list]
            return pd.DataFrame(rows, columns=cols)

        # 마지막 fallback: dir 기반(언더스코어/콜러블 제외)
        cols = [k for k in dir(first) if not k.startswith("_") and not callable(getattr(first, k, None))]
        rows = [self.obj_to_row(o, cols) for o in obj_list]
        return pd.DataFrame(rows, columns=cols)


    def save_obj_list_to_excel(self, filename, obj_list, columns=None, sheet_name="Sheet1"):
        """
        obj_list(객체/딕셔너리 리스트)를 엑셀 파일에 저장합니다.
        - 파일이 존재하면 같은 시트에 '이어쓰기'(header 없이)
        - 파일이 없거나 시트가 없으면 시트를 새로 만들고 header 포함 저장
        - columns 지정 시 해당 컬럼 순서/이름으로 저장
        - URL 포함된 값은 하이퍼링크로 변환
        """
        if not obj_list:
            return

        df = self.obj_list_to_dataframe(obj_list, columns=columns)
        if df is None or df.empty:
            if self.log_func:
                self.log_func("⚠️ 저장할 데이터가 없습니다.")
            return

        # 이어쓰기/신규 작성 처리
        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                ws = writer.sheets.get(sheet_name)
                if ws is not None:
                    start_row = ws.max_row if ws.max_row is not None else 0
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        # === URL 컬럼을 하이퍼링크로 변환 ===
        wb = load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=2):  # 1행은 header라 skip
                for cell in row:
                    val = str(cell.value) if cell.value else ""
                    if val.startswith("http://") or val.startswith("https://"):
                        cell.hyperlink = val
                        cell.style = "Hyperlink"

        wb.save(filename)

        # 원본 리스트 정리 및 로그
        obj_list.clear()
        if self.log_func:
            self.log_func("excel(객체 리스트) 저장완료 (URL 하이퍼링크 처리)")

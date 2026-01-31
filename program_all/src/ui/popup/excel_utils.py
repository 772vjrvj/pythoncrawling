import os
import pandas as pd
from openpyxl import load_workbook, Workbook


class ExcelUtils:
    def __init__(self, log_func=None):
        self.log_func = log_func

    # =========================
    # 기존 기능들(유지)
    # =========================
    def init_csv(self, filename, columns):
        df = pd.DataFrame(columns=columns)
        df.to_csv(filename, index=False, encoding="utf-8-sig")
        if self.log_func:
            self.log_func(f"CSV 초기화 완료: {filename}")

    def append_to_csv(self, filename, data_list, columns):
        if not data_list:
            return
        df = pd.DataFrame(data_list, columns=columns)
        df.to_csv(filename, mode='a', header=False, index=False, encoding="utf-8-sig")
        data_list.clear()
        if self.log_func:
            self.log_func("csv 저장완료")

    def append_to_excel(self, filename, data_list, columns, sheet_name="Sheet1"):
        if not data_list:
            return
        df = pd.DataFrame(data_list, columns=columns)

        if os.path.exists(filename):
            with pd.ExcelWriter(filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                start_row = writer.sheets[sheet_name].max_row if sheet_name in writer.sheets else 0
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
        else:
            with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        data_list.clear()
        if self.log_func:
            self.log_func("excel 저장완료")

    # ==========================================================
    # === 신규 === 원본 엑셀에 "직접" write-back
    # ==========================================================
    def update_rows_in_place(self, excel_path, rows, sheet_index=0, header_row=1):
        """
        excel_path: 원본 엑셀 경로
        rows: dict list
          - 반드시 "__row_idx" 포함 (엑셀 행 번호)
          - 업데이트할 키(한글헤더): 결과 파일명/수정 파일명/상태/메모/결과 파일 경로/수정 파일 경로
        sheet_index: 0이면 첫번째 시트
        header_row: 헤더 행 (기본 1)
        """
        if not rows:
            return
        if not excel_path or not os.path.exists(excel_path):
            if self.log_func:
                self.log_func(f"[EXCEL] 원본 파일 없음: {excel_path}")
            return

        wb = load_workbook(excel_path)
        ws = wb.worksheets[sheet_index]

        # 헤더 -> 컬럼번호 맵
        header_map = {}
        max_col = ws.max_column or 0
        for c in range(1, max_col + 1):
            v = ws.cell(header_row, c).value
            if v is None:
                continue
            header_map[str(v).strip()] = c

        def set_cell(row_idx, header_name, value):
            col = header_map.get(header_name)
            if not col:
                return
            ws.cell(row_idx, col).value = "" if value is None else str(value)

        for r in rows:
            row_idx = r.get("__row_idx")
            if not row_idx:
                continue

            set_cell(row_idx, "결과 파일명", r.get("결과 파일명"))
            set_cell(row_idx, "수정 파일명", r.get("수정 파일명"))
            set_cell(row_idx, "상태", r.get("상태"))
            set_cell(row_idx, "메모", r.get("메모"))
            set_cell(row_idx, "결과 파일 경로", r.get("결과 파일 경로"))
            set_cell(row_idx, "수정 파일 경로", r.get("수정 파일 경로"))

        wb.save(excel_path)

        if self.log_func:
            self.log_func(f"[EXCEL] 원본 반영 완료: {excel_path} (rows={len(rows)})")

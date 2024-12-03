import requests
from bs4 import BeautifulSoup
import re
import time
import random
import pandas as pd
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

def send_post_request(url, payload):
    """POST 요청을 보내고 응답을 반환하는 함수"""
    try:
        response = requests.post(url, data=payload)
        response.raise_for_status()  # 요청 성공 여부 확인
        return response.text
    except requests.RequestException as e:
        print(f"HTTP 요청 에러: {e}")
        return None

def parse_html_for_ids(html, page_index, id_index):
    """HTML을 파싱하여 필요한 ID들을 추출하는 함수"""
    try:
        soup = BeautifulSoup(html, 'html.parser')
        pattern = re.compile(r"javascript:dmoelRequst\('([^']+)'\);")

        results = []
        table = soup.find("table", class_="board")
        tbody = table.find("tbody") if table else None

        if tbody:
            for tr in tbody.find_all("tr"):
                tds = tr.find_all("td")
                if len(tds) < 6:
                    continue

                institution_name, homepage = tds[0].text.strip().split('\n', 1) if '\n' in tds[0].text.strip() else (tds[0].text.strip(), '')

                a_tag = tds[0].find("a")
                if a_tag and 'href' in a_tag.attrs and 'javascript:alert' in a_tag['href']:
                    id_value = None
                else:
                    id_value = pattern.search(a_tag['href']).group(1) if a_tag else None

                institution = {
                    "교육기관명": clean_text_for_excel(format_text(institution_name.strip())),
                    "홈페이지": clean_text_for_excel(format_text(homepage.strip())),
                    "ID": id_value,
                    "교육종목": clean_text_for_excel(format_text(tds[1].text.strip())),
                    "활동지역": clean_text_for_excel(format_text(tds[2].text.strip())),
                    "보유강사": clean_text_for_excel(format_text(tds[3].text.strip())),
                    "지정일자": clean_text_for_excel(format_text(tds[4].text.strip())),
                    "연락처": clean_text_for_excel(format_text(tds[5].text.strip())),
                }
                id_index += 1
                print(f"Page {page_index}, Index : {id_index}, Collected Obj - {institution}")

                results.append(institution)

        return results, id_index
    except Exception as e:
        print(f"HTML 파싱 에러: {e}")
        return []

def format_text(text):
    """텍스트를 원하는 형식으로 변환하는 함수"""
    formatted_text = text.replace("& lt;", "<").replace("& gt;", ">").replace("\n\n", "\n")
    return formatted_text

def clean_text_for_excel(text):
    """엑셀에서 사용 불가능한 문자를 제거하는 함수"""
    return re.sub(r'[\000-\010]|[\013-\014]|[\016-\037]', '', text)

def parse_institution_details(institution, detail_index):
    eclstNo = institution["ID"]
    url = "https://edu.kead.or.kr/aisd/search/EclstProfile.do"
    payload = {'eclstNo': eclstNo, 'pageNum': 1}
    response_text = send_post_request(url, payload)
    if response_text is None:
        return institution

    soup = BeautifulSoup(response_text, 'html.parser')
    details = institution.copy()

    try:
        t_con_sections = soup.find_all("div", class_="t-con")
        if len(t_con_sections) < 5:
            print(f"Expected at least 5 t-con sections for institution {eclstNo}, but found {len(t_con_sections)}")
            return institution

        # 교육기관 정보
        profile_section = t_con_sections[0]
        img_tag = profile_section.find("span", class_="profile-info").find("img")
        if img_tag:
            img_url = img_tag['src']
            if not img_url.startswith('http'):
                img_url = 'https://edu.kead.or.kr' + img_url
            details["이미지 URL"] = img_url
        else:
            details["이미지 URL"] = ""

        table_rows = profile_section.find("table", class_="view-table type-2").find("tbody").find_all("tr")
        for row in table_rows:
            ths = row.find_all("th")
            tds = row.find_all("td")
            for th, td in zip(ths, tds):
                th_text = clean_text_for_excel(format_text(th.text.strip()))
                td_text = clean_text_for_excel(format_text(td.text.strip()))
                if th_text in ["교육기관명", "교육종목", "연락처", "이메일", "소재지", "홈페이지"]:
                    details[th_text] = td_text
                elif th_text == "활동지역":
                    details["활동지역"] = ", ".join([clean_text_for_excel(format_text(li.text.strip())) for li in td.find("ul").find_all("li")])
                elif th_text == "활동요일":
                    details["활동요일"] = ", ".join([clean_text_for_excel(format_text(li.text.strip())) for li in td.find("ul").find_all("li")])

        # 홍보자료
        promo_section = t_con_sections[1]
        promo_rows = promo_section.find("table", class_="view-table type-2").find("tbody").find_all("tr")
        for row in promo_rows:
            ths = row.find_all("th")
            tds = row.find_all("td")
            for th, td in zip(ths, tds):
                th_text = clean_text_for_excel(format_text(th.text.strip()))
                td_text = clean_text_for_excel(format_text(td.text.strip()))
                if th_text in ["샘플 강의 동영상", "홍보자료1", "홍보자료2"]:
                    details[th_text] = td_text

        # 기관소개
        intro_section = t_con_sections[2]
        intro_rows = intro_section.find("table", class_="view-table type-2").find("tbody").find_all("tr")
        for row in intro_rows:
            ths = row.find_all("th")
            tds = row.find_all("td")
            for th, td in zip(ths, tds):
                th_text = clean_text_for_excel(format_text(th.text.strip()))
                td_text = clean_text_for_excel(td.find("textarea").text.strip() if td.find("textarea") else format_text(td.text.strip()))
                if th_text in ["주요연혁", "주요활동"]:
                    details[th_text] = td_text

        # 강의소개
        lecture_section = t_con_sections[3]
        lecture_rows = lecture_section.find("table", class_="view-table type-2").find("tbody").find_all("tr")
        for row in lecture_rows:
            ths = row.find_all("th")
            tds = row.find_all("td")
            for th, td in zip(ths, tds):
                th_text = clean_text_for_excel(format_text(th.text.strip()))
                td_text = clean_text_for_excel(td.find("textarea").text.strip() if td.find("textarea") else format_text(td.text.strip()))
                if th_text in ["강의소개", "강의목차"]:
                    details[th_text] = td_text

        # 최근 교육 실적
        recent_education_section = t_con_sections[4]
        recent_education_table = recent_education_section.find("table", class_="list-table")
        education_records = []
        no_records = recent_education_table.find("td", class_="ac")
        if no_records and "등록된 내용이 없습니다." in no_records.text:
            education_records.append({
                "교육일자": "",
                "교육시간": "",
                "교육 의뢰 기관": "",
                "교육장소": ""
            })
        else:
            for row in recent_education_table.find("tbody").find_all("tr"):
                record = {
                    "교육일자": clean_text_for_excel(format_text(row.find_all("td")[0].text.strip())),
                    "교육시간": clean_text_for_excel(format_text(row.find_all("td")[1].text.strip())),
                    "교육 의뢰 기관": clean_text_for_excel(format_text(row.find_all("td")[2].text.strip())),
                    "교육장소": clean_text_for_excel(format_text(row.find_all("td")[3].text.strip()))
                }
                education_records.append(record)
        details["최근 교육 실적"] = json.dumps(education_records, ensure_ascii=False)

    except Exception as e:
        print(f"상세 정보 파싱 에러: {e} for eclstNo {eclstNo}")
        return institution

    print(f"Detail {detail_index}: {details}")
    return details

def save_images_to_folder(details, folder="images"):
    """이미지 URL을 다운로드하여 폴더에 저장하고, 파일 경로를 반환하는 함수"""
    if not os.path.exists(folder):
        os.makedirs(folder)
    for detail in details:
        image_url = detail.get("이미지 URL", "")
        if image_url:
            image_name = os.path.join(folder, image_url.split("/")[-1])
            response = requests.get(image_url)
            with open(image_name, 'wb') as file:
                file.write(response.content)
            detail["이미지"] = image_name
        else:
            detail["이미지"] = ""
    return details

def save_to_excel_with_images(all_details, file_name="장애인 교육기관.xlsx"):
    df = pd.DataFrame(all_details)

    # 순서 맞추기: 교육기관 정보 -> 홍보자료 -> 강의소개 -> 기관소개 -> 최근 교육 실적 -> 페이지
    ordered_columns = [
        '교육기관명', '교육종목', '연락처', '이메일', '이미지', '이미지 URL',
        '한줄소개', '홈페이지', '활동요일', '활동지역', '소재지', '보유강사', '지정일자',
        '샘플 강의 동영상', '홍보자료1', '홍보자료2',
        '주요연혁', '주요활동',
        '강의소개', '강의목차',
        '최근 교육 실적'
    ]

    # 빈 데이터프레임에 컬럼 추가
    for col in ordered_columns:
        if col not in df.columns:
            df[col] = ''

    df = df[ordered_columns]

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # 헤더 추가
        ws.append(df.columns.tolist())

    # 데이터프레임을 엑셀 워크북으로 변환
    start_row = ws.max_row + 1
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # 이미지 파일을 엑셀에 삽입
    for idx, row in enumerate(df.itertuples(), start=start_row):
        img_path = row.이미지
        if img_path and os.path.exists(img_path):
            img = Image(img_path)
            img.width = 100  # 이미지 너비 조정
            img.height = 100  # 이미지 높이 조정
            ws.add_image(img, f"E{idx}")
            # 셀 크기를 이미지 크기에 맞게 조정
            ws.row_dimensions[idx].height = 100  # 높이 조정
            ws.column_dimensions['E'].width = 20  # 너비 조정

    wb.save(file_name)

def main():
    url = "https://edu.kead.or.kr/aisd/search/EclstSearchList.do"
    page_num = 1
    all_institutions = []
    id_index = 0

    # 교육기관 ID 수집
    while True:
        payload = {'pageNum': page_num, 'pageType': 'SEARCH'}
        response_text = send_post_request(url, payload)
        if response_text is not None:
            institutions, in_id_index = parse_html_for_ids(response_text, page_num, id_index)
            if not institutions:
                break
            all_institutions.extend(institutions)
            page_num += 1
            id_index = in_id_index
            time.sleep(1)
        else:
            print("데이터를 가져오지 못했습니다.")
            break

    all_details = []
    file_count = 1

    # 각 교육기관에 대한 상세 정보 수집
    for index, institution in enumerate(all_institutions, start=1):
        print(f"Processing institution {index}: {institution['ID']}")
        if institution['ID'] is not None:
            details = parse_institution_details(institution, index)
        else:
            details = institution  # 기본 데이터를 그대로 사용
        if details:
            all_details.append(details)
            print(f"Finished processing institution {index}: {institution['ID']}")

        # 100개마다 엑셀 파일로 저장
        if index % 100 == 0:
            print(f"Processing 100개 마다 저장")
            all_details = save_images_to_folder(all_details)
            save_to_excel_with_images(all_details, "장애인 교육기관.xlsx")
            file_count += 1
            all_details = []  # 저장 후 리스트 초기화
            print(f"Finished 100개 마다 저장")

        time.sleep(1)

    # 남은 데이터 저장
    if all_details:
        print(f"Processing 남은 데이터 저장")
        all_details = save_images_to_folder(all_details)
        save_to_excel_with_images(all_details, "장애인 교육기관.xlsx")
        print(f"Finished 남은 데이터 저장")

if __name__ == "__main__":
    main()

from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
import time
import traceback
from selenium.common.exceptions import ElementNotInteractableException
import re
import math
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import calendar


base_price = 0
unit = '개'
stnd_cnt = '1'
email_list = []

# 드라이버 설정 함수
def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1080,750")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
    except Exception as e:
        print(f"Error setting up the driver: {e}")
        driver = None
    return driver


# 숫자와 소수점을 포함한 숫자만 추출하는 함수
def extract_numbers(input_string):
    try:
        return ''.join(re.findall(r'\d+\.\d+|\d+', input_string))
    except Exception as e:
        print(f"Error in extracting numbers from {input_string}: {e}")
        return ''  # 에러 발생 시 빈 문자열 반환

# 문자와 단위(g, kg 등)만 추출하는 함수
def extract_non_numbers(input_string):
    try:
        return ''.join(re.findall(r'[^\d\.]+', input_string))
    except Exception as e:
        print(f"Error in extracting non-numbers from {input_string}: {e}")
        return ''  # 에러 발생 시 빈 문자열 반환

# 다나와 크롤링 함수
def scrape_danawa(driver, name, danawa_url, limit_count, on_and_off):
    global base_price
    try:
        if not danawa_url:
            return []
        driver.get(danawa_url)
        time.sleep(3)
        print(danawa_url)
        danawa_temp_list = []

        # 다른구성상품 더보기/접기
        try:
            if driver.find_elements(By.XPATH, '//*[@id="bundleProductMoreButton"]'):
                driver.find_element(By.XPATH, '//*[@id="bundleProductMoreButton"]').click()
                print("다른구성상품 더보기 클릭")
            else:
                print("다른구성상품 더보기 없는 상품")
        except ElementNotInteractableException as e:
            # 버튼이 있지만 상호작용이 불가능한 경우에만 넘어감
            print(f"다른구성상품 열기 버튼이 있지만 상호작용이 불가능하여 넘어갑니다: {e}")
            pass  # 다음 라인으로 넘어가서 계속 진행
        except Exception as e:
            # 다른 예외는 발생 시 빈 리스트 반환
            print(f"Error occurred while trying to open 다른구성상품: {e}")
            return []

        # 다른 구성
        danawa_opt_url_list = [e.get_attribute('href') for e in driver.find_elements(By.CSS_SELECTOR, '[class="list__variant-selector"] li a')] # class명 변경 2025-07-27
        danawa_opt_text_list = [e.text for e in driver.find_elements(By.CSS_SELECTOR, '[class="list__variant-selector"] li .text__spec')] # class명 변경 2025-07-27
        # ['1개', '2개', '3개', '4개', '5개'] #['https://prod.danawa.com/info/?pcode=5970722', 'https://prod.danawa.com/info/?pcode=5970724', 'https://prod.danawa.com/info/?pcode=5970731', 'https://prod.danawa.com/info/?pcode=5970748', 'https://prod.danawa.com/info/?pcode=5970738']

        print(danawa_opt_url_list)
        print(danawa_opt_text_list)

        # 기준 가격 1개 2025-07-27 추가
        if danawa_opt_text_list and danawa_opt_text_list[0]:
            driver.get(danawa_opt_url_list[0])
            time.sleep(2)
            driver.find_elements(By.CSS_SELECTOR, '.cardSaleChkbox')[0].click()  # 카드할인가 클릭
            driver.find_elements(By.CSS_SELECTOR, '.postPriceChkbox')[0].click()  # 배송비 클릭
            time.sleep(1)
            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            numeric_left_price = 0
            numeric_rgt_price = 0

            # 왼쪽 가격 추출
            left_pay_dil_prod_e_list = soup.select('.columm.left_col .diff_item')
            if left_pay_dil_prod_e_list:
                item = left_pay_dil_prod_e_list[0]
                card_price_tag = item.select_one('.card_line .prc_t')
                if card_price_tag:
                    price_str = card_price_tag.text
                else:
                    prc_price_tag = item.select_one('.prc_line .prc_c')
                    price_str = prc_price_tag.text if prc_price_tag else ''
                numeric_left_price = extract_numeric_price(price_str) if price_str else 0

            # 오른쪽 가격 추출
            rgt_pay_dil_prod_e_list = soup.select('.columm.rgt_col .diff_item')
            if rgt_pay_dil_prod_e_list:
                item = rgt_pay_dil_prod_e_list[0]
                prc_price_tag = item.select_one('.prc_c')
                price_str = prc_price_tag.text if prc_price_tag else ''
                numeric_rgt_price = extract_numeric_price(price_str) if price_str else 0

            # 더 작은 값을 base_price로 설정
            base_price = min(numeric_left_price, numeric_rgt_price)


        # 수량옵션이 없는경우 1개로 처리하기 위한 세팅
        if not danawa_opt_url_list:
            danawa_opt_url_list = [1]
            danawa_opt_text_list = [1]


        for ii in range(len(danawa_opt_url_list)):
            print(danawa_opt_url_list[ii])
            print(danawa_opt_text_list[ii])

            if danawa_opt_url_list[ii] != 1:
                if on_and_off == 0 and danawa_opt_text_list[ii] == f'{stnd_cnt}{unit}':  # 1개는 스킵 # 복수 구성이 없는 상품은 에러처리안나게 건너 뛰기
                    continue

                driver.get(danawa_opt_url_list[ii])
                time.sleep(2)

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.execute_script("window.scrollTo(0, 0);")
            driver.find_elements(By.CSS_SELECTOR, '.cardSaleChkbox')[0].click()  # 카드할인가 클릭 2025-07-27 제거
            driver.find_elements(By.CSS_SELECTOR, '.postPriceChkbox')[0].click()  # 배송비 클릭
            time.sleep(1)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 좌측: 무료배송
            pay_dil_prod_e_list = soup.select('.columm.left_col .diff_item')

            for ei in range(min(len(pay_dil_prod_e_list), limit_count)):
                try:
                    item = pay_dil_prod_e_list[ei]  # 변수에 저장하여 재사용

                    mall_name_elem = item.select('img')
                    mall_name = mall_name_elem[0].get('alt') if mall_name_elem else (
                        item.select('a .txt_logo')[0].text if item.select('a .txt_logo') else ""
                    )

                    prod_name = item.select('.info_line')[0].text.strip()


                    # 상품명에 해외가 들어가거나 배송비가 4000이상이면 제외
                    if '해외' in prod_name:
                        print(f"다나와 해외 상품 발견 {prod_name}")
                        continue

                    ship_tag = item.select_one('.prc_line .ship')
                    delivery_info = 0  # 기본값: 배송비 없음 또는 무료

                    if ship_tag:
                        ship_text = ship_tag.text.strip()

                        # 배송비가 숫자일 경우 추출
                        match = re.search(r'배송비\s?([\d,]+)원', ship_text)
                        if match:
                            delivery_info = int(match.group(1).replace(',', ''))
                            if delivery_info >= 4000:
                                print(f"다나와 배송비 4,000원 이상 상품 발견 {prod_name}")
                                continue

                        elif '무료배송' in ship_text:
                            delivery_info = '무료배송'

                    # card_line이 존재하는지 확인하고 적절한 price_str을 설정
                    card_line = item.select('.card_line')
                    if card_line:
                        price_str = card_line[0].select('.prc_t')[0].text
                    else:
                        price_str = item.select('.prc_line')[0].select('.prc_c')[0].text

                    # 숫자만 추출한 가격
                    numeric_price = extract_numeric_price(price_str)


                    # danawa_opt_text_list[ii] 값이 숫자를 포함하지 않으면 '1개' 할당
                    # (일반구매, 해외구매) case
                    qty = f'{stnd_cnt}{unit}'

                    if danawa_opt_text_list[ii] != 1:
                        if not any(char.isdigit() for char in danawa_opt_text_list[ii]):
                            qty = f'{stnd_cnt}{unit}'
                        else:
                            qty = danawa_opt_text_list[ii]

                    one_price = math.floor((numeric_price / convert_to_float(qty)) * float(stnd_cnt))

                    temp_list = [name, '다나와', qty, mall_name, delivery_info, prod_name, numeric_price, one_price]
                    print(f"다나와 {ei} : {temp_list}")
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (free shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

            # 우측: 유/무료 배송
            pay_dil_prod_e_list = soup.select('.columm.rgt_col .diff_item')

            for ei in range(min(len(pay_dil_prod_e_list), limit_count)):
                try:
                    item = pay_dil_prod_e_list[ei]  # 변수에 저장하여 재사용

                    mall_name_elem = item.select('img')
                    mall_name = mall_name_elem[0].get('alt') if mall_name_elem else (
                        item.select('a .txt_logo')[0].text if item.select('a .txt_logo') else ""
                    )

                    prod_name = item.select('.info_line')[0].text.strip()


                    # 상품명에 해외가 들어가거나 배송비가 4000이상이면 제외
                    if '해외' in prod_name:
                        print(f"다나와 해외 상품 발견 {prod_name}")
                        continue

                    ship_tag = item.select_one('.prc_line .ship')
                    delivery_info = 0  # 기본값: 배송비 없음 또는 무료

                    if ship_tag:
                        ship_text = ship_tag.text.strip()

                        # 배송비가 숫자일 경우 추출
                        match = re.search(r'배송비\s?([\d,]+)원', ship_text)
                        if match:
                            delivery_info = int(match.group(1).replace(',', ''))
                            if delivery_info >= 4000:
                                print(f"다나와 배송비 4,000원 이상 상품 발견 {prod_name}")
                                continue

                        elif '무료배송' in ship_text:
                            delivery_info = '무료배송'


                    price_str = item.select('.prc_c')[0].text
                    numeric_price = extract_numeric_price(price_str)  # 숫자만 추출한 가격

                    qty = f'{stnd_cnt}{unit}'
                    if danawa_opt_text_list[ii] != 1:
                        # (일반구매, 해외구매) case
                        if not any(char.isdigit() for char in danawa_opt_text_list[ii]):
                            qty = f'{stnd_cnt}{unit}'
                        else:
                            qty = danawa_opt_text_list[ii]

                    one_price = math.floor((numeric_price / convert_to_float(qty)) * float(stnd_cnt))

                    temp_list = [name, '다나와', qty, mall_name, delivery_info, prod_name, numeric_price, one_price]
                    print(f"다나와 {ei} : {temp_list}")
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (paid shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

        return danawa_temp_list

    except Exception as e:
        print(f"Error in Danawa scraping: {e}")
        return []



# 에누리 크롤링 함수
def scrape_enuri(driver, name, enuri_url, limit_count, on_and_off):
    try:
        # 테스트를 위한 에러 발생
        # raise Exception("테스트를 위한 에러 발생")
        merge_list = []
        if not enuri_url:
            return []
        print(f"enuri_url : {enuri_url}")

        driver.get(enuri_url)
        time.sleep(3)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # SCroll 맨 아래로
        driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로
        time.sleep(0.5)

        # 가격비교 구매옵션의 라디오: 1개 2개 3개 4개 5개 라디오 아래 더보기 버튼
        try:
            if len(driver.find_elements(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more')) != 0:
                driver.find_element(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more').click()
                print("더보기 버튼을 있는 상품.")
            else:
                print("더보기 버튼을 없는 상품.")
        except ElementNotInteractableException as e:
            # 더보기 버튼이 있지만 상호작용이 불가능한 경우에만 넘어감
            print(f"더보기 버튼이 있지만 상호작용이 불가능하여 넘어갑니다: {e}")
            pass  # 다음 라인으로 넘어가서 계속 진행
        except Exception as e:
            # 다른 예외는 발생 시 빈 리스트 반환
            print(f"Error occurred while trying to click 더보기 버튼: {e}")
            return []


        time.sleep(1)
        # 가격비교 구매옵션의 라디오 : 1개 2개 3개 4개 5개 라디오 아래 더보기 버튼
        radio_opts = driver.find_elements(By.CSS_SELECTOR, '[name="radioOPTION"]')
        xpath_list = ['//*[@for="' + e.get_attribute('id') + '"]' for e in radio_opts] # 라디오옵션을 감싸고 있는 label
        time.sleep(1)


        # 수량옵션이 없는경우 1개로 처리하기 위한 세팅
        if not radio_opts:
            radio_opts = [1]
            xpath_list = [1]

        list_names = []

        for eei, e in enumerate(radio_opts):
            if xpath_list[eei] != 1:
                elem = driver.find_element(By.XPATH, xpath_list[eei])
                elem_text = elem.text.split()

                if elem_text:
                    how_many = elem_text[0]
                    list_names.append(how_many)  # 리스트에 추가
                else:
                    print(f"Warning: No text found for element at index {eei}")
            else:
                list_names = [f'{stnd_cnt}{unit}']

        # 리스트를 쉼표로 구분한 문자열로 결합
        list_names_str = ', '.join(list_names)
        print(f'{list_names_str}')


        for eei, e in enumerate(radio_opts):

            #- 에누리 1개 / 배송비 포함 O / 카드할인 X
            #- 에누리 N개 배송비포함 X / 카드할인 O


            time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로

            how_many = f'{stnd_cnt}{unit}'

            if xpath_list[eei] != 1:

                elem = driver.find_element(By.XPATH, xpath_list[eei])  # 갯수 엘리먼트

                how_many = elem.text.split()[0]  # 갯수 텍스트

                print(how_many)
                if on_and_off == 0 and how_many == f'{stnd_cnt}{unit}':  # 1개는 스킵
                    continue

                elem.click()  # 갯수 클릭
                num = extract_number(how_many, unit)
                enuri_click(driver, num)

            else:
                enuri_click(driver, 1)


            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # SCroll 맨 아래로
            driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로
            time.sleep(1)

            max_wait_time = 10  # 최대 대기 시간 (초)
            start_time = time.time()

            while True:
                try:
                    # 로딩 상태가 아직 보이는지 확인
                    if driver.find_element(By.XPATH, '//*[@class="comm-loader"]').get_attribute('style') != "display: none;":
                        time.sleep(0.5)

                        # 최대 대기 시간이 넘으면 루프 탈출
                        if time.time() - start_time > max_wait_time:
                            print("Loader is taking too long. Skipping this step.")
                            break
                    else:
                        # 로딩이 끝났으면 루프 탈출
                        break
                except Exception as e:
                    # 에러가 발생했을 경우 에러 메시지 출력 후 루프 탈출
                    print(f"Error while waiting for loader: {e}")
                    print(traceback.format_exc())
                    break

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 배송비 포함, 카드할인 아래 이미지, 쇼핑몰, 상품명, 배송비, 판매가, 무이자 할부...  테이블 아래 list tr:  'tb-compare__list' 클래스 내부의 tbody 태그에서 모든 tr 태그를 찾음
            free_dil_prod_e_list = soup.select('.tb-compare__list tbody tr[data-plno]')

            for ei in range(min(len(free_dil_prod_e_list), limit_count)):
                try:
                    # 판매처 추출
                    mall_name = ""
                    shop_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--shop')
                    if shop_td:
                        img_tag = shop_td.find('img')
                        mall_name_origin = img_tag['alt'] if img_tag and 'alt' in img_tag.attrs else shop_td.get_text(strip=True)
                        mall_name = mall_name_change(mall_name_origin)

                    # 제품명 추출
                    prod_name = ""
                    name_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--name')
                    if name_td:
                        name_div = name_td.find('div', class_='tb-col__tx--name')
                        if name_div:
                            # <i> 태그 제거 후 텍스트만 추출
                            [i_tag.extract() for i_tag in name_div.find_all('i')]
                            prod_name = name_div.get_text(strip=True)

                    # 가격 추출
                    price_str = ""
                    price_div = free_dil_prod_e_list[ei].find('div', class_='tx-price')
                    if price_div:
                        em_tag = price_div.find('em')
                        price_str = em_tag.get_text(strip=True) if em_tag else ""

                    # 가격 추출 및 숫자만 저장
                    numeric_price = extract_numeric_price(price_str)  # 숫자만 추출한 가격


                    # how_many 값이 숫자를 포함하지 않으면 '1개' 할당
                    # (일반구매, 해외구매) case
                    if not any(char.isdigit() for char in how_many):
                        qty = f'{stnd_cnt}{unit}'
                    else:
                        qty = how_many

                    # 리스트에 데이터 추가
                    one_price = math.floor((numeric_price / convert_to_float(qty)) * float(stnd_cnt))

                    temp_list = [name, '에누리', qty, mall_name, '무료배송', prod_name, numeric_price, one_price]
                    print(f"에누리 {ei} : {temp_list}")
                    merge_list.append(temp_list)

                except Exception as e:
                    print(f"Error in scraping Enuri free shipping list (index {ei}): {e}")
                    continue

        return merge_list
    except Exception as e:
        print(f"Error in Enuri scraping: {e}")
        return []


def extract_number(how_many, unit):
    return int(re.sub(f'{unit}$', '', how_many).strip())


# - 에누리 1개 / 배송비 포함 O / 카드할인 X
# - 에누리 N개 배송비포함 X / 카드할인 O

def enuri_click(driver, num):


    if num == 1:

        # 카드 할인 토글 클릭
        try:
            # 두 개의 클래스를 가진 label 요소를 바로 찾습니다.
            card_label = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--card.inp-switch')) # 카드할인 토글 (배송비 포함 옆에)
            )

            # class 속성 가져오기
            card_label_classes = card_label.get_attribute("class")

            # is--on 클래스 포함 여부 확인
            if "is--on" in card_label_classes.split():

                # 카드할인 토글 요소가 화면에 보이도록 스크롤
                driver.execute_script("arguments[0].scrollIntoView(true);", card_label)

                # JavaScript로 강제로 클릭
                driver.execute_script("arguments[0].click();", card_label)
                print("카드할인 클릭")


            # 두 개의 클래스를 가진 label 요소를 바로 찾습니다.
            deli_label = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--delifee.inp-switch')) # 배송비포함 토글 (배송비 포함 옆에)
            )

            # class 속성 가져오기
            deli_label_classes = deli_label.get_attribute("class")

            # is--on 클래스 포함 여부 확인
            if "is--on" not in deli_label_classes.split():

                # 배송비포함 토글 요소가 화면에 보이도록 스크롤
                driver.execute_script("arguments[0].scrollIntoView(true);", deli_label)

                # JavaScript로 강제로 클릭
                driver.execute_script("arguments[0].click();", deli_label)
                print("배송비포함 클릭")

        except Exception as e:
            print(f"Error clicking the label with JavaScript: {e}")

    else:

        # 카드 할인 토글 클릭
        try:

            # 요소 찾기
            card_label = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--card.inp-switch'))
            )

            # class 속성 가져오기
            card_label_classes = card_label.get_attribute("class")

            # is--on 클래스 포함 여부 확인
            if "is--on" not in card_label_classes.split():

                # 카드할인 토글 요소가 화면에 보이도록 스크롤
                driver.execute_script("arguments[0].scrollIntoView(true);", card_label)

                # JavaScript로 강제로 클릭
                driver.execute_script("arguments[0].click();", card_label)
                print("카드할인 클릭")


            # 두 개의 클래스를 가진 label 요소를 바로 찾습니다.
            deli_label = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--delifee.inp-switch')) # 배송비포함 토글 (배송비 포함 옆에)
            )

            # class 속성 가져오기
            deli_label_classes = deli_label.get_attribute("class")

            # is--on 클래스 포함 여부 확인
            if "is--on" in deli_label_classes.split():

                # 배송비포함 토글 요소가 화면에 보이도록 스크롤
                driver.execute_script("arguments[0].scrollIntoView(true);", deli_label)

                # JavaScript로 강제로 클릭
                driver.execute_script("arguments[0].click();", deli_label)
                print("배송비포함 클릭")

        except Exception as e:
            print(f"Error clicking the label with JavaScript: {e}")




# 숫자를 추출할 때 빈 문자열을 처리하기 위한 함수
def extract_numeric_price(price_str):
    """ 가격 문자열에서 숫자만 추출하는 함수, '원' 앞에 있는 숫자만 추출 """
    try:
        # '원' 이전까지의 부분에서 숫자만 추출
        price_str = price_str.split('원')[0]
        numeric_part = ''.join(filter(str.isdigit, price_str))
        if numeric_part:
            return int(numeric_part)
        else:
            return None
    except Exception as e:
        print(f"Error extracting price from: {price_str}, {e}")
        return None


def mall_name_change(mall_name_origin):
    name_mapping = {
        "GS SHOP": "GS샵",
        # 다른 이름도 추가할 수 있음
    }
    return name_mapping.get(mall_name_origin, mall_name_origin) # 값이 없으면 원래 값 반환

# 숫자변환
def convert_to_float(value, default=1.0):
    """value에서 숫자와 소수점을 추출하여 float으로 변환, 변환할 수 없을 경우 기본값을 반환"""
    try:
        # 숫자와 소수점을 포함한 숫자 추출
        numeric_value = ''.join(re.findall(r'\d+\.\d+|\d+', value))
        return float(numeric_value) if numeric_value else default
    except (ValueError, AttributeError):
        return default  # 기본값을 반환


# 행별로 데이터를 처리하고 엑셀에 업데이트하는 함수
def save_row_to_excel(ws, merge_list, row_index, err_list, five_per_mall_name, open_market_list, producdt_count):
    global base_price
    try:
        # 기준가격(네이버) 계산 실제 사이트의 가격
        ws[f'E{row_index}'] = base_price

        # 5% 할인을 적용할 상점 이름과 비교
        for entry in merge_list:
            mall_name = entry[3]  # entry[3]은 mall_name을 의미
            if mall_name in five_per_mall_name:  # 배열 내에서 mall_name을 찾음
                price_value = entry[6]  # 이제 index 6은 숫자

                # 50000 미만일 경우에만 5% 할인 적용
                if price_value < 50000:
                    price_value *= 0.95  # 5% 할인 적용

                # 할인된 가격 저장
                entry[6] = price_value

        # 나머지 판매처 및 상품 업데이트 (정렬 후) [가격 / 수량] 해서 오름차순으로 정렬 맨위가 가격이 가장 쌈
        merge_list.sort(key=lambda x: x[6] / convert_to_float(x[2]) if x[6] else float('inf'))

        if len(merge_list) > 0:
            numeric_qty_0 = convert_to_float(merge_list[0][2])  # 숫자로 변환
            numeric_price_0 = merge_list[0][6]
            ws[f'F{row_index}'] = f"{merge_list[0][1]}-{merge_list[0][2]}-{merge_list[0][3]}"  # 판매처1
            ws[f'G{row_index}'] = merge_list[0][5]  # 상품명1
            ws[f'H{row_index}'] = int((numeric_price_0 / numeric_qty_0) * float(stnd_cnt))  # 가격1

        if len(merge_list) > 1 and producdt_count > 1:
            numeric_qty_1 = convert_to_float(merge_list[1][2])  # 숫자로 변환
            numeric_price_1 = merge_list[1][6]
            ws[f'I{row_index}'] = f"{merge_list[1][1]}-{merge_list[1][2]}-{merge_list[1][3]}"  # 판매처2
            ws[f'J{row_index}'] = merge_list[1][5]  # 상품명2
            ws[f'K{row_index}'] = int((numeric_price_1 / numeric_qty_1) * float(stnd_cnt)) # 가격2

        # 배경색 설정 (빨간색)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        set_cell_color(ws, 'B', row_index, err_list, 0, red_fill, white_fill)
        set_cell_color(ws, 'C', row_index, err_list, 1, red_fill, white_fill)

        print(f"Row {row_index} 엑셀에 업데이트됨.")

    except Exception as e:
        print(f"Error saving row {row_index} to Excel: {e}")
        traceback.print_exc()  # 에러 위치와 호출 스택을 출력

def set_cell_color(ws, cell_name, row_index, err_list, err_index, red_fill, white_fill):
    if err_list[err_index] == 1:
        ws[f'{cell_name}{row_index}'].fill = red_fill
    else:
        ws[f'{cell_name}{row_index}'].fill = white_fill


# 중복 체크 함수
def is_duplicate(entry, merge_list):
    for merge_entry in merge_list:
        if (merge_entry[2] == entry[2] and  # 구성 개수
                merge_entry[3] == entry[3] and  # 판매처
                merge_entry[5] == entry[5]):    # 상품명
            print(f'중복 발견: {merge_entry}')
            return True
    return False


# 크롤링 결과를 확인하고 에러 리스트에 반영하는 함수
def handle_scraping_result(result_list, index, err_list, merge_list):
    if len(result_list) == 0:
        err_list[index] = 1  # 크롤링 실패 시 해당 인덱스에 에러 표시
    else:
        add_non_duplicates(merge_list, result_list)


# 중복 체크 후 리스트에 추가하는 함수
def add_non_duplicates(merge_list, new_entries):
    for entry in new_entries:
        if not is_duplicate(entry, merge_list):
            merge_list.append(entry)



def result_print(result, name):
    for idx, obj in enumerate(result):
        print(f'{name} {idx + 1}: {obj[7]} {obj}')



def send_email(email_data):
    global email_list
    for recip_email in email_data['수신자이메일']:
        sender_email = email_data['발신자이메일']
        sender_password = email_data['발신자비밀번호']
        recipient_email = recip_email
        subject = email_data['제목']

        body = create_email_table(email_list)

        send_naver_email(sender_email, sender_password, recipient_email, subject, body, attachment_path=None)
        time.sleep(1)



def send_naver_email(sender_email, sender_password, recipient_email, subject, body, attachment_path=None):
    try:
        # SMTP 서버 설정 (네이버)
        smtp_server = "smtp.naver.com"
        smtp_port = 587

        # 이메일 메시지 생성
        message = MIMEMultipart()
        message["From"] = sender_email
        message["To"] = recipient_email
        message["Subject"] = subject

        # 본문 추가
        message.attach(MIMEText(body, "html"))

        # 첨부 파일 추가 (옵션)
        if attachment_path:
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    "Content-Disposition",
                    f"attachment; filename={attachment_path.split('/')[-1]}",
                )
                message.attach(part)

        # SMTP 서버에 연결
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # TLS 보안 활성화
        server.login(sender_email, sender_password)  # 로그인

        # 이메일 전송
        server.sendmail(sender_email, recipient_email, message.as_string())
        print("이메일이 성공적으로 전송되었습니다.")

        # 서버 종료
        server.quit()

    except Exception as e:
        print(f"이메일 전송 중 오류 발생: {e}")



def email_setting(ws, index, email_data):
    global email_list
    columns = [
        "수집제외몰", "다나와 URL", "에누리 URL", "상품명", "기준가격",  # 0, 1, 2, 3, 4, 5
        "판매처1", "상품명1", "가격1",  # 6, 7, 8
        "판매처2", "상품명2", "가격2",  # 9, 10, 11
        "키워드" # 12
    ]

    # 기본값 설정
    default_values = {
        "수집제외몰": "", "다나와 URL": "", "에누리 URL": "",
        "상품명": "", "기준가격": 0.0,
        "판매처1": "", "상품명1": "", "가격1": 0.0,
        "판매처2": "", "상품명2": "", "가격2": 0.0,
        "키워드": ""
    }

    # 특정 index의 행 가져오기 (Excel은 1-based index이므로 min_row와 조정)
    row = ws[index]

    # Cell 값 추출
    row_values = [cell.value for cell in row]

    # 행 데이터를 딕셔너리로 매핑
    row_object = {
        columns[i]: row_values[i] if i < len(row_values) and row_values[i] is not None else default_values[columns[i]]
        for i in range(len(columns))
    }

    # 판매처에 대해 처리
    for seller in email_data['판매처']:
        process_seller(index, row_object, seller, email_data, '판매처1', '상품명1', '가격1')


    # 중복 제거 (excel_row 기준)
    unique_rows = {}
    for item in email_list:
        if item["excel_row"] not in unique_rows:
            unique_rows[item["excel_row"]] = item

    # 결과 리스트로 변환
    email_list = list(unique_rows.values())

    # 전송 조건 확인
    if len(email_list) >= email_data['전송기준수']:
        print('전송가능한 데이터 수가 전송 기준수 보다 큽니다. 전송 시도 하겠습니다.')
        print(f'전송가능한 데이터 수 : {len(email_list)}')
        send_email(email_data)
        email_list = [] # 이메일 전송후 초기화
        print(f'전송 후 초기화 : {len(email_list)}')
    else:
        print('전송가능한 데이터 수가 전송 기준수 보다 작습니다. 다음에 전송 하겠습니다.')
        print(f'전송가능한 데이터 수 : {len(email_list)}')



def process_seller(index, row, seller, email_data, seller_key, product_key, price_key):
    global email_list, global_month_review_cnt

    keywords = row['키워드'].split(',') if row['키워드'] else []
    product_string = row[product_key]

    if seller in row[seller_key] or any(keyword in product_string for keyword in keywords):
        # 순수익 계산
        net_profit = row['기준가격'] - row[price_key] - (row['기준가격'] * (email_data['수수료율'] / 100)) - email_data['배송비']

        # 마진율 계산
        if row['기준가격'] > 0:
            net_profit_rate = (net_profit / row['기준가격']) * 100
            net_profit_rate = math.floor(net_profit_rate * 100) / 100  # 소수점 두 자리까지 내림
        else:
            net_profit_rate = 0.0  # 기준가격이 0일 때는 마진율 0으로 처리

        if email_data['마진율시작'] <= net_profit_rate <= email_data['마진율끝']:
            excel_row = index

            # URL 결정
            email_url = ''
            if '에누리' in row[seller_key]:
                email_url = row['에누리 URL']
            elif '다나와' in row[seller_key]:
                email_url = row['다나와 URL']

            # 엑셀 행 내용 생성
            email_content = {
                'excel_row': excel_row,
                'product_name': row['상품명'],
                'danawa_url': row['다나와 URL'],
                'product': row[product_key],
                'seller': row[seller_key],
                'net_profit_rate': net_profit_rate,
                'email_url': email_url
            }

            email_list.append(email_content)


# 이메일 테이블 형식으로
def create_email_table(email_list):

    table_rows = ""

    for row in email_list:
        table_rows += f"""
        <tr style="border: 1px solid #dddddd; text-align: left; padding: 12px;">
            <td style="border: 1px solid #dddddd; padding: 12px;">{row['excel_row']}</td>
            <td style="border: 1px solid #dddddd; padding: 12px;">{row['product_name']}</td>
            <td style="border: 1px solid #dddddd; padding: 12px;"><a href=\"{row['danawa_url']}\" target=\"_blank\">{row['danawa_url']}</a></td>
            <td style="border: 1px solid #dddddd; padding: 12px;">{row['product']}</td>
            <td style="border: 1px solid #dddddd; padding: 12px;">{row['seller']}</td>
            <td style="border: 1px solid #dddddd; padding: 12px;">{row['net_profit_rate']}%</td>
            <td style="border: 1px solid #dddddd; padding: 12px;"><a href=\"{row['email_url']}\" target=\"_blank\">{row['email_url']}</a></td>
        </tr>
        """

    email_content = f"""
    <html>
    <body>
        <p>안녕하세요, 상품 정보입니다:)</p>
        <br>
        <br>
        <table style="border-collapse: collapse; width: 100%; margin: 20px 0; font-size: 16px; font-family: Arial, sans-serif; box-shadow: 0 2px 5px rgba(0, 0, 0, 0.1);">
            <tr style="border: 1px solid #dddddd; background-color: #f4f4f4; font-weight: bold;">
                <th style="border: 1px solid #dddddd; padding: 12px;">엑셀 번호</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">상품명</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">다나와 URL</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">상품명1</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">판매처</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">마진율</th>
                <th style="border: 1px solid #dddddd; padding: 12px;">URL</th>
            </tr>
            {table_rows}
        </table>
    </body>
    </html>
    """
    return email_content


def previous_month_date(today):
    current_day = today.day
    current_month = today.month
    current_year = today.year

    if current_month == 1:
        previous_month = 12
        previous_year = current_year - 1
    else:
        previous_month = current_month - 1
        previous_year = current_year

    last_day_of_previous_month = calendar.monthrange(previous_year, previous_month)[1]

    start_date_of_previous_month = today.replace(
        year=previous_year,
        month=previous_month,
        day=min(current_day, last_day_of_previous_month)
    )

    return start_date_of_previous_month



def check_previous_month_add(today, start_date_of_previous_month, input_date):
    date_object = datetime.strptime(input_date.strip(), '%y.%m.%d.')
    if start_date_of_previous_month <= date_object <= today:
        return 1
    elif date_object > today:
        return 0
    else:
        return -1


# 메인 함수
def main(excel_path, limit_count, on_and_off, five_per_mall_name, start_row, end_row, repeat, email_data, open_market_list, producdt_count):
    global global_month_review_cnt
    # 엑셀 파일 열기 (openpyxl로 읽기)
    wb = load_workbook(excel_path)
    ws = wb.active

    driver = setup_driver()
    if not driver:
        print("Failed to initialize the web driver.")
        return

    if start_row < 2:
        start_row = 2

    if end_row > ws.max_row:
        end_row = ws.max_row


    while True:  # 무한 루프

        # 엑셀의 각 행을 처리
        for i in range(start_row, end_row + 1):
            global_month_review_cnt = 0
            if i > ws.max_row:
                print("값을 초과하였습니다.")
                break

            # 시작행 ~ 끝행 사이에서만 작업함
            try:
                name        = ws[f'D{i}'].value # 상품이름
                danawa_url  = ws[f'B{i}'].value # 다나와 URL
                enuri_url   = ws[f'C{i}'].value # 에누리 URL

                # 에러 리스트 초기화 (각 크롤링 사이트별 에러 체크)
                err_list = [0, 0]

                # 전체 병합 리스트 초기화
                merge_list = []

                # 2. 다나와 크롤링 처리
                print("============================== 다나와 시작 ==============================")
                danawa_result = scrape_danawa(driver, name, danawa_url, limit_count, on_and_off)
                sorted_merge_list = sorted(danawa_result, key=lambda x: x[-1])
                result_print(sorted_merge_list, '다나와')
                print(f'다나와 수: {len(danawa_result)}')
                print('===================================================')
                handle_scraping_result(danawa_result, 1, err_list, merge_list)
                print("============================== 다나와 끝 ==============================")
                print(f'\n\n')

                # 3. 에누리 크롤링 처리
                print("============================== 에누리 시작 ==============================")
                enuri_result = scrape_enuri(driver, name, enuri_url, limit_count, on_and_off)
                sorted_merge_list = sorted(enuri_result, key=lambda x: x[-1])
                result_print(sorted_merge_list, '에누리')
                print(f'에누리 수: {len(enuri_result)}')
                handle_scraping_result(enuri_result, 2, err_list, merge_list)
                print("============================== 에누리 끝 ==============================")
                print(f'\n\n')

                # 4. 전체 목록
                print("============================== 전체 시작 ==============================")
                # 가격 오름차순 정렬
                sorted_merge_list = sorted(merge_list, key=lambda x: x[-1])
                result_print(sorted_merge_list, '전체')
                print(f'전체 수: {len(merge_list)}')
                print("============================== 전체 끝 ==============================")

                # 엑셀로 저장 (각 행별로 저장)
                save_row_to_excel(ws, merge_list, i, err_list, five_per_mall_name, open_market_list, producdt_count)

                # 엑셀 파일을 즉시 저장
                wb.save(excel_path)

                # 이메일 발송
                email_setting(ws, i, email_data)

            except Exception as e:
                print(f"Error processing row {i}: {e}")
                traceback.print_exc()  # 스택 트레이스 출력
                # 특정 행에서 에러가 발생하더라도 계속 진행하도록 함


        # 무한 반복 여부 확인
        if not repeat:  # repeat가 False면 1회 반복 후 종료
            break

    driver.quit()


if __name__ == "__main__":

    # 원하는 값을 여기에 입력하세요.
    excel_path = "프로그램_테스트.xlsx"     # 파일 이름 (프로그램이 실행되는 경로에 파일이 있어야 합니다.)
    limit_count = 5                      # 사이트의 갯수별 수집 갯수
    producdt_count = 2                   # 판매처 갯수 1~3까지 입력가능
    on_and_off = 0                       # 1개 수집 : on (수집 변수 1) / off 미수집 변수 0 (기본 미수집 0)
    five_per_mall_name = ['11번가', '옥션']        # 5% 할인 적용 판매처 (옥션, G마켓, 11번가...)

    start_row = 2    # 첫 row값은 기본 2이상으로 설정 - 엑셀2번줄
    end_row = 100    # 실제 row수보다 작거나 같게 설정

    # repeat가 False면 1회 반복 후 종료 True면 무한반복
    repeat = False

    # 자사몰인 경우만 메일 발송 및 엑셀에 넣기
    # 수집 제외몰은 사용하지 않음
    # 자사몰은 처음 네이버에서 기준 가격할때만 사용
    open_market_list = ['옥션', # auction
                        '지마켓', #gmarket
                        'G마켓', #gmarket
                        '쿠팡', #coupang
                        '11번가', #11st
                        '인터파크', #interpark
                        '롯데온', #lotteon
                        '롯데ON', #lotteon
                        '티몬', #tmon
                        'GSSHOP', #gsshop
                        '오늘의집', #ohou
                        'SSG닷컴', #ssg.com
                        '이마트몰', #ssg.com
                        '신세계몰', #ssg.com
                        '롯데홈쇼핑', #lotteimall
                        '현대Hmall', #hmall
                        'CJ온스타일', #cjonstyle
                        '홈앤쇼핑', #hnsmall
                        '무신사', #musinsa
                        '컬리', #kurly
                        '멸치쇼핑', #smelchi
                        'NS홈쇼핑', #nsmall
                        '삼성닷컴', #samsung
                        '신세계라이브쇼핑', #shinsegaetvshopping
                        '하프클럽', #halfclub
                        '패션플러스', #fashionplus
                        'LFmall',
                        '현대Hmall'
                        ]

    # 이메일 설정
    email_data = {
        '수수료율': 1, #(단위 %)
        '배송비': 1,        #(단위 원)
        '판매처': ['G마켓', '이마트몰', '쿠팡', '11번가'], #(명확히 입력)
        # '판매처': [], #(명확히 입력)
        '마진율시작': 1,        #(단위 %)
        '마진율끝': 100,        #(단위 %)
        '전송기준수': 0,     #(단위 개 매진률수 이상이 되면 메일 발송)
        '발신자이메일': '',
        '발신자비밀번호': '',
        '수신자이메일': [],
        '제목': '특정 마진률 이상이면 메일 전송',
        '내용': '', # 엑셀행/ 상품명 / 네or다or에-N개-판매처 / 마진% / URL(네or다or에) 이 형식으로 바뀔것임 초기값은 공백
    }

    # 메인실행 함수
    main(
        excel_path,
        limit_count,
        on_and_off,
        five_per_mall_name,
        start_row,
        end_row,
        repeat,
        email_data,
        open_market_list,
        producdt_count
    )



# 변경이력 history

# 2024-09-11 ver_1

# 2024-09-12 ver_2
# 5% 할인 적용 판매처 (five_per_mall_name) 문자열 -> 배열로 변경
# 갯수없는 것 일반구매 해외구매 -> 1개로 수정
# [6] 숫자로만 수정


# 2024-10-01 ver_3


# 2024-10-04 ver_4
# 에누리 카드할인 반영 안됨 X -> 중복데이터 해결


# 2024-10-12 ver_5


# 2024-10-19 ver_5
# 기준가격 수정
# 단위수정


# 2025-01-12 ver_6
# 오류 1 (기준가격 수집 오류) - 네이버 class이름 태그 바뀜, 카드할인 부분 코드 제거
# 오류 2 (기준가격 미 수집 오류) / 빨간색으로 표시가 안 되어있는데 미수집
# 특정 마진률 이상이면 메일 전송되도록
# 개선사항2. 코딩상에서 엑셀 번호 적어서 거기서부터 돌아가도록
# 개선사항3. 프로그램이 끝나는게 아니라 계속 돌아가도록


# 2025-01-22 ver_7
# 마진율 사이값 10 ~ 50%
# star_row end_row -> end_row가 엑셀 row안넘기만 하면 되게 수정
# 상품명의 특정 키워드가 들어가면 판매처에 안 들어가 있어도, 다른 오픈마켓도 메일 발송하도록 "키워드" Q열에 적어주세요
# 전송규칙에 상품명 / 네이버 URL 추가
# 메일좀 예쁘게

# 2025-01-27 ver_8
# 자사몰 추가
# 기타 오류 수정


# 2025-02-03 ver_9
# 리플 수정
# 배송비포함/카드할인버튼 이슈
# 빨간색 표시 수정
# 중복전송 수정
# 기준가격 수정
# 자사몰 수정
# 네이버 복수


# 2025-02-03 ver_10
# 리플 딜레이 시간 추가
# 배송비포함/카드할인버튼 이슈 추가 수정


# 2025-02-24 ver_11
# 상품없음 수정
# 리다이렉트 수정
# 기준가격 수정


# 2027-07-27 ver_12
# 네이버 전체제거
# 다나와 배송비 클릭진행
# 다나와 배송비 4,000원 이상 제외, 다나와 상품명 '해외'들어가면 제외
# 다나와 기준가격 1개 상품에서 좌우 비교하여 가격 작은것


# 2027-11-01 ver_13
# 다나와 에누리 더보기 로그 수정
# 다나와 수정 다시 원복
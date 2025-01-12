from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
import time
import traceback
from selenium.common.exceptions import ElementNotInteractableException
import re

unit = '개'
stnd_cnt = '1'

# 드라이버 설정 함수
def setup_driver():
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--window-size=1080,750")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    try:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
            'source': '''
                Object.defineProperty(navigator, 'webdriver', {
                  get: () => undefined
                })
            '''
        })
    except Exception as e:
        print(f"Error setting up the driver: {e}")
        driver = None
    return driver


# 숫자와 소수점을 포함한 숫자만 추출하는 함수
def extract_numbers(input_string):
    try:
        return ''.join(re.findall(r'\d+\.\d+|\d+', input_string))
    except Exception as e:
        print(f"Error in extracting numbers from {input_string}: {e}")
        return ''  # 에러 발생 시 빈 문자열 반환

# 문자와 단위(g, kg 등)만 추출하는 함수
def extract_non_numbers(input_string):
    try:
        return ''.join(re.findall(r'[^\d\.]+', input_string))
    except Exception as e:
        print(f"Error in extracting non-numbers from {input_string}: {e}")
        return ''  # 에러 발생 시 빈 문자열 반환


# 네이버 크롤링 함수
def extract_product_info(idx, element, name, qty):
    """ 제품 정보를 추출하는 함수 """
    try:
        # 상점 이름 추출
        try:
            mall_name = element.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.mall"] img').get_attribute('alt')
        except NoSuchElementException:
            mall_name = element.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.mall"]').text.split('\n')[0]

        # 제품 이름 추출
        prod_name = element.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.pd"]').text

        # 가격 추출
        price_str = element.find_element(By.CSS_SELECTOR, '[data-shp-area="prc.price"]').text.replace('최저', '').strip()
        numeric_price = extract_numeric_price(price_str)  # 숫자만 추출한 가격

        # qty 값이 숫자를 포함하지 않으면 '1개' 할당
        if not any(char.isdigit() for char in qty):
            qt = f'{stnd_cnt}{unit}'
        else:
            qt = qty

        one_price = (numeric_price / convert_to_float(qt)) * float(stnd_cnt)

        temp_list = [name, '네이버', qt, mall_name, '', prod_name, numeric_price, one_price]

        print(f"네이버 제품 {idx}: {temp_list}")

        return temp_list
    except Exception as e:
        print(f"Error in extracting product info: {e}")
        return None  # 에러 발생 시 None 반환


def extract_ul_class(driver):
    """ ul 클래스 이름 추출 함수 """
    try:
        uls = driver.find_elements(By.CSS_SELECTOR, '#section_price ul')  # ul 목록
        for e in uls:  # ul 안에 li class 이름 가져오기
            if 'productList_list_seller' in e.get_attribute('class'):
                return e.get_attribute('class').replace(' ', '.')
    except Exception as e:
        print(f"Error in extracting ul class: {e}")
    return ""


def process_product_list(driver, ul_class, name, qty, naver_temp_list):
    """ 제품 목록 처리 함수 """
    try:
        prod_list = driver.find_elements(By.CSS_SELECTOR, f'#section_price .{ul_class} li')

        # 첫 번째 제품 목록으로 스크롤
        action = ActionChains(driver)
        action.move_to_element(prod_list[0]).perform()
        time.sleep(1)

        # 공통 함수 호출하여 제품 정보 추출
        for idx, e in enumerate(prod_list):

            # 기준인건 1개까지
            if qty == f'{stnd_cnt}{unit}' and idx > 0:
                break

            # 기준이 아닌건 3개까지
            if qty != f'{stnd_cnt}{unit}' and idx > 2:
                break

            temp_list = extract_product_info(idx, e, name, qty)
            if temp_list:
                naver_temp_list.append(temp_list)
            else:
                print(f"Error in scraping 네이버: index {idx}")
    except Exception as e:
        print(f"Error in processing product list: {e}")


def scrape_naver(driver, name, naver_url):
    global unit, stnd_cnt
    try:
        if not naver_url:
            return []
        driver.get(naver_url)
        print(naver_url)
        naver_temp_list = []
        time.sleep(3)


        # 카드할인 토글 클릭 ON으로 둘 다 변경 (위, 아래 있음)
        discount_elements = driver.find_elements(By.CSS_SELECTOR, '[data-shp-contents-type="카드할인가 정렬"]')
        if discount_elements:
            discount_elements[0].click()  # 카드할인 클릭
            time.sleep(0.5)
        else:
            print("카드할인가 정렬 옵션을 찾을 수 없습니다. 중지합니다.")
            return []


        # 옵션 이름 ex) 수량, 개수, 상품구성 등등 / 개당 중량, 수량 : 2개입, 1개
        opt_name = driver.execute_script('return document.querySelector("#section_price em").closest("div");').text.split(" : ")[0].split(',')[-1]
        print(opt_name)

        # 상품구성: 1개, 2개, 3개 등 옵션 처리
        qtys = []
        if len(driver.find_elements(By.CSS_SELECTOR, f'.condition_area a[data-shp-contents-type="{opt_name}"] .info')) != 0:
            qtys = driver.find_elements(By.CSS_SELECTOR, f'.condition_area a[data-shp-contents-type="{opt_name}"] .info')
        elif len(driver.find_elements(By.CSS_SELECTOR, '.condition_area a .info')) != 0:
            qtys = driver.find_elements(By.CSS_SELECTOR, '.condition_area a .info')  # 수량, 개수 옵션이 없다면 2번째 옵션으로 지정

        ul_class = extract_ul_class(driver)  # ul 클래스 추출

        if len(qtys) > 0:
            # ['1개', '2개', '3개', '4개', '5개']
            qlist = [q.text for q in qtys]
            print(qlist)

            unit = extract_non_numbers(qlist[0])
            print(f'단위 : {unit}')

            stnd_cnt = extract_numbers(qlist[0])
            print(f'기준 수량 : {stnd_cnt}')

            delivery_option = 0

            for p in range(len(qtys)):
                print(qlist[p])
                driver.find_element(By.CSS_SELECTOR, f'[data-shp-contents-id="{qlist[p]}"]').click()
                time.sleep(2)
                driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                driver.execute_script("window.scrollTo(0, 0);")


                if qlist[p] != f'{stnd_cnt}{unit}' and delivery_option == 0:
                    delivery_elements = driver.find_elements(By.CSS_SELECTOR, '[data-shp-contents-type="배송비포함 필터"]')
                    if delivery_elements:
                        delivery_elements[0].click()  # 배송비포함 클릭
                        time.sleep(0.5)
                        delivery_option = 1
                    else:
                        print("카드할인가 정렬 옵션을 찾을 수 없습니다. 중지합니다.")
                        return []

                if not ul_class:
                    continue

                process_product_list(driver, ul_class, name, qlist[p], naver_temp_list)

                # 1개는 1개 , 1개 이상은 3개로 제한을 두었기 때문에 페이징 처리 필요 없음
                # try:
                #     # 페이지네이션 처리
                #     pagination_wrap = driver.find_element(By.CLASS_NAME, 'productList_seller_wrap__FZtUS')
                #     pagination = pagination_wrap.find_element(By.CLASS_NAME, 'pagination_pagination__JW7zT')
                #     page_links = pagination.find_elements(By.TAG_NAME, 'a')
                #
                #     # 페이지가 있는 경우 처리
                #     for page_link in page_links:
                #         page_link.click()  # 페이지 클릭
                #         time.sleep(2)  # 페이지 로드 대기
                #         # 제품 목록 처리
                #         process_product_list(driver, ul_class, name, qlist[p], naver_temp_list)
                #
                # except Exception as e:
                #     # 페이지네이션이 없는 경우 예외 처리
                #     print(f"Pagination not found for item {qlist[p]}. Processing product list without pagination.")
                #     process_product_list(driver, ul_class, name, qlist[p], naver_temp_list)

        else:
            # 수량 옵션이 없는 경우 제품 목록 처리
            process_product_list(driver, ul_class, name, f'{stnd_cnt}{unit}', naver_temp_list)

        return naver_temp_list

    except (NoSuchElementException, TimeoutException) as e:
        print(f"Error in Naver scraping: {e}")
        return []


# 다나와 크롤링 함수
def scrape_danawa(driver, name, danawa_url, limit_count, on_and_off):
    try:
        if not danawa_url:
            return []
        driver.get(danawa_url)
        print(danawa_url)
        danawa_temp_list = []

        # 구성 상품 열기 클릭 시 에러 발생 시 빈 리스트 반환
        try:
            if driver.find_elements(By.XPATH, '//*[@id="bundleProductMoreOpen"]'):
                driver.find_element(By.XPATH, '//*[@id="bundleProductMoreOpen"]').click()
                print("구성 상품을 성공적으로 열었습니다.")
            else:
                print("구성 상품을 찾을 수 없습니다. 계속 진행합니다.")
        except ElementNotInteractableException as e:
            # 버튼이 있지만 상호작용이 불가능한 경우에만 넘어감
            print(f"구성 상품 열기 버튼이 있지만 상호작용이 불가능하여 넘어갑니다: {e}")
            pass  # 다음 라인으로 넘어가서 계속 진행
        except Exception as e:
            # 다른 예외는 발생 시 빈 리스트 반환
            print(f"Error occurred while trying to open 구성 상품: {e}")
            return []


        danawa_opt_url_list = [e.get_attribute('href') for e in driver.find_elements(By.CSS_SELECTOR, '[class="othr_list"] li .chk a')]
        danawa_opt_text_list = [e.text for e in driver.find_elements(By.CSS_SELECTOR, '[class="othr_list"] li .chk a')]
        # ['1개', '2개', '3개', '4개', '5개'] #['https://prod.danawa.com/info/?pcode=5970722', 'https://prod.danawa.com/info/?pcode=5970724', 'https://prod.danawa.com/info/?pcode=5970731', 'https://prod.danawa.com/info/?pcode=5970748', 'https://prod.danawa.com/info/?pcode=5970738']

        print(danawa_opt_url_list)
        print(danawa_opt_text_list)

        # 수량옵션이 없는경우 1개로 처리하기 위한 세팅
        if not danawa_opt_url_list:
            danawa_opt_url_list = [1]
            danawa_opt_text_list = [1]

        for ii in range(len(danawa_opt_url_list)):
            print(danawa_opt_url_list[ii])
            print(danawa_opt_text_list[ii])

            if danawa_opt_url_list[ii] != 1:
                if on_and_off == 0 and danawa_opt_text_list[ii] == f'{stnd_cnt}{unit}':  # 1개는 스킵 # 복수 구성이 없는 상품은 에러처리안나게 건너 뛰기
                    continue

                driver.get(danawa_opt_url_list[ii])
                time.sleep(2)

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            driver.execute_script("window.scrollTo(0, 0);")
            driver.find_elements(By.CSS_SELECTOR, '.cardSaleChkbox')[0].click()  # 카드할인가 클릭
            time.sleep(1)

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 좌측: 무료배송
            free_dil_prod_e_list = soup.select('.columm.left_col .diff_item')
            time.sleep(1)

            for ei in range(min(len(free_dil_prod_e_list), limit_count)):
                try:
                    try:
                        mall_name = soup.select('.columm.left_col .diff_item')[ei].select('img')[0].get('alt')
                    except:
                        mall_name = soup.select('.columm.left_col .diff_item')[ei].select('a .txt_logo')[0].text

                    prod_name = soup.select('.columm.left_col .diff_item')[ei].select('.info_line')[0].text.strip()

                    # card_line이 존재하는지 확인하고 적절한 price_str을 설정
                    card_line = soup.select('.columm.left_col .diff_item')[ei].select('.card_line')
                    if card_line:
                        price_str = card_line[0].select('.prc_t')[0].text
                    else:
                        price_str = soup.select('.columm.left_col .diff_item')[ei].select('.prc_line')[0].select('.prc_c')[0].text

                    # 숫자만 추출한 가격
                    numeric_price = extract_numeric_price(price_str)


                    # danawa_opt_text_list[ii] 값이 숫자를 포함하지 않으면 '1개' 할당
                    # (일반구매, 해외구매) case
                    qty = f'{stnd_cnt}{unit}'

                    if danawa_opt_text_list[ii] != 1:
                        if not any(char.isdigit() for char in danawa_opt_text_list[ii]):
                            qty = f'{stnd_cnt}{unit}'
                        else:
                            qty = danawa_opt_text_list[ii]

                    one_price = (numeric_price / convert_to_float(qty)) * float(stnd_cnt)

                    temp_list = [name, '다나와', qty, mall_name, '무료배송', prod_name, numeric_price, one_price]
                    print(f"다나와 {ei} : {temp_list}")
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (free shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

            # 우측: 유/무료 배송
            pay_dil_prod_e_list = soup.select('.columm.rgt_col .diff_item')
            time.sleep(1)

            for ei in range(min(len(pay_dil_prod_e_list), limit_count)):
                try:
                    try:
                        mall_name = soup.select('.columm.rgt_col .diff_item')[ei].select('img')[0].get('alt')
                    except:
                        mall_name = soup.select('.columm.rgt_col .diff_item')[ei].select('a .txt_logo')[0].text

                    prod_name = soup.select('.columm.rgt_col .diff_item')[ei].select('.info_line')[0].text.strip()
                    price_str = soup.select('.columm.rgt_col .diff_item')[ei].select('.prc_c')[0].text
                    numeric_price = extract_numeric_price(price_str)  # 숫자만 추출한 가격

                    qty = f'{stnd_cnt}{unit}'
                    if danawa_opt_text_list[ii] != 1:
                        # (일반구매, 해외구매) case
                        if not any(char.isdigit() for char in danawa_opt_text_list[ii]):
                            qty = f'{stnd_cnt}{unit}'
                        else:
                            qty = danawa_opt_text_list[ii]

                    one_price = (numeric_price / convert_to_float(qty)) * float(stnd_cnt)

                    temp_list = [name, '다나와', qty, mall_name, '유/무료배송', prod_name, numeric_price, one_price]
                    print(f"다나와 {ei} : {temp_list}")
                    danawa_temp_list.append(temp_list)
                except Exception as e:
                    print(f"Error in scraping Danawa (paid shipping): {e}")
                    continue  # 에러 발생 시 다음 루프 항목으로 이동

        return danawa_temp_list

    except Exception as e:
        print(f"Error in Danawa scraping: {e}")
        return []


# 에누리 크롤링 함수
def scrape_enuri(driver, name, enuri_url, limit_count, on_and_off):
    try:
        # 테스트를 위한 에러 발생
        # raise Exception("테스트를 위한 에러 발생")
        merge_list = []
        if not enuri_url:
            return []
        print(f"enuri_url : {enuri_url}")

        driver.get(enuri_url)
        time.sleep(0.5)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # SCroll 맨 아래로
        driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로
        time.sleep(0.5)

        # 가격비교 구매옵션의 라디오: 1개 2개 3개 4개 5개 라디오 아래 더보기 버튼
        try:
            if len(driver.find_elements(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more')) != 0:
                driver.find_element(By.CSS_SELECTOR, '#prod_option .adv-search__btn--more').click()
                print("더보기 버튼을 성공적으로 클릭했습니다.")
            else:
                print("더보기 버튼을 찾을 수 없습니다. 계속 진행합니다.")
        except ElementNotInteractableException as e:
            # 더보기 버튼이 있지만 상호작용이 불가능한 경우에만 넘어감
            print(f"더보기 버튼이 있지만 상호작용이 불가능하여 넘어갑니다: {e}")
            pass  # 다음 라인으로 넘어가서 계속 진행
        except Exception as e:
            # 다른 예외는 발생 시 빈 리스트 반환
            print(f"Error occurred while trying to click 더보기 버튼: {e}")
            return []


        time.sleep(1)
        # 가격비교 구매옵션의 라디오 : 1개 2개 3개 4개 5개 라디오 아래 더보기 버튼
        radio_opts = driver.find_elements(By.CSS_SELECTOR, '[name="radioOPTION"]')
        xpath_list = ['//*[@for="' + e.get_attribute('id') + '"]' for e in radio_opts] # 라디오옵션을 감싸고 있는 label
        time.sleep(1)

        # 카드 할인 토글 클릭
        try:
            # 두 개의 클래스를 가진 label 요소를 바로 찾습니다.
            label = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'label.model__cb--card.inp-switch')) # 카드할인 토글 (배송비 포함 옆에)
            )

            # 카드할인 토글 요소가 화면에 보이도록 스크롤
            driver.execute_script("arguments[0].scrollIntoView(true);", label)
            time.sleep(1)  # 스크롤 후 잠시 대기

            # JavaScript로 강제로 클릭
            driver.execute_script("arguments[0].click();", label)
            print("카드할인 클릭")

        except Exception as e:
            print(f"Error clicking the label with JavaScript: {e}")


        # 수량옵션이 없는경우 1개로 처리하기 위한 세팅
        if not radio_opts:
            radio_opts = [1]
            xpath_list = [1]

        list_names = []

        for eei, e in enumerate(radio_opts):
            if xpath_list[eei] != 1:
                elem = driver.find_element(By.XPATH, xpath_list[eei])
                elem_text = elem.text.split()

                if elem_text:
                    how_many = elem_text[0]
                    list_names.append(how_many)  # 리스트에 추가
                else:
                    print(f"Warning: No text found for element at index {eei}")
            else:
                list_names = [f'{stnd_cnt}{unit}']

        # 리스트를 쉼표로 구분한 문자열로 결합
        list_names_str = ', '.join(list_names)
        print(f'{list_names_str}')


        for eei, e in enumerate(radio_opts):

            time.sleep(0.5)
            driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로

            how_many = f'{stnd_cnt}{unit}'

            if xpath_list[eei] != 1:

                elem = driver.find_element(By.XPATH, xpath_list[eei])  # 갯수 엘리먼트

                how_many = elem.text.split()[0]  # 갯수 텍스트

                print(how_many)
                if on_and_off == 0 and how_many == f'{stnd_cnt}{unit}':  # 1개는 스킵
                    continue

                elem.click()  # 갯수 클릭


            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") # SCroll 맨 아래로
            driver.execute_script("window.scrollTo(0, 0);") # SCroll 맨 위로
            time.sleep(1)

            max_wait_time = 10  # 최대 대기 시간 (초)
            start_time = time.time()

            while True:
                try:
                    # 로딩 상태가 아직 보이는지 확인
                    if driver.find_element(By.XPATH, '//*[@class="comm-loader"]').get_attribute('style') != "display: none;":
                        time.sleep(0.5)

                        # 최대 대기 시간이 넘으면 루프 탈출
                        if time.time() - start_time > max_wait_time:
                            print("Loader is taking too long. Skipping this step.")
                            break
                    else:
                        # 로딩이 끝났으면 루프 탈출
                        break
                except Exception as e:
                    # 에러가 발생했을 경우 에러 메시지 출력 후 루프 탈출
                    print(f"Error while waiting for loader: {e}")
                    print(traceback.format_exc())
                    break

            html = driver.page_source
            soup = BeautifulSoup(html, 'html.parser')

            # 배송비 포함, 카드할인 아래 이미지, 쇼핑몰, 상품명, 배송비, 판매가, 무이자 할부...  테이블 아래 list tr:  'tb-compare__list' 클래스 내부의 tbody 태그에서 모든 tr 태그를 찾음
            free_dil_prod_e_list = soup.select('.tb-compare__list tbody tr[data-plno]')

            for ei in range(min(len(free_dil_prod_e_list), limit_count)):
                try:
                    # 판매처 추출
                    mall_name = ""
                    shop_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--shop')
                    if shop_td:
                        img_tag = shop_td.find('img')
                        mall_name_origin = img_tag['alt'] if img_tag and 'alt' in img_tag.attrs else shop_td.get_text(strip=True)
                        mall_name = mall_name_change(mall_name_origin)

                    # 제품명 추출
                    prod_name = ""
                    name_td = free_dil_prod_e_list[ei].find('td', class_='tb-col--name')
                    if name_td:
                        name_div = name_td.find('div', class_='tb-col__tx--name')
                        if name_div:
                            # <i> 태그 제거 후 텍스트만 추출
                            [i_tag.extract() for i_tag in name_div.find_all('i')]
                            prod_name = name_div.get_text(strip=True)

                    # 가격 추출
                    price_str = ""
                    price_div = free_dil_prod_e_list[ei].find('div', class_='tx-price')
                    if price_div:
                        em_tag = price_div.find('em')
                        price_str = em_tag.get_text(strip=True) if em_tag else ""

                    # 가격 추출 및 숫자만 저장
                    numeric_price = extract_numeric_price(price_str)  # 숫자만 추출한 가격


                    # how_many 값이 숫자를 포함하지 않으면 '1개' 할당
                    # (일반구매, 해외구매) case
                    if not any(char.isdigit() for char in how_many):
                        qty = f'{stnd_cnt}{unit}'
                    else:
                        qty = how_many

                    # 리스트에 데이터 추가
                    one_price = (numeric_price / convert_to_float(qty)) * float(stnd_cnt)

                    temp_list = [name, '에누리', qty, mall_name, '무료배송', prod_name, numeric_price, one_price]
                    print(f"에누리 {ei} : {temp_list}")
                    merge_list.append(temp_list)

                except Exception as e:
                    print(f"Error in scraping Enuri free shipping list (index {ei}): {e}")
                    continue

        return merge_list
    except Exception as e:
        print(f"Error in Enuri scraping: {e}")
        return []


# 숫자를 추출할 때 빈 문자열을 처리하기 위한 함수
def extract_numeric_price(price_str):
    """ 가격 문자열에서 숫자만 추출하는 함수, '원' 앞에 있는 숫자만 추출 """
    try:
        # '원' 이전까지의 부분에서 숫자만 추출
        price_str = price_str.split('원')[0]
        numeric_part = ''.join(filter(str.isdigit, price_str))
        if numeric_part:
            return int(numeric_part)
        else:
            return None
    except Exception as e:
        print(f"Error extracting price from: {price_str}, {e}")
        return None


def mall_name_change(mall_name_origin):
    name_mapping = {
        "GS SHOP": "GS샵",
        # 다른 이름도 추가할 수 있음
    }
    return name_mapping.get(mall_name_origin, mall_name_origin) # 값이 없으면 원래 값 반환

# 숫자변환
def convert_to_float(value, default=1.0):
    """value에서 숫자와 소수점을 추출하여 float으로 변환, 변환할 수 없을 경우 기본값을 반환"""
    try:
        # 숫자와 소수점을 포함한 숫자 추출
        numeric_value = ''.join(re.findall(r'\d+\.\d+|\d+', value))
        return float(numeric_value) if numeric_value else default
    except (ValueError, AttributeError):
        return default  # 기본값을 반환


# 행별로 데이터를 처리하고 엑셀에 업데이트하는 함수
def save_row_to_excel(ws, merge_list, row_index, err_list, five_per_mall_name):
    try:
        except_list = ws[f'A{row_index}'].value.split(',') if ws[f'A{row_index}'].value else []

        # 기준가격(네이버) 계산
        # 이 코드에서는 next()를 사용하여 조건에 맞는 첫 번째 항목을 찾아 filtered_naver에 저장하고, 값이 존재하면 계산하여 ws에 저장합니다.
        filtered_naver = next((entry for entry in merge_list if entry[1] == '네이버' and entry[3]), None)
        if filtered_naver:
            naver_price = filtered_naver[6] / convert_to_float(filtered_naver[2])  # 숫자로 변환, 실패 시 기본값 사용
            ws[f'G{row_index}'] = naver_price * int(stnd_cnt)  # 기준가격(네이버) 셀

        # except_list에 포함된 mall_name 제외
        merge_list = [entry for entry in merge_list if entry[3] not in except_list]

        # 5% 할인을 적용할 상점 이름과 비교
        for entry in merge_list:
            mall_name = entry[3]  # entry[3]은 mall_name을 의미
            if mall_name in five_per_mall_name:  # 배열 내에서 mall_name을 찾음
                price_value = entry[6]  # 이제 index 6은 숫자

                # 50000 미만일 경우에만 5% 할인 적용
                if price_value < 50000:
                    price_value *= 0.95  # 5% 할인 적용

                # 할인된 가격 저장
                entry[6] = price_value

        # 나머지 판매처 및 상품 업데이트 (정렬 후) [가격 / 수량] 해서 오름차순으로 정렬 맨위가 가격이 가장 쌈
        merge_list.sort(key=lambda x: x[6] / convert_to_float(x[2]) if x[6] else float('inf'))

        # 1개짜리 수집인 경우 네이버는 제외 2024-10-01
        filtered_list = [
            item for item in merge_list
            if not (item[1] == '네이버' and convert_to_float(item[2]) == 1)
        ]

        if len(filtered_list) > 0:
            numeric_qty_0 = convert_to_float(filtered_list[0][2])  # 숫자로 변환
            numeric_price_0 = filtered_list[0][6]
            ws[f'H{row_index}'] = f"{filtered_list[0][1]}-{filtered_list[0][2]}-{filtered_list[0][3]}"  # 판매처1
            ws[f'I{row_index}'] = filtered_list[0][5]  # 상품명1
            ws[f'J{row_index}'] = (numeric_price_0 / numeric_qty_0) * float(stnd_cnt)  # 가격1

        if len(filtered_list) > 1:
            numeric_price_1 = filtered_list[1][6]
            numeric_qty_1 = convert_to_float(filtered_list[1][2])  # 숫자로 변환
            ws[f'K{row_index}'] = f"{filtered_list[1][1]}-{filtered_list[1][2]}-{filtered_list[1][3]}"  # 판매처2
            ws[f'L{row_index}'] = filtered_list[1][5]  # 상품명2
            ws[f'M{row_index}'] = (numeric_price_1 / numeric_qty_1) * float(stnd_cnt) # 가격2

        if len(filtered_list) > 2:
            numeric_price_2 = filtered_list[2][6]
            numeric_qty_2 = convert_to_float(filtered_list[2][2])  # 숫자로 변환
            ws[f'N{row_index}'] = f"{filtered_list[2][1]}-{filtered_list[2][2]}-{filtered_list[2][3]}"  # 판매처3
            ws[f'O{row_index}'] = filtered_list[2][5]  # 상품명3
            ws[f'P{row_index}'] = (numeric_price_2 / numeric_qty_2) * float(stnd_cnt)  # 가격3

        # 배경색 설정 (빨간색)
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # err_list 값에 따라 배경색 변경
        if err_list[0] == 1:
            ws[f'B{row_index}'].fill = red_fill  # "네이버 URL" 셀
        if err_list[1] == 1:
            ws[f'C{row_index}'].fill = red_fill  # "다나와 URL" 셀
        if err_list[2] == 1:
            ws[f'D{row_index}'].fill = red_fill  # "에누리 URL" 셀

        print(f"Row {row_index} 엑셀에 업데이트됨.")

    except Exception as e:
        print(f"Error saving row {row_index} to Excel:")
        traceback.print_exc()  # 에러 위치와 호출 스택을 출력


# 중복 체크 함수
def is_duplicate(entry, merge_list):
    for merge_entry in merge_list:
        if (merge_entry[2] == entry[2] and  # 구성 개수
                merge_entry[3] == entry[3] and  # 판매처
                merge_entry[5] == entry[5]):    # 상품명
            print(f'중복 발견: {merge_entry}')
            return True
    return False


# 크롤링 결과를 확인하고 에러 리스트에 반영하는 함수
def handle_scraping_result(result_list, index, err_list, merge_list):
    if len(result_list) == 0:
        err_list[index] = 1  # 크롤링 실패 시 해당 인덱스에 에러 표시
    else:
        add_non_duplicates(merge_list, result_list)


# 중복 체크 후 리스트에 추가하는 함수
def add_non_duplicates(merge_list, new_entries):
    for entry in new_entries:
        if not is_duplicate(entry, merge_list):
            merge_list.append(entry)


def result_print(result, name):
    for idx, obj in enumerate(result):
        print(f'{name} {idx + 1}: {obj[7]} {obj}')


# 메인 함수
def main(excel_path, limit_count, on_and_off, five_per_mall_name):
    # 엑셀 파일 열기 (openpyxl로 읽기)
    wb = load_workbook(excel_path)
    ws = wb.active

    driver = setup_driver()
    if not driver:
        print("Failed to initialize the web driver.")
        return

    # 엑셀의 각 행을 처리
    for i in range(2, ws.max_row + 1):
        try:
            name = ws[f'E{i}'].value
            naver_url = ws[f'B{i}'].value
            danawa_url = ws[f'C{i}'].value
            enuri_url = ws[f'D{i}'].value

            # 에러 리스트 초기화 (각 크롤링 사이트별 에러 체크)
            err_list = [0, 0, 0]

            # 전체 병합 리스트 초기화
            merge_list = []

            # 1. 네이버 크롤링 처리
            print("============================== 네이버 시작 ==============================")
            naver_result = scrape_naver(driver, name, naver_url)
            sorted_merge_list = sorted(naver_result, key=lambda x: x[-1])
            result_print(sorted_merge_list, '네이버')
            print(f'네이버 수: {len(naver_result)}')
            print("============================== 네이버 끝 ==============================")
            handle_scraping_result(naver_result, 0, err_list, merge_list)
            print(f'\n\n')

            # 2. 다나와 크롤링 처리
            print("============================== 다나와 시작 ==============================")
            danawa_result = scrape_danawa(driver, name, danawa_url, limit_count, on_and_off)
            sorted_merge_list = sorted(danawa_result, key=lambda x: x[-1])
            result_print(sorted_merge_list, '다나와')
            print(f'다나와 수: {len(danawa_result)}')
            print('===================================================')
            handle_scraping_result(danawa_result, 1, err_list, merge_list)
            print("============================== 다나와 끝 ==============================")
            print(f'\n\n')

            # 3. 에누리 크롤링 처리
            print("============================== 에누리 시작 ==============================")
            enuri_result = scrape_enuri(driver, name, enuri_url, limit_count, on_and_off)
            sorted_merge_list = sorted(enuri_result, key=lambda x: x[-1])
            result_print(sorted_merge_list, '에누리')
            print(f'에누리 수: {len(enuri_result)}')
            handle_scraping_result(enuri_result, 2, err_list, merge_list)
            print("============================== 에누리 끝 ==============================")
            print(f'\n\n')

            # 4. 전체 목록
            print("============================== 전체 시작 ==============================")
            sorted_merge_list = sorted(merge_list, key=lambda x: x[-1])
            result_print(sorted_merge_list, '전체')
            print(f'전체 수: {len(merge_list)}')
            print("============================== 전체 끝 ==============================")

            # 엑셀로 저장 (각 행별로 저장)
            save_row_to_excel(ws, merge_list, i, err_list, five_per_mall_name)

            # 엑셀 파일을 즉시 저장
            wb.save(excel_path)

        except Exception as e:
            print(f"Error processing row {i}: {e}")
            # 특정 행에서 에러가 발생하더라도 계속 진행하도록 함

    driver.quit()


if __name__ == "__main__":

    # 원하는 값을 여기에 입력하세요.
    excel_path = "프로그램.xlsx"    # 파일 이름 (프로그램이 실행되는 경로에 파일이 있어야 합니다.)
    limit_count = 3                    # 수집 갯수
    on_and_off = 0                     # 1개 수집 : on (수집 변수 1) / off 미수집 변수 0 (기본 미수집 0)
    five_per_mall_name = ['11번가', '옥션']        # 5% 할인 적용 판매처 (옥션, G마켓, 11번가...)

    # 메인실행 함수
    main(
        excel_path,
        limit_count,
        on_and_off,
        five_per_mall_name
    )



# 변경이력 history

# 2024-09-11 ver_1

# 2024-09-12 ver_2
# 5% 할인 적용 판매처 (five_per_mall_name) 문자열 -> 배열로 변경
# 갯수없는 것 일반구매 해외구매 -> 1개로 수정
# [6] 숫자로만 수정


# 2024-10-01 ver_3


# 2024-10-04 ver_4
# 에누리 카드할인 반영 안됨 X -> 중복데이터 해결


# 2024-10-12 ver_5


# 2024-10-19 ver_5
# 기준가격 수정
# 단위수정
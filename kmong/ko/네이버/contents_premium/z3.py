import os
import time

import pdfkit
import psutil
import pyautogui
import pyperclip
from docx import Document
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


def close_chrome_processes():
    """모든 Chrome 프로세스를 종료합니다."""
    for proc in psutil.process_iter(['pid', 'name']):
        try:
            if 'chrome' in proc.info['name'].lower():
                proc.kill()  # Chrome 프로세스를 종료
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess):
            pass

def setup_driver():
    try:
        close_chrome_processes()

        chrome_options = Options()
        user_data_dir = f"C:\\Users\\{os.getlogin()}\\AppData\\Local\\Google\\Chrome\\User Data"
        profile = "Default"

        chrome_options.add_argument(f"user-data-dir={user_data_dir}")
        chrome_options.add_argument(f"profile-directory={profile}")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-software-rasterizer")
        chrome_options.add_argument("--window-size=960,720")

        user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        chrome_options.add_argument(f'user-agent={user_agent}')
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

        script = '''
                Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                window.navigator.chrome = { runtime: {} };
                Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
                Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                Object.defineProperty(navigator, 'userAgent', { get: () => 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36' });
            '''
        driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {'source': script})

        return driver
    except WebDriverException as e:
        print(f"Error setting up the WebDriver: {e}")
        return None


if __name__ == "__main__":
    excel_path = "비플랜 목차 테스트.xlsx"
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"엑셀 파일이 존재하지 않습니다: {excel_path}")

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    urls = [row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)]

    driver = setup_driver()

    for url in urls:
        try:
            driver.get(url)
            time.sleep(3)  # 페이지 로드 대기

            # 뷰티풀숲으로 HTML 가져오기
            html = driver.page_source
            soup = BeautifulSoup(html, "html.parser")

            # 특정 요소 제거
            selectors_to_remove = [
                "div.container_aside._CONTAINER_ASIDE",
                "header.header_wrap",
                "footer[role='contentinfo']",
                "div._GRID_TEMPLATE_COLUMN_OUTSIDE",
                "div.comment_u_cbox_wrap",
                "div.viewer_bottom_section",
            ]
            for selector in selectors_to_remove:
                for element in soup.select(selector):
                    element.decompose()

            # 남은 HTML을 PDF로 저장
            cleaned_html = soup.prettify()
            output_pdf = f"{url.split('/')[-1]}.pdf"
            pdfkit.from_string(cleaned_html, output_pdf)
            print(f"PDF 파일 생성: {output_pdf}")



        except Exception as e:
            print(f"Error processing URL {url}: {e}")

    driver.quit()

from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import ttk, filedialog, font, messagebox
import tkinter as tk
import pandas as pd
import threading
import requests
import random
import ctypes
import time
import json
import math
import re
from bs4 import BeautifulSoup
from urllib.parse import quote
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font


# region
# ══════════════════════════════════════════════════════
# 전역 변수
# ══════════════════════════════════════════════════════

# 엑셀 업로드 후 url 리스트
id_list = []

# 추출값 페센트 리스트
extracted_data_list = []  # 모든 데이터 저장용

# 일시 중지
stop_flag = False

# 네이버 로그인 쿠키
global_naver_cookies = None

# 완료후 깜빡임 상태
flashing = True

# 엑셀 경로
filepath = None

# 셀렉트 초기값을 1로 설정
selected_value = 5
selected_cont_value = 30
time_sleep_val = 1

bearer_token = ""
refresh_token = ""

global_server_cookies = {}  # 다른 서버 로그인 쿠키를 저장
URL = "http://vjrvj.cafe24.com"
login_server_check = ''
stop_thread = threading.Event()

# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 셀레니움
# ══════════════════════════════════════════════════════

# 드라이버 세팅
def setup_driver():
    """
    Selenium 웹 드라이버를 설정하고 반환하는 함수입니다.
    """
    chrome_options = Options()
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")

    # 사용자 에이전트 설정
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    chrome_options.add_argument(f'user-agent={user_agent}')

    # 자동화 탐지 방지 설정
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)

    # 크롬 드라이버 실행 및 자동화 방지 우회
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''
            Object.defineProperty(navigator, 'webdriver', {
              get: () => undefined
            })
        '''
    })

    # 브라우저 위치와 크기 설정
    driver.set_window_position(0, 0)  # 왼쪽 위 (0, 0) 위치로 이동
    driver.set_window_size(500, 800)  # 크기를 500x800으로 설정

    return driver
# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 엑셀 관련
# ══════════════════════════════════════════════════════

# 드래그 엑셀 파일로 이름 업로드
def on_drop(event):
    global id_list, filepath
    filepath = event.data.strip('{}')
    id_list = read_excel_file(filepath)
    update_log(id_list)
    # 리스트 상태 확인 및 버튼 활성화
    check_list_and_toggle_button(id_list)

# 윈도우 엑셀 파일로 이름 업로드
def browse_file():
    global id_list, filepath
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    if filepath:
        id_list = read_excel_file(filepath)
        update_log(id_list)
        # 리스트 상태 확인 및 버튼 활성화
        check_list_and_toggle_button(id_list)

# 엑셀에 url 리스트를 가져오는 함수
def update_log(id_list):
    log_text_widget.delete(1.0, tk.END)
    for url in id_list:
        log_text_widget.insert(tk.END, url + "\n")
    log_text_widget.insert(tk.END, f"\n총 {len(id_list)}개의 URL이 있습니다.\n")
    log_text_widget.see(tk.END)

# url들이 정상인지를 확인하여 버튼을 활성 비활성화 한다.
def check_list_and_toggle_button(id_list):
    if id_list:
        start_button.config(state=tk.NORMAL)
    else:
        start_button.config(state=tk.DISABLED)

# url에서 id를 추출
def extract_blog_id(url):
    # 마지막 '://' 뒤의 내용만 남기기
    url = url.rsplit('://', 1)[-1]

    # "PostList.naver"에서 blogId 추출
    postlist_pattern = r'[?&]blogId=([^&]+)'
    postlist_match = re.search(postlist_pattern, url)

    if postlist_match:
        return postlist_match.group(1)  # blogId 값 추가

    # 기본 URL에서 ID 추출 (www 포함)
    base_pattern = r'^(?:www\.)?(?:blog|m\.blog)\.naver\.com/([^/?&]+)'
    match = re.search(base_pattern, url)

    if match:
        return match.group(1)  # 첫 번째 그룹 (ID)

    return ''  # ID가 없을 경우 빈 문자열 반환

# 엑셀의 url리스트를 읽어오는 함수.
def read_excel_file(filepath):
    global id_list
    df = pd.read_excel(filepath, sheet_name=0)
    id_list = []
    url_list = df.iloc[:, 1].tolist()
    for url in url_list:
        url_blog_id = extract_blog_id(url)
        id_list.append(url_blog_id.strip())
    return id_list


def save_excel_file_sheet2(results):
    global filepath
    # file_name = "블로그 빅.xlsx"
    new_print("엑셀 저장 시작")

    try:
        # 열 순서를 정의
        columns_order = [
            '아이디', '블로그 URL', '게시글 URL',
            '태그1 이름', '태그1 PC 검색수', '태그1 모바일 검색수', '태그1 순위',
            '태그2 이름', '태그2 PC 검색수', '태그2 모바일 검색수', '태그2 순위',
            '태그3 이름', '태그3 PC 검색수', '태그3 모바일 검색수', '태그3 순위',
            '태그4 이름', '태그4 PC 검색수', '태그4 모바일 검색수', '태그4 순위',
            '태그5 이름', '태그5 PC 검색수', '태그5 모바일 검색수', '태그5 순위'
        ]

        # 누락된 열을 기본값으로 채우기
        for result in results:
            for column in columns_order:
                if column not in result:
                    result[column] = ''  # 기본값 설정

        # 기존 데이터 처리
        if os.path.exists(filepath):
            existing_df = pd.read_excel(filepath)
            new_df = pd.DataFrame(results, columns=columns_order)  # 순서 고정
            df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            df = pd.DataFrame(results, columns=columns_order)  # 순서 고정

        # 열 순서 명시적으로 다시 설정
        df = df[columns_order]

        # 엑셀 파일 저장 (Sheet1에 저장)
        df.to_excel(filepath, index=False, engine="openpyxl", sheet_name='Sheet2')

        # 색상 조건 적용 및 왼쪽 정렬
        apply_color_and_alignment_to_excel(filepath, df)

        new_print(f"엑셀 저장 완료: {filepath}")
    except Exception as e:
        new_print(f"엑셀 저장 실패: {e}")

# 엑셀저장
def save_excel_file_sheet1(new_data):
    global filepath
    # 기존 엑셀 파일 읽기
    df = pd.read_excel(filepath, sheet_name='Sheet1')

    # new_data 복사본 생성
    new_data_copy = new_data.copy()

    # new_data_copy의 길이가 df의 길이보다 짧을 경우, 부족한 부분을 ''으로 채움
    if len(new_data_copy) < len(df):
        new_data_copy.extend([''] * (len(df) - len(new_data_copy)))
    elif len(new_data_copy) > len(df):
        new_data_copy = new_data_copy[:len(df)]  # new_data_copy가 더 길면 df 길이에 맞게 자름

    # 새로운 데이터를 '퍼센트' 열에 추가
    df['퍼센트'] = new_data_copy  # '퍼센트'라는 이름으로 F 컬럼에 데이터 저장

    # 업데이트된 데이터 저장
    df.to_excel(filepath, index=False)


def apply_color_and_alignment_to_excel(file_name, df):
    # 엑셀 파일 로드
    workbook = load_workbook(file_name)
    sheet = workbook.active

    # 색상 정의
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

    # 폰트 정의 (검은색 배경일 때 흰색 글씨)
    white_font = Font(color="FFFFFF")
    default_font = Font(color="000000")  # 기본 검은색 글씨

    # 컬럼 이름 추출
    tag_columns = [f"태그{i} PC 검색수" for i in range(1, 6)] + [f"태그{i} 모바일 검색수" for i in range(1, 6)]
    tag_name_columns = [f"태그{i} 이름" for i in range(1, 6)]

    for row in range(2, sheet.max_row + 1):  # 데이터 시작 행부터 끝까지
        for i in range(5):  # 태그 1~5에 대해 처리
            pc_col = tag_columns[i]  # PC 검색수 컬럼
            mobile_col = tag_columns[i + 5]  # 모바일 검색수 컬럼
            name_col = tag_name_columns[i]  # 이름 컬럼

            pc_value = df.iloc[row - 2][pc_col] if pc_col in df.columns else 0
            mobile_value = df.iloc[row - 2][mobile_col] if mobile_col in df.columns else 0

            # 값이 숫자가 아니면 흰색으로 처리
            try:
                pc_value = int(pc_value)
                mobile_value = int(mobile_value)
            except ValueError:
                max_value = 0  # 숫자가 아닌 경우 기본 흰색 처리
            else:
                max_value = max(pc_value, mobile_value)

            # 검색수 조건 판단
            if max_value >= 50000:
                fill = black_fill
                font = white_font  # 글씨를 흰색으로 설정
            elif max_value >= 10000:
                fill = blue_fill
                font = default_font
            elif max_value >= 5000:
                fill = green_fill
                font = default_font
            elif max_value >= 3000:
                fill = yellow_fill
                font = default_font
            else:
                fill = white_fill
                font = default_font

            # 셀 색상 및 글꼴 적용
            col_letter = chr(65 + df.columns.get_loc(name_col))  # 컬럼 이름을 엑셀 열 문자로 변환
            cell = sheet[f"{col_letter}{row}"]
            cell.fill = fill
            cell.font = font

    # 전체 왼쪽 정렬
    apply_left_alignment_to_excel(sheet)

    # 저장
    workbook.save(file_name)
    workbook.close()


def apply_left_alignment_to_excel(sheet):
    # 왼쪽 정렬 정의
    left_alignment = Alignment(horizontal="left", vertical="center")

    # 전체 셀 왼쪽 정렬 적용
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = left_alignment


# 글수
def on_select_cont(event):
    global selected_cont_value
    selected_cont_value = int(value_cont_select.get())


# 태그
def on_select(event):
    global selected_value
    selected_value = int(value_select.get())

# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 네이버 API
# ══════════════════════════════════════════════════════

def requests_get(url, headers, payload=None):
    global global_naver_cookies
    if payload:  # payload가 존재할 때만 params를 포함
        rs = requests.get(url, headers=headers, cookies=global_naver_cookies, params=payload)
        return rs
    else:
        rs = requests.get(url, headers=headers, cookies=global_naver_cookies)
        return rs

# 네이버 로그아웃
def naver_logout():
    global global_naver_cookies, login_button, login_board, id_list, extracted_data_list
    global_naver_cookies = None
    login_button.config(text="로그인", command=naver_login)
    login_board.config(text="관리자님 로그인을 진행해주세요.", fg="red")
    messagebox.showinfo("알림", "로그인 아웃: 정상 로그아웃 되었습니다.")
    start_button.config(text="시작", bg="#d0f0c0", fg="black", state=tk.DISABLED)
    id_list = []
    extracted_data_list = []  # 모든 데이터 저장용
    stop_thread.set()  # 종료 신호 설정


def periodic_token_update():
    """
    5분마다 update_bearer_token 함수를 실행하는 스레드.
    """
    while True:
        try:
            update_bearer_token()  # Bearer Token 갱신
        except Exception as e:
            new_print(f"토큰 업데이트 중 오류 발생: {e}")
        time.sleep(300)  # 5분 대기


# 네이버 로그인
def naver_login():
    global global_naver_cookies, login_button, login_board, bearer_token, refresh_token
    try:
        driver = setup_driver()
        driver.get("https://nid.naver.com/nidlogin.login")
        time.sleep(2)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "id")))

        logged_in = False
        max_wait_time = 300
        start_time = time.time()

        while not logged_in:
            time.sleep(1)
            elapsed_time = time.time() - start_time

            if elapsed_time > max_wait_time:
                messagebox.showwarning("경고", "로그인 실패: 300초 내에 로그인하지 않았습니다.")
                break

            cookies = {cookie['name']: cookie['value'] for cookie in driver.get_cookies()}
            if 'NID_AUT' in cookies and 'NID_SES' in cookies:
                global_naver_cookies = cookies
                logged_in = True
                break

        # 로그인 후 관리 페이지로 이동
        if logged_in:
            driver.get("https://manage.searchad.naver.com/customers/3216661/tool/keyword-planner")
            time.sleep(3)

            # 로컬 스토리지에서 tokens 값 가져오기
            tokens_json = driver.execute_script(
                "return window.localStorage.getItem('tokens');"
            )

            if tokens_json:
                # JSON 파싱
                tokens = json.loads(tokens_json)

                if "3216661" in tokens:
                    account_data = tokens["3216661"]
                    bearer_token = f'Bearer {account_data.get("bearer")}'
                    refresh_token = account_data.get("refreshToken")
                    messagebox.showinfo("알림", "로그인 성공: 정상 로그인 되었습니다.")
                    login_button.config(text="로그아웃", command=naver_logout)
                    login_board.config(text="관리자님 로그인 되었습니다. 작업을 진행해주세요.", fg="blue")

                    update_thread = threading.Thread(target=periodic_token_update, daemon=True)  # Daemon 스레드로 설정
                    update_thread.start()

                else:
                    new_print("경고: 지정된 키가 없습니다.")
            else:
                new_print("경고: 값이 없습니다.")

    except Exception as e:
        messagebox.showwarning("경고", f"로그인 중 오류가 발생했습니다.{e}")
    finally:
        driver.quit()

# 내 블로그 30개 번호 가져오기
def fetch_naver_blog_my_logNos(blog_id, current_page):
    """주어진 블로그 ID와 현재 페이지를 사용하여 게시글 제목을 가져옵니다."""
    url = f"https://m.blog.naver.com/api/blogs/{blog_id}/post-list?categoryNo=0&itemCount=30&page={current_page}&userId="
    headers = {
        "authority": "m.blog.naver.com",
        "method": "GET",
        "path": "/api/blogs/roketmissile/post-list?categoryNo=0&itemCount=10&page=1&userId=",
        "scheme": "https",
        "accept": "application/json, text/plain, */*",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "priority": "u=1, i",
        "referer": "https://m.blog.naver.com/roketmissile?tab=1",
        "sec-ch-ua": '"Chromium";v="130", "Google Chrome";v="130", "Not?A_Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
    }
    try:
        response = requests_get(url, headers)
        if response.status_code == 200:
            json_data = response.json()
            if json_data.get("isSuccess"):
                items = json_data['result']['items']
                logNos = []
                for item in items:
                    obj = {
                        'logNo': item.get("logNo"),
                        'title': item.get("titleWithInspectMessage")
                    }
                    logNos.append(obj)
                return logNos[:selected_cont_value]
    except requests.RequestException as e:
        messagebox.showwarning("경고", "게시글을 가져오는중 에러가 발생했습니다.")
        return []
    except json.JSONDecodeError as e:
        messagebox.showwarning("경고", "게시글을 가져오는중 에러가 발생했습니다.")
        return []
    except Exception as e:
        messagebox.showwarning("경고", e)
        return []

# 블로그 게시글에서 해시태그들을 가져오기
def fetch_gs_tag_name(blog_id, logNo):

    url = f'https://m.blog.naver.com/{blog_id}/{logNo}?referrerCode=1'

    # HTTP GET 요청
    headers = {
        'authority': 'm.blog.naver.com',
        'method': 'GET',
        'path': f'/{blog_id}/{logNo}?referrerCode=1',
        'scheme': 'https',
        'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
        'accept-encoding': 'gzip, deflate, br, zstd',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'cache-control': 'max-age=0',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
    }

    response = requests_get(url, headers)

    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')

        # <script> 태그 찾기
        for script in soup.find_all('script'):
            if 'gsTagName' in script.text:
                # gsTagName 변수 찾기
                start = script.text.find('gsTagName = "') + len('gsTagName = "')
                end = script.text.find('";', start)
                gs_tag_value = script.text[start:end]

                # 콤마로 나누고 최대 5개 반환
                tags = gs_tag_value.split(',')
                return tags[:selected_value]  # 최대 5개 반환

    return []  # 요청 실패 시 빈 리스트 반환

# 검색 블로그 20개 번호 가져오기
def fetch_naver_blog_search_logNos(query, page):
    url = "https://s.search.naver.com/p/review/48/search.naver"

    # 페이로드를 딕셔너리 형태로 정의
    payload = {
        "ssc": "tab.blog.all",
        "api_type": 8,
        "query": f"{query}",
        "start": f"{page + 1}",
        "sm": "tab_hty.top",
        "prank": f'{page}',
        "ngn_country": "KR"
    }
    query_encoding = quote(query)

    # 헤더 설정
    headers = {
        "authority": "s.search.naver.com",
        "method": "GET",
        "path": "/p/review/48/search.naver",
        "scheme": "https",
        "accept": "application/json, text/javascript, */*; q=0.01",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "origin": "https://search.naver.com",
        "referer": f"https://search.naver.com/search.naver?sm=tab_hty.top&ssc=tab.blog.all&query={query_encoding}&oquery={query_encoding}&tqi=iyLxLlqo1awssNDx7HsssssstkG-146063",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
    }

    # GET 요청 보내기
    response = requests_get(url, headers, payload)
    if response.status_code == 200:
        # JSON 응답 파싱
        json_data = response.json()
        contents = json_data.get("contents", "")

        # HTML 파싱
        soup = BeautifulSoup(contents, 'html.parser')

        # class="detail_box" 안에 있는 title_area의 a 태그 찾기
        detail_boxes = soup.find_all(class_="detail_box")
        logNos = []  # 제목과 LogNo를 담을 리스트

        for box in detail_boxes:
            title_area = box.find(class_="title_area")
            if title_area:
                a_tag = title_area.find('a')
                if a_tag:
                    # href 속성에서 LogNo 추출
                    href = a_tag['href']
                    log_no = href.split('/')[-1]  # URL의 마지막 부분이 LogNo
                    logNos.append(log_no)

        return logNos[:30]  # LogNo 리스트
    else:
        new_print("조회된 해시태크 목록이 없습니다.")
        return []

# 네이버 검색광고 토큰 업데이트 5분에 한번씩 호출
def update_bearer_token():
    global bearer_token, refresh_token

    # URL 설정
    url = f"https://atower.searchad.naver.com/auth/local/extend"

    # 헤더 설정
    headers = {
        "authority": "atower.searchad.naver.com",
        "method": "PUT",
        "path": f"/auth/local/extend?refreshToken={refresh_token}",
        "scheme": "https",
        "accept": "*/*",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "content-length": "0",
        "origin": "https://manage.searchad.naver.com",
        "priority": "u=1, i",
        "referer": "https://manage.searchad.naver.com/customers/3216661/tool/keyword-planner",
        "sec-ch-ua": "\"Google Chrome\";v=\"131\", \"Chromium\";v=\"131\", \"Not_A Brand\";v=\"24\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\"",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-site",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    }

    params = {
        "refreshToken": refresh_token
    }

    # PUT 요청 보내기
    response = requests.put(url, headers=headers, params=params, cookies=global_naver_cookies)

    # 응답 처리
    if response.status_code == 200:
        try:
            # JSON 응답 파싱
            data = response.json()
            token = data.get("token")
            refresh_token = data.get("refreshToken")

            if token:
                bearer_token = f"Bearer {token}"  # Bearer Token 업데이트
                new_print("Token 업데이트 완료:")
            else:
                new_print("응답에 token이 없습니다.")

        except ValueError:
            new_print("응답이 JSON 형식이 아닙니다.")
    else:
        new_print(f"요청 실패: 상태 코드 {response.status_code}, 응답 내용: {response.text}")

# 검색수 가져오기
def search_keyword_cnt(keyword):

    # 기본 URL과 동적 키워드 설정
    base_url = "https://manage.searchad.naver.com/keywordstool"
    params = {
        "format": "json",
        "hintKeywords": keyword,
        "siteId": "",
        "month": "",
        "biztpId": "",
        "event": "",
        "includeHintKeywords": "0",
        "showDetail": "1",
        "keyword": "",
    }

    # 헤더 설정
    headers = {
        "accept": "application/json, text/plain, */*",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "authorization": bearer_token,
        "priority": "u=1, i",
        "referer": "https://manage.searchad.naver.com/customers/3216661/tool/keyword-planner",
        "sec-ch-ua": "\"Google Chrome\";v=\"131\", \"Chromium\";v=\"131\", \"Not_A Brand\";v=\"24\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\"",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        "x-accept-language": "ko",
    }

    # GET 요청 보내기
    response = requests.get(base_url, headers=headers, params=params, cookies=global_naver_cookies)


    # 응답 처리
    if response.status_code == 200:
        try:
            data = response.json()
            keyword_list = data.get("keywordList", [])
            for kw in keyword_list:
                if kw['relKeyword'] == keyword:
                    return kw
        except ValueError:
            new_print("응답이 JSON 형식이 아닙니다.")
    else:
        new_print(f"요청 실패: 상태 코드 {response.status_code}")


# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 윈도우 처리
# ══════════════════════════════════════════════════════

# 완료후 깜빡임 상태를 관리하는 함수
def flash_window(root):
    global flashing

    # FLASHWINFO 구조체 정의
    class FLASHWINFO(ctypes.Structure):
        _fields_ = [('cbSize', ctypes.c_uint),
                    ('hwnd', ctypes.c_void_p),
                    ('dwFlags', ctypes.c_uint),
                    ('uCount', ctypes.c_uint),
                    ('dwTimeout', ctypes.c_uint)]

    FLASHW_ALL = 3  # 모든 플래시
    hwnd = root.winfo_id()  # Tkinter 창의 윈도우 핸들 얻기
    flash_info = FLASHWINFO(ctypes.sizeof(FLASHWINFO), hwnd, FLASHW_ALL, 0, 0)

    def flash():
        while flashing:
            ctypes.windll.user32.FlashWindowEx(ctypes.byref(flash_info))
            time.sleep(0.5)  # 0.5초 간격으로 깜빡임

    threading.Thread(target=flash, daemon=True).start()  # 깜빡임을 별도의 쓰레드에서 실행

# 중지후 깜빡임 상태를 나태나는 윈도우 함수
def stop_flash_window(root):
    global flashing
    flashing = False

    # FLASHWINFO 구조체 정의
    class FLASHWINFO(ctypes.Structure):
        _fields_ = [('cbSize', ctypes.c_uint),
                    ('hwnd', ctypes.c_void_p),
                    ('dwFlags', ctypes.c_uint),
                    ('uCount', ctypes.c_uint),
                    ('dwTimeout', ctypes.c_uint)]

    hwnd = root.winfo_id()
    flash_info = FLASHWINFO(ctypes.sizeof(FLASHWINFO), hwnd, 0, 0, 0)
    ctypes.windll.user32.FlashWindowEx(ctypes.byref(flash_info))

# 시작 버튼 누르면 실행되는 함수
def toggle_start_stop():
    if global_naver_cookies:
        global id_list
        new_print("작업시작")
        if not id_list:
            messagebox.showwarning("경고", "엑셀 목록이 없습니다.")
            return
        if start_button.config('text')[-1] == '시작':
            start_button.config(text="중지", bg="red", fg="white")
            threading.Thread(target=start_processing).start()
        else:
            stop_processing()
    else:
        messagebox.showinfo("알림", "로그인을 진행해주세요.")

# 정지 버튼 정지 처리
def stop_processing():
    global stop_flag
    stop_flag = True
    start_button.config(text="시작", bg="#d0f0c0", fg="black", state=tk.DISABLED)
    new_print(f'작업중지')
    messagebox.showinfo("알림", "중지되었습니다..")

# 화면 로그
def new_print(text, level="INFO"):
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    formatted_text = f"[{timestamp}] [{level}] {text}"
    log_text_widget.insert(tk.END, f"{formatted_text}\n")
    log_text_widget.see(tk.END)

# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 실제 메인 실행 로직
# ══════════════════════════════════════════════════════

# 대기시간 처리
def time_sleep():
    global time_sleep_val
    time.sleep(random.uniform(0.7, time_sleep_val))

# 진행률 업데이트
def remaining_time_update(now_cnt, total_contents):
    global selected_value, time_sleep_val, id_list

    progress["value"] = now_cnt + 1
    progress_rate = math.floor((now_cnt + 1) / total_contents * 100 * 100) / 100  # 소수점 셋째 자리까지 버림
    progress_label.config(text=f"진행률: {progress_rate:.2f}%")  # 소수점 둘째 자리까지 표시

    # 태그조회5개 조회 (1) + 검색수 태그갯수만큼 (5) + 검색창 20개 (5)  + 1개당 30개 글을 가져오니깐  30개로 나눔
    blog_tag_cnt_time = ((1 + (selected_value * 2)) * time_sleep_val) + (time_sleep_val/selected_cont_value)

    remaining_time = int((total_contents - (now_cnt + 1)) * blog_tag_cnt_time)  # 소수점 제거
    hours = remaining_time // 3600             # 초를 시간으로 변환
    minutes = (remaining_time % 3600) // 60    # 남은 초를 분으로 변환
    seconds = remaining_time % 60              # 남은 초

    formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"
    eta_label.config(text=f"남은 시간: {formatted_time}")

# 실제 시작 처리 메인 로직
def start_processing():
    global stop_flag, extracted_data_list, id_list, extracted_data_per_list

    stop_flag = False
    # 기존 로그 화면 초기화
    log_text_widget.delete(1.0, tk.END)

    extracted_data_list = []
    extracted_data_per_list = []
    total_contents = len(id_list) * selected_cont_value
    progress["maximum"] = total_contents
    remaining_time_update(-1, total_contents)

    new_print(f'전체 ID 수 : {len(id_list)} ============================================================')
    new_print(f'전체 게시글 수 : {total_contents} ============================================================')

    # 전체 블로그 주소 id
    for index, blog_id in enumerate(id_list):
        if index != 0 and index % 5 == 0:
            new_print(f'{index} 번까지 임시 저장 ============================================================')
            save_excel_file_sheet1(extracted_data_per_list)
            save_excel_file_sheet2(extracted_data_list)

        new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 계산 시작 ============================================================')
        hash_tag_cnt = 0
        if stop_flag:
            completed_process(extracted_data_list, extracted_data_per_list)
            return
        try:
            # 내 블로그 30개 번호 가져오기
            logNos = fetch_naver_blog_my_logNos(blog_id, 1)
            new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 수 {len(logNos)} ============================================================')
            time_sleep()

            # 각 블로그안에 게시글 30개 리스트
            for idx, blog in enumerate(logNos):
                obj = {
                    '아이디': f'{blog_id}',
                    '블로그 URL': f'https://blog.naver.com/{blog_id}',
                    '게시글 URL': f'https://blog.naver.com/{blog_id}/{blog['logNo']}',
                    '태그1 이름': '',
                    '태그1 PC 검색수': '',
                    '태그1 모바일 검색수': '',
                    '태그1 순위': '',
                    '태그2 이름': '',
                    '태그2 PC 검색수': '',
                    '태그2 모바일 검색수': '',
                    '태그2 순위': '',
                    '태그3 이름': '',
                    '태그3 PC 검색수': '',
                    '태그3 모바일 검색수': '',
                    '태그3 순위': '',
                    '태그4 이름': '',
                    '태그4 PC 검색수': '',
                    '태그4 모바일 검색수': '',
                    '태그4 순위': '',
                    '태그5 이름': '',
                    '태그5 PC 검색수': '',
                    '태그5 모바일 검색수': '',
                    '태그5 순위': ''
                }
                if stop_flag:
                    completed_process(extracted_data_list, extracted_data_per_list)
                    return
                tags = fetch_gs_tag_name(blog_id, blog['logNo'])
                filter_tags = list(filter(lambda tag: tag in blog['title'], tags))
                new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 태그 목록 : {tags}')
                new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 제목 : {blog['title']}')
                new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 추출 태그 목록 : {filter_tags}')

                time_sleep()

                exit_loops = False
                for ix, tag in enumerate(filter_tags):
                    if stop_flag:
                        completed_process(extracted_data_list, extracted_data_per_list)
                        return
                    kw = search_keyword_cnt(tag)
                    time_sleep()
                    new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 태그 : [{ix + 1}/{len(filter_tags)}] - {tag}, PC 검색수 : {kw['monthlyPcQcCnt']}')
                    new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 태그 : [{ix + 1}/{len(filter_tags)}] - {tag}, 모바일 검색수 : {kw['monthlyMobileQcCnt']}')

                    obj[f'태그{ix+1} 이름'] = tag
                    obj[f'태그{ix+1} PC 검색수'] = kw['monthlyPcQcCnt']
                    obj[f'태그{ix+1} 모바일 검색수'] = kw['monthlyMobileQcCnt']

                    # 검색 블로그 20개 번호 가져오기
                    search_logNos1 = fetch_naver_blog_search_logNos(tag, 0)
                    time_sleep()
                    search_logNos2 = fetch_naver_blog_search_logNos(tag, 1)

                    # 합집합을 순서 유지하며 생성
                    union_logNos = []
                    for item in search_logNos1 + search_logNos2:  # a와 b를 이어붙인 후 순회
                        if item not in union_logNos:
                            union_logNos.append(item)

                    union_logNos = union_logNos[:30]

                    new_print(f'아이디 : {blog_id} - [{index + 1}/{len(id_list)}], 게시글 번호 : {blog['logNo']} - [{idx + 1}/{len(logNos)}], 태그 : [{ix + 1}/{len(filter_tags)}] - {tag}, 검색글 수 : {len(union_logNos)}')

                    for i, search_logNo in enumerate(union_logNos):
                        if stop_flag:
                            completed_process(extracted_data_list, extracted_data_per_list)
                            return
                        if str(search_logNo) == str(blog['logNo']):
                            new_print(f'■ 찾은 검색글 위치 : {i+1} 번째, URL : https://m.blog.naver.com/{blog_id}/{search_logNo}')
                            obj[f'태그{ix+1} 순위'] = i+1
                            if not exit_loops:
                                hash_tag_cnt += 1
                                exit_loops = True
                            break

                extracted_data_list.append(obj)
                now_cnt = (index * selected_cont_value) + idx
                remaining_time_update(now_cnt, total_contents)

            hash_tag_per = math.floor((hash_tag_cnt / selected_cont_value) * 100 * 100) / 100
            extracted_data_per_list.append(hash_tag_per)
            new_print(f'작업한 전체목록 수 : {len(extracted_data_list)}')
        except Exception as e:
            extracted_data_list.append(0)
            new_print(e, level="WARN")

        # 진행률 업데이트
        remaining_time_update(((index+1) * selected_cont_value)-1, total_contents)

    if not stop_flag:
        completed_process(extracted_data_list, extracted_data_per_list)
        return

# 완료 처리
def completed_process(extracted_data_list, extracted_data_per_list):
    global root
    save_excel_file_sheet1(extracted_data_per_list)
    save_excel_file_sheet2(extracted_data_list)
    new_print("작업 완료.", level="SUCCESS")
    start_button.config(text="시작", bg="#d0f0c0", fg="black")
    flash_window(root)
    messagebox.showinfo("알림", "작업이 완료되었습니다.")
    stop_flash_window(root)  # 메시지박스 확인 후 깜빡임 중지

# ══════════════════════════════════════════════════════
# endregion




# region
# ══════════════════════════════════════════════════════
# 로그인 페이지
# ══════════════════════════════════════════════════════
def login_window():
    global login_root, id_entry, pw_entry

    login_root = tk.Tk()
    login_root.title("로그인")

    # 창 크기 설정
    login_root.geometry("300x200")
    screen_width = login_root.winfo_screenwidth()
    screen_height = login_root.winfo_screenheight()
    window_width = 300
    window_height = 200

    # 창을 화면의 가운데로 배치
    position_top = int(screen_height / 2 - window_height / 2)
    position_left = int(screen_width / 2 - window_width / 2)
    login_root.geometry(f'{window_width}x{window_height}+{position_left}+{position_top}')

    # ID 입력
    id_label = tk.Label(login_root, text="ID:")
    id_label.pack(pady=10)
    id_entry = tk.Entry(login_root, width=20)
    id_entry.pack(pady=5)

    # PW 입력
    pw_label = tk.Label(login_root, text="PW:")
    pw_label.pack(pady=10)
    pw_entry = tk.Entry(login_root, show="*", width=20)
    pw_entry.pack(pady=5)

    # 버튼 프레임
    button_frame = tk.Frame(login_root)
    button_frame.pack(pady=10)

    # 로그인 버튼
    login_button = tk.Button(button_frame, text="로그인", command=on_login)
    login_button.grid(row=0, column=0, padx=5)

    # 비밀번호 변경 버튼
    change_pw_button = tk.Button(button_frame, text="비밀번호 변경", command=open_change_password_window)
    change_pw_button.grid(row=0, column=1, padx=5)

    login_root.mainloop()

# 로그인
def on_login():
    user_id = id_entry.get()
    user_pw = pw_entry.get()

    # 서버 세션 관리
    session = requests.Session()
    if login_to_server(user_id, user_pw, session):
        messagebox.showinfo("로그인 성공", "로그인 성공!")
        login_root.destroy()
        # 로그인 성공 후, 세션 상태를 주기적으로 체크하는 쓰레드 시작
        session_thread = threading.Thread(target=check_session_periodically, args=(session, "server"), daemon=True)
        session_thread.start()

        # 로그인 후 네이버 로그인 창을 띄운다
        main()
    else:
        messagebox.showerror("로그인 실패", "아이디 또는 비밀번호가 틀렸습니다.")

# 비밀번호 변경 창 열기
def open_change_password_window():
    login_root.destroy()  # 로그인 창 닫기
    change_password_window()  # 비밀번호 변경 창 열기

# 서버 로그인 함수 (네이버와 다른 서버 구분)
def login_to_server(username, password, session):
    global global_server_cookies
    url = f"{URL}/auth/login"
    payload = {
        "username": username,
        "password": password
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    try:
        # JSON 형식으로 서버에 POST 요청으로 로그인 시도
        response = session.post(url, json=payload, headers=headers)  # 헤더 추가

        # 요청이 성공했는지 확인
        if response.status_code == 200:
            # 세션 관리로 쿠키는 자동 처리
            global_server_cookies = session.cookies.get_dict()
            return True
        else:
            return False
    except Exception as e:
        return False

# 세션 실시간 요청
def check_session_periodically(session, server_type="server"):
    while True:
        if global_naver_cookies:
            check_session(session, server_type)  # 세션 상태를 체크
        time.sleep(60)  # 2분 대기

# 비밀번호 변경 창 열기
def open_change_password_window():
    login_root.destroy()  # 로그인 창 닫기
    change_password_window()  # 비밀번호 변경 창 열기


# 비밀번호 변경 창
def change_password_window():
    global change_pw_root, id_entry_cp, current_pw_entry, new_pw_entry

    change_pw_root = tk.Tk()
    change_pw_root.title("비밀번호 변경")

    # 창 크기 설정
    change_pw_root.geometry("300x250")
    screen_width = change_pw_root.winfo_screenwidth()
    screen_height = change_pw_root.winfo_screenheight()
    window_width = 300
    window_height = 270

    # 창을 화면의 가운데로 배치
    position_top = int(screen_height / 2 - window_height / 2)
    position_left = int(screen_width / 2 - window_width / 2)
    change_pw_root.geometry(f'{window_width}x{window_height}+{position_left}+{position_top}')

    # 창 닫기 이벤트 처리
    change_pw_root.protocol("WM_DELETE_WINDOW", return_to_login)

    # ID 입력
    id_label_cp = tk.Label(change_pw_root, text="ID:")
    id_label_cp.pack(pady=10)
    id_entry_cp = tk.Entry(change_pw_root, width=20)
    id_entry_cp.pack(pady=5)

    # 현재 비밀번호 입력
    current_pw_label = tk.Label(change_pw_root, text="현재 비밀번호:")
    current_pw_label.pack(pady=10)
    current_pw_entry = tk.Entry(change_pw_root, show="*", width=20)
    current_pw_entry.pack(pady=5)

    # 변경할 비밀번호 입력
    new_pw_label = tk.Label(change_pw_root, text="변경 비밀번호:")
    new_pw_label.pack(pady=10)
    new_pw_entry = tk.Entry(change_pw_root, show="*", width=20)
    new_pw_entry.pack(pady=5)

    # 버튼 프레임
    button_frame = tk.Frame(change_pw_root)
    button_frame.pack(pady=10)

    # 비밀번호 변경 버튼
    change_pw_button = tk.Button(button_frame, text="비밀번호 변경", command=on_change_password)
    change_pw_button.grid(row=0, column=0, padx=5)

    # 취소 버튼
    cancel_button = tk.Button(button_frame, text="취소", command=return_to_login)
    cancel_button.grid(row=0, column=1, padx=5)

    change_pw_root.mainloop()


# 취소 버튼 클릭 이벤트
def return_to_login():
    change_pw_root.destroy()
    login_window()

# 비밀번호 변경 버튼 클릭 이벤트
def on_change_password():
    user_id = id_entry_cp.get()
    current_pw = current_pw_entry.get()
    new_pw = new_pw_entry.get()

    # 여기에 비밀번호 변경 API 호출 로직 추가
    session = requests.Session()
    result = change_password(session, user_id, current_pw, new_pw)

    if result:
        messagebox.showinfo("비밀번호 변경", "비밀번호가 변경되었습니다.")
        change_pw_root.destroy()
        login_window()
    else:
        messagebox.showerror("비밀번호 변경", "비밀번호 변경에 실패하였습니다.")

# 비밀변경
def change_password(session, username, current_password, new_password):
    url = f"{URL}/auth/change-password"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }
    payload = {
        "id": username,
        "currentPassword": current_password,
        "newPassword": new_password
    }

    try:
        # PUT 요청을 사용하여 비밀번호 변경
        response = session.put(url, params=payload, headers=headers)

        if response.status_code == 200:
            return True
        else:
            return False
    except Exception as e:
        return False

# 세션체크
def check_session(session, server_type="server"):
    global login_server_check
    cookies = global_server_cookies if server_type == "server" else global_naver_cookies
    url = f"{URL}/session/check-me"
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    # /check-me 엔드포인트를 호출하여 세션 상태 확인
    response = session.get(url, headers=headers, cookies=cookies)

    if response.status_code == 200:
        login_server_check = response.text

# ══════════════════════════════════════════════════════
# endregion



# region
# ══════════════════════════════════════════════════════
# 메인 초기화
# ══════════════════════════════════════════════════════

# 초기화
def main():
    global log_text_widget, start_button, progress, progress_label, eta_label, root, login_button, login_board, value_select, value_cont_select


    root = TkinterDnD.Tk()
    root.title("블로그 빅 프로그램")
    root.geometry("600x780")

    font_large = font.Font(size=10)

    # 시작 버튼
    login_button = tk.Button(root, text="로그인", command=naver_login, font=font_large, bg="#d0f0c0", fg="black", width=25)
    login_button.pack(pady=10)

    login_board = tk.Label(root, text="관리자님 로그인을 진행해주세요.\n(로그인 완료 되면 잠시 기다려주세요...)", width=40, height=5, font=font_large, bg="white", fg="red",
                           borderwidth=1, relief="solid", highlightbackground="black", highlightthickness=1)
    login_board.pack(pady=10)

    btn_browse = tk.Button(root, text="엑셀 파일 선택", command=browse_file, font=font_large, width=20)
    btn_browse.pack(pady=10)

    lbl_or = tk.Label(root, text="또는", font=font_large)
    lbl_or.pack(pady=5)

    lbl_drop = tk.Label(root, text="여기에 파일을 드래그 앤 드롭하세요", relief="solid", width=40, height=5, font=font_large, bg="white")
    lbl_drop.pack(pady=10)

    lbl_drop.drop_target_register(DND_FILES)
    lbl_drop.dnd_bind('<<Drop>>', on_drop)

    # 시작 버튼
    start_button = tk.Button(root, text="시작", command=toggle_start_stop, font=font_large, bg="#d0f0c0", fg="black", width=25, state=tk.DISABLED)
    start_button.pack(pady=10)


    # 글수 레이블 추가
    cont_count_label = tk.Label(root, text="# 글 수", font=font_large)
    cont_count_label.pack(pady=5)

    # 글수 셀렉트 박스 추가
    value_cont_select = ttk.Combobox(root, values=list(range(1, 31)), font=font_large, state="readonly")
    value_cont_select.set(selected_cont_value)  # 초기값 설정
    value_cont_select.pack(pady=10)
    value_cont_select.bind("<<ComboboxSelected>>", on_select_cont)  # 함수 연결


    # 태그 레이블 추가
    tag_count_label = tk.Label(root, text="# 태그 수", font=font_large)
    tag_count_label.pack(pady=5)

    # 태그  셀렉트 박스 추가
    value_select = ttk.Combobox(root, values=[1, 2, 3, 4, 5], font=font_large, state="readonly")
    value_select.set(selected_value)  # 초기값 설정
    value_select.pack(pady=10)
    value_select.bind("<<ComboboxSelected>>", on_select)  # 함수 연결


    log_label = tk.Label(root, text="로그 화면", font=font_large)
    log_label.pack(fill=tk.X, padx=10)

    log_frame = tk.Frame(root)
    log_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

    x_scrollbar = tk.Scrollbar(log_frame, orient=tk.HORIZONTAL)
    x_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

    y_scrollbar = tk.Scrollbar(log_frame, orient=tk.VERTICAL)
    y_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    log_text_widget = tk.Text(log_frame, wrap=tk.NONE, height=10, font=font_large, xscrollcommand=x_scrollbar.set, yscrollcommand=y_scrollbar.set)
    log_text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    x_scrollbar.config(command=log_text_widget.xview)
    y_scrollbar.config(command=log_text_widget.yview)

    # 진행률
    progress_frame = tk.Frame(root)
    progress_frame.pack(fill=tk.X, padx=10, pady=10)

    progress_label = tk.Label(progress_frame, text="진행률: 0%", font=font_large)
    eta_label = tk.Label(progress_frame, text="남은 시간: 00:00:00", font=font_large)

    progress_label.pack(side=tk.TOP, padx=5)
    eta_label.pack(side=tk.TOP, padx=5)

    style = ttk.Style()
    style.configure("TProgressbar", thickness=30, troughcolor='white', background='green')
    progress = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate", style="TProgressbar")
    progress.pack(fill=tk.X, padx=10, pady=10, expand=True)

    root.mainloop()

# 최초 시작 메인
if __name__ == "__main__":
    login_window()
    # main()

# ══════════════════════════════════════════════════════
# endregion
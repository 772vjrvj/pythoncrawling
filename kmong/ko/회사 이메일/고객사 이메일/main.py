import openpyxl
import requests
from bs4 import BeautifulSoup
import re

# 이메일 주소를 찾기 위한 정규식 패턴
email_pattern = re.compile(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}')

# email.xlsx 파일을 열어 Website 컬럼을 읽어옴
def load_websites_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    websites = []

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1, values_only=True):
        websites.append(row[0])

    return websites, workbook, sheet

# 웹사이트에서 이메일 주소를 추출하는 함수
def get_email_from_website(url):
    try:
        response = requests.get(url, timeout=10)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'lxml')
            text = soup.get_text()
            emails = email_pattern.findall(text)
            return emails[0] if emails else None
        else:
            print(f"Failed to access {url}: Status Code {response.status_code}")
            return None
    except requests.RequestException as e:
        print(f"Error accessing {url}: {e}")
        return None

# 엑셀 파일에 이메일 주소 업데이트
def update_emails_in_excel(workbook, sheet, emails):
    for row, email in enumerate(emails, start=2):
        sheet.cell(row=row, column=2).value = email  # 이메일 컬럼 업데이트
    workbook.save("email_updated.xlsx")

# 메인 작업 함수
def main():
    file_path = "email.xlsx"  # 엑셀 파일 경로
    websites, workbook, sheet = load_websites_from_excel(file_path)

    emails = []
    for website in websites:
        print(f"Processing: {website}")
        email = get_email_from_website(website)
        emails.append(email)

    update_emails_in_excel(workbook, sheet, emails)
    print("작업이 완료되었습니다.")

if __name__ == "__main__":
    main()

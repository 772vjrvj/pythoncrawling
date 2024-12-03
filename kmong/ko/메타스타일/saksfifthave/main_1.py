import requests
import os
import pandas as pd
from PIL import Image as PILImage
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
import re
import time

# 엑셀 파일 저장을 위한 전역 리스트
product_data = []

# 디렉토리가 없으면 생성
def create_directory(dir_name):
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

# HTML 태그 제거 함수
def remove_html_tags(text):
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text).replace('\n', '\r\n')  # 줄바꿈 유지

# 이미지 저장 함수
def save_image_from_url(url, filename):
    try:
        response = requests.get(url, stream=True)
        if response.status_code == 200:
            img = PILImage.open(BytesIO(response.content))
            img.save(filename)
            print(f"Image saved: {filename}")
            return True  # 이미지 저장 성공 시 True 반환
        else:
            print(f"Failed to fetch image: {url} (Status code: {response.status_code})")
            return False  # 이미지 저장 실패 시 False 반환
    except Exception as e:
        print(f"Error fetching image from {url}: {e}")
        return False  # 이미지 저장 실패 시 False 반환

# 상품 상세 정보를 가져오는 함수
def fetch_product_details(product_id):
    url = f"https://www.saksfifthave.kr/api/product/s/0{product_id}?lang=en&siteTag=SA_KR"
    print(f"url : {url}")

    headers = {
        "authority": "www.saksfifthave.kr",
        "method": "GET",
        "scheme": "https",
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "cache-control": "max-age=0",
        "priority": "u=0, i",
        "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "sec-fetch-dest": "document",
        "sec-fetch-mode": "navigate",
        "sec-fetch-site": "none",
        "sec-fetch-user": "?1",
        "upgrade-insecure-requests": "1",
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            print("성공")
            return response.json(), url
        else:
            print(f"Error fetching product details for ID {product_id} (Status code: {response.status_code})")
            return None
    except Exception as e:
        print(f"An error occurred while fetching product details: {e}")
        return None

# 페이지별 상품 리스트 가져오는 함수
def fetch_products(page, per_page):
    url = "https://1zeqokjjx6-dsn.algolia.net/1/indexes/prd-product-SA-KR/query?x-algolia-agent=Algolia%20for%20JavaScript%20(4.24.0)%3B%20Browser"

    headers = {
        "accept": "*/*",
        "accept-encoding": "gzip, deflate, br, zstd",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "x-algolia-api-key": "1ebbbdc18b025ec3b2d4e296a82f97d9",
        "x-algolia-application-id": "1ZEQOKJJX6"
    }

    payload = {
        "query": "",
        "analyticsTags": ["productlistingpage", "womenclothing-promo"],
        "attributesToRetrieve": "*",
        "clickAnalytics": True,
        "facetFilters": [["categories.tag:saks/womens-apparel"]],
        "facetingAfterDistinct": False,
        "facets": [
            "brand.tag", "categories.tag", "collections.tag",
            "color.tag", "size.tag", "productId", "price.KRW-KR.sale"
        ],
        "filters": "availabilityFlag < 3",
        "length": per_page,
        "numericFilters": ["price.KRW-KR.sale > 0"],
        "offset": (page - 1) * per_page,
        "optionalFilters": [],
        "analytics": True,
        "userToken": "8c595326-01c3-400f-87bc-dc1e0b065a04"
    }

    try:
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Error fetching product list (Status code: {response.status_code})")
            return None
    except Exception as e:
        print(f"An error occurred while fetching products: {e}")
        return None

# 페이지별 데이터를 엑셀에 추가 저장하는 함수
def save_to_excel_append(data, filename="products.xlsx"):
    df = pd.DataFrame(data)

    # 엑셀 파일이 이미 존재하면 추가로 데이터를 기록
    if os.path.exists(filename):
        existing_df = pd.read_excel(filename)
        df = pd.concat([existing_df, df], ignore_index=True)

    df.to_excel(filename, index=False)

    # 이미지 삽입을 위해 openpyxl 로드
    wb = load_workbook(filename)
    ws = wb.active

    row_num = ws.max_row - len(data) + 1  # 현재 행부터 이미지 삽입 시작
    for index, row in enumerate(data):
        real_second_image_path = row['second_image']
        real_last_image_path = row['last_image']

        # second_image 삽입
        if os.path.exists(real_second_image_path):
            img = ExcelImage(real_second_image_path)
            img.width, img.height = 100, 100  # 이미지 크기 조정
            ws.add_image(img, f"H{row_num + index}")  # F열에 이미지 삽입

        # last_image 삽입
        if os.path.exists(real_last_image_path):
            img = ExcelImage(real_last_image_path)
            img.width, img.height = 100, 100  # 이미지 크기 조정
            ws.add_image(img, f"I{row_num + index}")  # G열에 이미지 삽입

    wb.save(filename)
    print(f"Data and images saved to {filename}")

# 모든 페이지 처리 함수
def fetch_all_pages(total_pages, per_page=120):
    global product_data
    for page in range(1, total_pages + 1):
        print(f"Fetching page {page}...")
        result = fetch_products(page, per_page)

        if result and 'hits' in result:
            hits = result['hits']

            for hit in hits:
                product_id = str(int(hit.get('productId', '0')))  # '0444' => '444'
                print(f"Fetching details for product {product_id}...")
                time.sleep(2)
                product_details, url = fetch_product_details(product_id)

                if product_details:
                    name = product_details.get('name', '')
                    description = remove_html_tags(product_details.get('description', ''))
                    brand_name = "See all " + product_details.get('brand', {}).get('name', '')

                    # 이미지 처리
                    options = product_details.get('options', [])
                    if options and len(options) > 0:
                        media = options[0].get('media', {}).get('standard', [])
                        second_image = media[1] if len(media) > 1 else ''
                        last_image = media[-1] if len(media) > 1 else ''

                        # 이미지 저장
                        product_dir = f"{product_id}"
                        create_directory(product_dir)
                        second_image_path = ""
                        last_image_path = ""

                        if second_image:
                            second_image_path = os.path.join(product_dir, f"{product_id}_second_image.jpg")
                            second_image_success = save_image_from_url(second_image, second_image_path)
                        else:
                            second_image_success = False

                        if last_image:
                            last_image_path = os.path.join(product_dir, f"{product_id}_last_image.jpg")
                            last_image_success = save_image_from_url(last_image, last_image_path)
                        else:
                            last_image_success = False

                        # 엑셀에 삽입할 이미지가 존재하는지 확인
                        if second_image_success or last_image_success:
                            obj = {
                                "productId": str(product_id),  # product_id를 문자열로 저장
                                "url": url,
                                "name": name,
                                "description": description,
                                "tag": brand_name,
                                "second_image_url": second_image,  # 실제 이미지 파일 경로
                                "last_image_url": last_image,  # 실제 이미지 파일 경로
                                "second_image": second_image_path if second_image_success else '',  # 이미지 경로가 없으면 빈 문자열
                                "last_image": last_image_path if last_image_success else '',  # 이미지 경로가 없으면 빈 문자열
                            }

                            print(f"obj : {obj}")

                            # 엑셀 데이터 저장
                            product_data.append(obj)

            # 각 페이지 끝날 때마다 엑셀에 저장
            save_to_excel_append(product_data)
            product_data = []  # 데이터를 비워서 다음 페이지 준비

        else:
            print(f"Failed to fetch page {page}")

        time.sleep(2)  # 요청 사이에 2초 대기

if __name__ == "__main__":
    # 1페이지부터 총 2페이지까지 데이터를 가져오는 예시
    fetch_all_pages(total_pages=1, per_page=50)
